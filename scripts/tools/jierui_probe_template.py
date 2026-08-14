"""只读解析捷锐凭证导入模板 XLSX，输出结构事实（P0-B 探测工具）。

用法：
    python scripts/tools/jierui_probe_template.py <模板路径> [--json <输出facts草稿路径>]

只读打开工作簿，打印 sheet 名、合并单元格、表头行、示例行和数据校验，
并可生成 voucher-import-template.facts.json 草稿（confirmed 字段留空，等人工确认）。
"""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def probe(path: Path) -> dict:
    wb = load_workbook(path, read_only=False, data_only=False)
    facts: dict = {
        "source": "捷锐查看凭证页【导入/导出→导入→下载凭证导入模板】",
        "template_file": path.name,
        "sheets": [],
    }
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12)):
            rows.append([
                None if c.value is None else str(c.value) for c in row
            ])
        facts["sheets"].append(
            {
                "sheet_name": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "merged_cells": [str(r) for r in ws.merged_cells.ranges],
                "rows_head": rows,
                "data_validations": [
                    {
                        "ranges": str(dv.sqref),
                        "type": dv.type,
                        "formula1": dv.formula1,
                    }
                    for dv in ws.data_validations.dataValidation
                ],
            }
        )
    return facts


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("template", type=Path)
    parser.add_argument("--json", type=Path, default=None)
    args = parser.parse_args()

    facts = probe(args.template)
    text = json.dumps(facts, ensure_ascii=False, indent=2)
    print(text)
    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(text + "\n", encoding="utf-8")
        print(f"\n[written] {args.json}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
