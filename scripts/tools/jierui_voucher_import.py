#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any
from urllib import request

from openpyxl import load_workbook

from invoice_hub.bookkeeping.batches import immutable_batch_payload
from invoice_hub.bookkeeping.paths import company_bookkeeping_paths
from invoice_hub.bookkeeping.repository import file_sha256, strict_read_json_object
from invoice_hub.bookkeeping.status import load_voucher_status
from invoice_hub.domain.models import ImportBatch

DEFAULT_API_BASE = "http://127.0.0.1:8766"
MONEY_QUANT = Decimal("0.01")


class DryRunError(RuntimeError):
    def __init__(self, message: str, summary: dict[str, Any]) -> None:
        super().__init__(message)
        self.summary = summary


@dataclass(frozen=True)
class ImportGroup:
    voucher_type: str
    voucher_no: str
    voucher_date: str
    row_numbers: list[int]
    rows: list[dict[str, str]]
    debit_total: Decimal
    credit_total: Decimal
    signature: dict[str, Any]

    def as_summary(self) -> dict[str, Any]:
        return {
            "voucher_type": self.voucher_type,
            "voucher_no": self.voucher_no,
            "voucher_date": self.voucher_date,
            "row_numbers": self.row_numbers,
            "line_count": len(self.rows),
            "debit_total": money_text(self.debit_total),
            "credit_total": money_text(self.credit_total),
        }


@dataclass(frozen=True)
class ExpectedVoucher:
    voucher_key: str
    voucher_no: str
    export_file: str
    signature: dict[str, Any]
    debit_total: Decimal
    credit_total: Decimal

    def as_summary(self) -> dict[str, Any]:
        return {
            "voucher_key": self.voucher_key,
            "voucher_no": self.voucher_no,
            "export_file": self.export_file,
            "debit_total": money_text(self.debit_total),
            "credit_total": money_text(self.credit_total),
        }


def money(value: object) -> Decimal:
    text = str(value if value is not None else "").strip()
    if not text:
        return Decimal("0.00")
    try:
        return Decimal(text).quantize(MONEY_QUANT, rounding=ROUND_HALF_UP)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"金额不是合法数字: {value}") from exc


def money_text(value: Decimal | object) -> str:
    number = value if isinstance(value, Decimal) else money(value)
    return format(number.quantize(MONEY_QUANT, rounding=ROUND_HALF_UP), "f")


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def normalize_voucher_no(value: object) -> str:
    text = cell_text(value)
    return text.zfill(3) if text.isdigit() and len(text) < 3 else text


def same_path(left: str | Path, right: str | Path) -> bool:
    try:
        return Path(left).expanduser().resolve() == Path(right).expanduser().resolve()
    except OSError:
        return str(left) == str(right)


def parse_import_workbook(import_file: Path) -> list[ImportGroup]:
    if not import_file.exists() or not import_file.is_file():
        raise FileNotFoundError(f"导入文件不存在: {import_file}")
    workbook = load_workbook(import_file, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"导入文件为空: {import_file}")
    headers = [cell_text(value) for value in rows[0]]
    required = {"凭证类别", "凭证号", "凭证日期", "摘要", "科目编码", "借方金额", "贷方金额"}
    missing = sorted(required.difference(headers))
    if missing:
        raise ValueError(f"导入文件缺少必需列: {', '.join(missing)}")

    groups: list[ImportGroup] = []
    current_key: tuple[str, str, str] | None = None
    current_rows: list[dict[str, str]] = []
    current_row_numbers: list[int] = []
    for row_number, values in enumerate(rows[1:], start=2):
        record = {header: cell_text(values[index]) if index < len(values) else "" for index, header in enumerate(headers)}
        if not any(record.values()):
            continue
        key = (record.get("凭证类别", ""), normalize_voucher_no(record.get("凭证号", "")), record.get("凭证日期", ""))
        record["凭证号"] = key[1]
        if current_rows and key != current_key:
            groups.append(build_group(current_key, current_rows, current_row_numbers))
            current_rows = []
            current_row_numbers = []
        current_key = key
        current_rows.append(record)
        current_row_numbers.append(row_number)
    if current_rows:
        groups.append(build_group(current_key, current_rows, current_row_numbers))
    return groups


def build_group(key: tuple[str, str, str] | None, rows: list[dict[str, str]], row_numbers: list[int]) -> ImportGroup:
    voucher_type, voucher_no, voucher_date = key or ("", "", "")
    debit_total = sum((money(row.get("借方金额")) for row in rows), Decimal("0.00"))
    credit_total = sum((money(row.get("贷方金额")) for row in rows), Decimal("0.00"))
    return ImportGroup(
        voucher_type=voucher_type,
        voucher_no=voucher_no,
        voucher_date=voucher_date,
        row_numbers=list(row_numbers),
        rows=list(rows),
        debit_total=debit_total,
        credit_total=credit_total,
        signature=signature_from_import_rows(voucher_date, rows),
    )


def signature_from_import_rows(voucher_date: str, rows: list[dict[str, str]]) -> dict[str, Any]:
    lines = []
    for row in rows:
        debit = row.get("借方金额", "")
        credit = row.get("贷方金额", "")
        direction = "debit" if debit else "credit"
        lines.append(
            {
                "summary": row.get("摘要", ""),
                "account_code": row.get("科目编码", ""),
                "direction": direction,
                "amount": money_text(debit or credit),
            }
        )
    return {"voucher_date": voucher_date, "lines": lines}


def signature_from_snapshot(snapshot: dict[str, Any]) -> dict[str, Any]:
    lines = []
    for line in snapshot.get("lines") or []:
        lines.append(
            {
                "summary": str(line.get("summary") or ""),
                "account_code": str(line.get("account_code") or ""),
                "direction": str(line.get("direction") or ""),
                "amount": money_text(line.get("amount")),
            }
        )
    return {"voucher_date": str(snapshot.get("voucher_date") or ""), "lines": lines}


def totals_from_signature(signature: dict[str, Any]) -> tuple[Decimal, Decimal]:
    debit = Decimal("0.00")
    credit = Decimal("0.00")
    for line in signature.get("lines") or []:
        amount = money(line.get("amount"))
        if line.get("direction") == "debit":
            debit += amount
        elif line.get("direction") == "credit":
            credit += amount
    return debit, credit


def load_batch_manifest(path: str | Path) -> ImportBatch:
    manifest_path = Path(path).expanduser().resolve()
    if not manifest_path.is_file():
        raise FileNotFoundError(f"批次 manifest 不存在: {manifest_path}")
    return ImportBatch.model_validate(strict_read_json_object(manifest_path))


def resolve_batch_manifest(paths, batch_manifest: str | Path | None, import_file: str | Path | None) -> Path:
    if batch_manifest:
        return Path(batch_manifest).expanduser().resolve()
    if import_file:
        inferred = Path(import_file).expanduser().resolve().parent / "manifest.json"
        if inferred.is_file():
            return inferred
    raise ValueError("W8 runner 必须显式提供 --batch-manifest；不再猜测最新 XLSX")


def load_expected_vouchers(status_path: Path, batch: ImportBatch) -> list[ExpectedVoucher]:
    store = load_voucher_status(status_path)
    authoritative = store.batches.get(batch.batch_id)
    if authoritative is None:
        raise ValueError(f"状态文件未登记批次: {batch.batch_id}")
    if immutable_batch_payload(authoritative) != immutable_batch_payload(batch):
        raise ValueError("状态文件批次与 immutable manifest 不一致")
    expected: list[ExpectedVoucher] = []
    for batch_item in batch.items:
        key = batch_item.posting_key
        item = store.items.get(key)
        if item is None or item.batch_id != batch.batch_id:
            raise ValueError(f"状态文件缺少批次凭证: {key}")
        if item.status not in {"exported", "importing", "import_unknown", "import_failed_confirmed", "imported"}:
            raise ValueError(f"批次凭证状态不可用于 dry-run: {key}={item.status}")
        snapshot = dict(item.snapshot or {})
        if str(snapshot.get("proposal_revision_hash") or "") != batch_item.proposal_revision_hash:
            raise ValueError(f"批次凭证 revision 已漂移: {key}")
        signature = signature_from_snapshot(snapshot)
        debit_total, credit_total = totals_from_signature(signature)
        expected.append(
            ExpectedVoucher(
                voucher_key=key,
                voucher_no=normalize_voucher_no(batch_item.planned_voucher_no),
                export_file=batch.file_path,
                signature=signature,
                debit_total=debit_total,
                credit_total=credit_total,
            )
        )
    return expected


def duplicate_voucher_numbers(groups: list[ImportGroup]) -> list[dict[str, Any]]:
    seen: dict[str, ImportGroup] = {}
    duplicates = []
    for group in groups:
        if group.voucher_no in seen:
            duplicates.append(
                {
                    "voucher_no": group.voucher_no,
                    "first_rows": seen[group.voucher_no].row_numbers,
                    "duplicate_rows": group.row_numbers,
                }
            )
        else:
            seen[group.voucher_no] = group
    return duplicates


def compare_import_to_status(groups: list[ImportGroup], expected: list[ExpectedVoucher]) -> dict[str, list[dict[str, Any]]]:
    errors: dict[str, list[dict[str, Any]]] = {
        "missing": [],
        "duplicates": duplicate_voucher_numbers(groups),
        "extra": [],
        "unbalanced": [],
    }
    for group in groups:
        if group.debit_total != group.credit_total:
            errors["unbalanced"].append(group.as_summary())

    groups_by_no: dict[str, list[ImportGroup]] = {}
    for group in groups:
        groups_by_no.setdefault(group.voucher_no, []).append(group)
    expected_by_no = {item.voucher_no: item for item in expected}
    matched_group_ids: set[int] = set()

    for item in expected:
        candidates = groups_by_no.get(item.voucher_no) or []
        match = next((group for group in candidates if group.signature == item.signature), None)
        if match:
            matched_group_ids.add(id(match))
            continue
        detail = item.as_summary()
        detail["reason"] = "file_group_missing" if not candidates else "file_group_mismatch"
        if candidates:
            detail["file_candidates"] = [group.as_summary() for group in candidates]
        errors["missing"].append(detail)

    for group in groups:
        expected_item = expected_by_no.get(group.voucher_no)
        if not expected_item or id(group) not in matched_group_ids:
            detail = group.as_summary()
            detail["reason"] = "status_missing" if not expected_item else "content_mismatch"
            errors["extra"].append(detail)

    return {key: value for key, value in errors.items() if value}


def append_audit_log(company_dir: Path, summary: dict[str, Any]) -> Path:
    paths = company_bookkeeping_paths(company_dir)
    paths.log_dir.mkdir(parents=True, exist_ok=True)
    log_path = paths.log_dir / f"rpa_{datetime.now().strftime('%Y%m%d')}.jsonl"
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    actions = list(summary.get("actions") or [])
    final_action = "dry_run_passed" if summary.get("ok") else "dry_run_failed"
    actions.append({"action": final_action, "errors": summary.get("errors", {})})
    with log_path.open("a", encoding="utf-8") as handle:
        for action in actions:
            handle.write(json.dumps({"ts": now, "mode": "dry-run", **action}, ensure_ascii=False, default=str) + "\n")
    return log_path


def dry_run(
    company_dir: str | Path,
    import_file: str | Path | None = None,
    api_base: str = DEFAULT_API_BASE,
    *,
    batch_manifest: str | Path | None = None,
) -> dict[str, Any]:
    company = Path(company_dir).expanduser().resolve()
    paths = company_bookkeeping_paths(company)
    manifest_path = resolve_batch_manifest(paths, batch_manifest, import_file)
    batch = load_batch_manifest(manifest_path)
    expected_batch_dir = (paths.batch_dir / batch.batch_id).resolve()
    expected_manifest = expected_batch_dir / "manifest.json"
    if manifest_path != expected_manifest or Path(batch.manifest_path).expanduser().resolve() != expected_manifest:
        raise ValueError("batch manifest 路径未绑定当前公司和 batch_id")
    selected_import = Path(batch.file_path).expanduser().resolve()
    if selected_import != expected_batch_dir / "凭证导入.xlsx":
        raise ValueError("批次 XLSX 路径未绑定 manifest 所在目录")
    if import_file and not same_path(selected_import, import_file):
        raise ValueError("--import-file 与 batch manifest 绑定文件不一致")
    if not selected_import.is_file():
        raise FileNotFoundError(f"批次导入文件不存在: {selected_import}")
    actual_file_sha256 = file_sha256(selected_import)
    if actual_file_sha256 != batch.file_sha256:
        summary = {
            "ok": False,
            "dry_run": True,
            "apply": False,
            "batch_id": batch.batch_id,
            "batch_manifest": str(manifest_path),
            "company_dir": str(company),
            "import_file": str(selected_import),
            "file_sha256": actual_file_sha256,
            "expected_file_sha256": batch.file_sha256,
            "errors": {"file_hash": [{"reason": "batch_file_changed"}]},
            "actions": [],
        }
        summary["audit_log_path"] = str(append_audit_log(company, summary))
        raise DryRunError("批次 XLSX SHA256 已变化，原授权失效", summary)
    groups = parse_import_workbook(selected_import)
    expected = load_expected_vouchers(paths.voucher_status_json, batch)
    errors = compare_import_to_status(groups, expected)
    debit_total = sum((group.debit_total for group in groups), Decimal("0.00"))
    credit_total = sum((group.credit_total for group in groups), Decimal("0.00"))
    actions = [
        {"action": "read_import_file", "path": str(selected_import), "voucher_count": len(groups)},
        {"action": "read_voucher_status", "path": str(paths.voucher_status_json), "exported_count": len(expected)},
        {
            "action": "plan_jierui_import",
            "target_hint": f"请在捷锐中核对目标账套精确为「{batch.ledger_name}」且期间为 {batch.period}",
            "voucher_count": len(groups),
            "debit_total": money_text(debit_total),
            "credit_total": money_text(credit_total),
        },
    ]
    summary: dict[str, Any] = {
        "ok": not errors,
        "dry_run": True,
        "apply": False,
        "batch_id": batch.batch_id,
        "batch_manifest": str(manifest_path),
        "file_sha256": actual_file_sha256,
        "ledger_name": batch.ledger_name,
        "period": batch.period,
        "facts_version": batch.template_facts_version,
        "company_dir": str(company),
        "import_file": str(selected_import),
        "api_base": api_base,
        "status_path": str(paths.voucher_status_json),
        "voucher_count": len(groups),
        "expected_exported_count": len(expected),
        "debit_total": money_text(debit_total),
        "credit_total": money_text(credit_total),
        "target_hint": f"请在捷锐中核对目标账套精确为「{batch.ledger_name}」且期间为 {batch.period}",
        "actions": actions,
        "errors": errors,
    }
    log_path = append_audit_log(company, summary)
    summary["audit_log_path"] = str(log_path)
    if errors:
        raise DryRunError("捷锐导入 dry-run 对账失败", summary)
    return summary


def patch_import_result(
    api_base: str,
    voucher_key: str,
    result: str,
    *,
    voucher_no: str = "",
    detail: str = "",
    timeout: int = 10,
) -> dict[str, Any]:
    raise RuntimeError("BATCH_FINALIZE_REQUIRED: W8 已停用逐张导入结果回写")


def _post_json(url: str, payload: dict[str, Any], timeout: int = 10) -> dict[str, Any]:
    encoded = json.dumps(payload, ensure_ascii=False).encode("utf-8")
    req = request.Request(url, data=encoded, method="POST", headers={"Content-Type": "application/json"})
    with request.urlopen(req, timeout=timeout) as response:
        return json.loads(response.read().decode("utf-8"))


def record_dry_run(api_base: str, batch_id: str, summary: dict[str, Any], timeout: int = 10) -> dict[str, Any]:
    payload = {
        "ok": summary.get("ok") is True,
        "file_sha256": summary.get("file_sha256", ""),
        "voucher_count": summary.get("voucher_count", 0),
        "debit_total": summary.get("debit_total", "0.00"),
        "credit_total": summary.get("credit_total", "0.00"),
        "errors": summary.get("errors", {}),
    }
    return _post_json(f"{api_base.rstrip('/')}/api/v1/bookkeeping/import-batches/{batch_id}/dry-run", payload, timeout)


def finalize_batch(api_base: str, batch_id: str, payload: dict[str, Any], timeout: int = 10) -> dict[str, Any]:
    return _post_json(f"{api_base.rstrip('/')}/api/v1/bookkeeping/import-batches/{batch_id}/finalize", payload, timeout)


def apply_import(
    company_dir: str | Path,
    import_file: str | Path | None = None,
    api_base: str = DEFAULT_API_BASE,
    *,
    batch_manifest: str | Path | None = None,
) -> dict[str, Any]:
    summary = dry_run(company_dir, import_file=import_file, api_base=api_base, batch_manifest=batch_manifest)
    summary["apply"] = True
    raise NotImplementedError("W8 仅发布 batch-bound dry-run；真实捷锐 apply 与 reconcile-only driver 留待 W10。")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="捷锐凭证批次 runner（W8 仅开放 dry-run）")
    parser.add_argument("--company-dir", required=True, help="公司资料夹路径")
    parser.add_argument("--batch-manifest", required=True, help="凭证/批次/<batch_id>/manifest.json")
    parser.add_argument("--import-file", default="", help="可选复核路径；必须与 manifest 绑定文件一致")
    parser.add_argument("--api-base", default=DEFAULT_API_BASE, help="InvoiceHub API 地址")
    parser.add_argument("--mode", choices=("dry-run", "apply", "reconcile-only"), default="dry-run")
    parser.add_argument("--record-api", action="store_true", help="dry-run 通过后回写批次状态")
    parser.add_argument("--apply", action="store_true", help=argparse.SUPPRESS)
    return parser


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    import_file = args.import_file or None
    try:
        mode = "apply" if args.apply else args.mode
        if mode == "apply":
            summary = apply_import(
                args.company_dir,
                import_file=import_file,
                api_base=args.api_base,
                batch_manifest=args.batch_manifest,
            )
        elif mode == "reconcile-only":
            raise NotImplementedError("W8 已保留 reconcile-only 协议，但真实只读 driver 留待 W10。")
        else:
            summary = dry_run(
                args.company_dir,
                import_file=import_file,
                api_base=args.api_base,
                batch_manifest=args.batch_manifest,
            )
            if args.record_api:
                summary["api_result"] = record_dry_run(args.api_base, summary["batch_id"], summary)
        print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
        return 0
    except DryRunError as exc:
        print(json.dumps(exc.summary, ensure_ascii=False, indent=2, default=str))
        print(str(exc), file=sys.stderr)
        return 1
    except NotImplementedError as exc:
        print(str(exc), file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
