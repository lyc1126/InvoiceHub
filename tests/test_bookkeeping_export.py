import csv
import hashlib
import json
from pathlib import Path

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from invoice_hub.api.app import create_app
from invoice_hub.bookkeeping.catalogs import canonical_ledger_identity
from invoice_hub.bookkeeping.import_file import load_jierui_import_facts, write_jierui_import_xlsx
from invoice_hub.bookkeeping.mapping import MappingStoreBinding, write_mapping
from invoice_hub.bookkeeping.paths import company_bookkeeping_paths
from invoice_hub.bookkeeping.repository import canonical_sha256, file_sha256
from invoice_hub.bookkeeping.status import load_voucher_status, write_voucher_status
from invoice_hub.domain.models import AccountMappingRule, VoucherDraft, VoucherLine, VoucherStatusItem
from invoice_hub.projections.cost_analysis import DETAIL_HEADERS
from invoice_hub.services.app_state import AppState

FACTS_COLUMNS_SHA1 = "c62bac018b0e7bd3f3e4f77a1254ad1d0656ecf1"
COMPANY_ID = "company-export-test"


def _repo_facts_path() -> Path:
    return Path(__file__).resolve().parents[1] / "docs" / "jierui" / "voucher-import-template.facts.json"


def _draft(key: str, date: str = "2026-07-06") -> VoucherDraft:
    return VoucherDraft(
        voucher_key=key,
        voucher_date=date,
        lines=[
            VoucherLine(summary="采购供应商 A", account_code="1405", account_name="库存商品", direction="debit", amount="100.00"),
            VoucherLine(summary="采购供应商 A", account_code="222101", account_name="进项税额", direction="debit", amount="13.00"),
            VoucherLine(summary="采购供应商 A", account_code="2202", account_name="应付账款", direction="credit", amount="113.00"),
        ],
        source_invoice_nos=["12345678901234567890"],
        source_rows=["row:1"],
        balance_ok=True,
        review_tier="auto",
        generated_at="",
    )


def test_jierui_import_xlsx_matches_facts_columns_and_grouping(tmp_path: Path) -> None:
    facts = load_jierui_import_facts(_repo_facts_path())
    output = tmp_path / "凭证导入.xlsx"

    result = write_jierui_import_xlsx([_draft("k1"), _draft("k2", "2026-07-07")], {"1405": "库存商品"}, output, facts)

    assert result["voucher_count"] == 2
    assert hashlib.sha1(json.dumps(facts["columns"], ensure_ascii=False, separators=(",", ":")).encode("utf-8")).hexdigest() == FACTS_COLUMNS_SHA1
    workbook = load_workbook(output)
    worksheet = workbook[facts["sheet_name"]]
    headers = [cell.value for cell in worksheet[1]]
    assert headers == facts["columns"]
    rows = [[cell.value for cell in row] for row in worksheet.iter_rows(min_row=2)]
    assert [row[0:3] for row in rows[:3]] == [["记", "001", "2026-07-06"]] * 3
    assert [row[0:3] for row in rows[3:6]] == [["记", "002", "2026-07-07"]] * 3


def _ready_client(tmp_path: Path, monkeypatch, dates: tuple[str, ...] = ("2026-07-06",)) -> tuple[TestClient, object]:
    monkeypatch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
    company = tmp_path / "公司 A"
    watch_dir = company / "成本发票"
    watch_dir.mkdir(parents=True)
    with (watch_dir / "成本发票明细.csv").open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_HEADERS)
        writer.writeheader()
        for index, invoice_date in enumerate(dates, start=1):
            source = f"source-{index}.pdf"
            (watch_dir / source).write_bytes(f"invoice-{index}".encode())
            writer.writerow(
                {
                    "销售方": "供应商 A",
                    "购买方": "本公司",
                    "发票号码": f"1234567890123456789{index}",
                    "开票日期": invoice_date,
                    "内部项目名称": "项目 A",
                    "规格型号": "M1",
                    "单位": "吨",
                    "数量": "1",
                    "单价(除税)": "100.00",
                    "金额(除税)": "100.00",
                    "税率": "13%",
                    "税金": "13.00",
                    "价税合计": "113.00",
                    "源文件": source,
                }
            )
    facts = json.loads(_repo_facts_path().read_text(encoding="utf-8"))
    for item in facts["readiness"].values():
        item["status"] = "ready"
        item["evidence"] = "test fixture"
    facts_path = tmp_path / "docs" / "jierui" / "voucher-import-template.facts.json"
    facts_path.parent.mkdir(parents=True)
    facts_path.write_text(json.dumps(facts, ensure_ascii=False), encoding="utf-8")

    paths = company_bookkeeping_paths(company)
    paths.voucher_dir.mkdir(parents=True)
    periods = sorted({date[:7] for date in dates})
    capture_id = "capture-export-fixture"
    profile_seed = {
        "company_id": COMPANY_ID,
        "company_name": "公司 A",
        "company_tax_id": "91320000EXPORTTEST",
        "ledger_environment": "test",
        "ledger_provider": "jierui",
        "ledger_instance_key": "jierui-test-ledger",
        "ledger_name": "测试账套",
        "identity_method": "native_id",
        "accounting_standard": "小企业会计准则",
        "currency": "CNY",
    }
    ledger_identity = canonical_ledger_identity(profile_seed)
    account_records = [
        {"code": "1405", "name": "库存商品", "enabled": True, "is_leaf": True, "balance_direction": "debit", "required_aux_dimensions": [], "quantity_enabled": False, "foreign_currency_enabled": False},
        {"code": "222101", "name": "进项税额", "enabled": True, "is_leaf": True, "balance_direction": "debit", "required_aux_dimensions": [], "quantity_enabled": False, "foreign_currency_enabled": False},
        {"code": "2202", "name": "应付账款", "enabled": True, "is_leaf": True, "balance_direction": "credit", "required_aux_dimensions": [], "quantity_enabled": False, "foreign_currency_enabled": False},
    ]
    paths.account_table_json.write_text(
        json.dumps(
            {
                "schema_version": 2,
                "catalog_kind": "accounts",
                "company_id": COMPANY_ID,
                "ledger_environment": "test",
                "ledger_identity_sha256": ledger_identity,
                "capture_id": capture_id,
                "captured_at": "2026-07-10T00:00:00Z",
                "captured_by": "fixture",
                "content_sha256": canonical_sha256(account_records),
                "records": account_records,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    aux_records: list[dict] = []
    paths.aux_catalog_json.write_text(
        json.dumps(
            {
                "schema_version": 2,
                "catalog_kind": "auxiliary",
                "company_id": COMPANY_ID,
                "ledger_environment": "test",
                "ledger_identity_sha256": ledger_identity,
                "capture_id": capture_id,
                "captured_at": "2026-07-10T00:00:00Z",
                "captured_by": "fixture",
                "content_sha256": canonical_sha256(aux_records),
                "records": aux_records,
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    paths.company_facts_json.write_text(
        json.dumps(
            {
                "schema_version": 1,
                "revision": 1,
                "company_id": COMPANY_ID,
                "company_name": "公司 A",
                "company_tax_id": "91320000EXPORTTEST",
                "confirmed_by": "fixture",
                "confirmed_at": "2026-07-10T00:00:00Z",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    paths.ledger_profile_json.write_text(
        json.dumps(
            {
                **profile_seed,
                "schema_version": 2,
                "revision": 1,
                "ledger_identity_sha256": ledger_identity,
                "capture_id": capture_id,
                "taxpayer_profile": "一般纳税人",
                "open_periods": periods,
                "closed_through": "",
                "default_voucher_type": "记",
                "voucher_write_permission_confirmed": True,
                "account_table_sha256": file_sha256(paths.account_table_json),
                "aux_catalog_sha256": file_sha256(paths.aux_catalog_json),
                "confirmed_by": "tester",
                "confirmed_at": "2026-07-10T00:00:00Z",
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    mapping_binding = MappingStoreBinding(
        company_id=COMPANY_ID,
        ledger_environment="test",
        ledger_identity_sha256=ledger_identity,
        ledger_profile_sha256=file_sha256(paths.ledger_profile_json),
        account_table_sha256=file_sha256(paths.account_table_json),
        aux_catalog_sha256=file_sha256(paths.aux_catalog_json),
    )
    write_mapping(
        paths.account_mapping_json,
        [
            AccountMappingRule(
                rule_id="pending",
                match_seller="供应商 A",
                debit_account_code="1405",
                debit_account_name="库存商品",
                credit_account_code="2202",
                credit_account_name="应付账款",
                tax_account_code="222101",
                source="manual",
                confirmed_at="2026-07-06T00:00:00Z",
                confirmed_by="fixture",
            )
        ],
        expected_revision=0,
        binding=mapping_binding,
    )
    write_voucher_status(
        paths.voucher_status_json,
        {},
        company_id=COMPANY_ID,
        ledger_environment="test",
        ledger_identity_sha256=ledger_identity,
        ledger_profile_sha256=file_sha256(paths.ledger_profile_json),
    )
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps({"host": "127.0.0.1", "port": 8766, "watch_dir": str(watch_dir), "runtime_dir": "./runtime"}, ensure_ascii=False),
        encoding="utf-8",
    )
    client = TestClient(create_app(tmp_path, str(config_path)))
    assert client.post("/api/v1/bookkeeping/generate").json()["ok"] is True
    while True:
        listing = client.get("/api/v1/bookkeeping/vouchers").json()["items"]
        item = next((value for value in listing if value["status"] == "draft"), None)
        if item is None:
            break
        snapshot = item["snapshot"]
        invoice_no = snapshot["source_invoice_nos"][0]
        tax_evidence_path = company / "进项抵扣明细" / f"{invoice_no}.json"
        tax_evidence_path.parent.mkdir(exist_ok=True)
        tax_evidence_path.write_text('{"status":"confirmed"}', encoding="utf-8")
        (company / "入库证据").mkdir(exist_ok=True)
        for source_line in snapshot["source_lines"]:
            (company / "入库证据" / f"{source_line['source_line_id']}.json").write_text(
                '{"coverage":"full"}',
                encoding="utf-8",
            )
        receiving_refs = [
            {
                "evidence_id": f"receipt-{source_line['source_line_id']}",
                "evidence_type": "inventory_receipt",
                "subject_id": source_line["source_line_id"],
                "source_path": str((company / "入库证据" / f"{source_line['source_line_id']}.json").relative_to(company)),
                "source_sha256": file_sha256(company / "入库证据" / f"{source_line['source_line_id']}.json"),
                "source_revision": "receipt-v1",
                "coverage_state": "full",
                "confirmed_by": "tester",
                "confirmed_at": "2026-07-10T00:00:00Z",
            }
            for source_line in snapshot["source_lines"]
        ]
        response = client.put(
            f"/api/v1/bookkeeping/vouchers/{item['voucher_key']}/decision",
            json={
                "expected_store_revision": item["store_revision"],
                "expected_proposal_revision_hash": item["proposal_revision_hash"],
                "command_id": f"decision-{item['voucher_key']}",
                "decided_by": "tester",
                "business_class": "inventory_purchase",
                "payment_state": "unmatched",
                "payment_evidence_refs": [],
                "tax_treatment": "deductible",
                "tax_evidence_refs": [
                    {
                        "evidence_id": f"tax-{invoice_no}",
                        "evidence_type": "tax_usage_confirmation",
                        "subject_id": invoice_no,
                        "source_path": str(tax_evidence_path.relative_to(company)),
                        "source_sha256": file_sha256(tax_evidence_path),
                        "source_revision": "tax-v1",
                        "confirmed_by": "tester",
                        "confirmed_at": "2026-07-10T00:00:00Z",
                    }
                ],
                "receiving_state": "full",
                "receiving_evidence_refs": receiving_refs,
                "project_allocations": snapshot["project_allocations"],
                "lines": [
                    {"line_id": line["line_id"], "account_code": line["account_code"], "aux": line.get("aux", {})}
                    for line in snapshot["lines"]
                ],
            },
        )
        assert response.status_code == 200, response.text
    return client, paths


def _approve_all(client: TestClient) -> list[dict]:
    approved = []
    while True:
        listing = client.get("/api/v1/bookkeeping/vouchers").json()["items"]
        item = next((candidate for candidate in listing if candidate["status"] == "review_pending"), None)
        if item is None:
            return approved
        assert item["blockers"] == []
        response = client.post(
            f"/api/v1/bookkeeping/vouchers/{item['voucher_key']}/review",
            json={
                "action": "approve",
                "proposal_revision_hash": item["proposal_revision_hash"],
                "expected_store_revision": item["store_revision"],
                "reviewed_by": "tester",
                "command_id": f"approve-{len(approved) + 1}",
            },
        )
        assert response.status_code == 200
        approved.append(item)


def test_export_creates_one_immutable_single_period_batch(tmp_path: Path, monkeypatch) -> None:
    client, paths = _ready_client(tmp_path, monkeypatch)
    _approve_all(client)
    export_status = client.get("/api/v1/bookkeeping/export-status").json()

    response = client.post(
        "/api/v1/bookkeeping/export-import-file",
        json={**export_status["export_plan"], "requested_by": "tester", "command_id": "export-1"},
    )

    assert response.status_code == 200
    result = response.json()
    assert result["exported"] is True
    assert file_sha256(Path(result["file_path"])) == result["file_sha256"]
    manifest = Path(result["batch"]["manifest_path"])
    assert manifest.is_file()
    assert json.loads(manifest.read_text(encoding="utf-8"))["batch_id"] == result["batch_id"]
    store = load_voucher_status(paths.voucher_status_json)
    assert set(item.status for item in store.items.values()) == {"exported"}
    assert len(store.batches) == 1
    assert client.get("/api/v1/bookkeeping/export-status").json()["export_plan"] is None
    assert len(list(paths.batch_dir.glob("*/manifest.json"))) == 1


def test_multi_period_approved_items_cannot_enter_one_batch(tmp_path: Path, monkeypatch) -> None:
    client, _paths = _ready_client(tmp_path, monkeypatch, dates=("2026-06-30", "2026-07-01"))
    _approve_all(client)
    listing = client.get("/api/v1/bookkeeping/vouchers").json()
    items = [
        {"posting_key": item["posting_key"], "proposal_revision_hash": item["proposal_revision_hash"]}
        for item in listing["items"]
    ]

    response = client.post(
        "/api/v1/bookkeeping/export-import-file",
        json={
            "period": "2026-07",
            "items": items,
            "expected_store_revision": listing["store_revision"],
            "requested_by": "tester",
            "command_id": "export-multi",
        },
    )

    assert response.status_code == 409
    assert response.json()["error"]["blockers"][0]["code"] == "MULTI_PERIOD_BATCH"


@pytest.mark.parametrize("changed_target", ["source", "xlsx"])
def test_begin_revalidates_source_and_export_file_after_dry_run(tmp_path: Path, monkeypatch, changed_target: str) -> None:
    client, paths = _ready_client(tmp_path, monkeypatch)
    _approve_all(client)
    export_status = client.get("/api/v1/bookkeeping/export-status").json()
    exported = client.post(
        "/api/v1/bookkeeping/export-import-file",
        json={**export_status["export_plan"], "requested_by": "tester", "command_id": "export-before-begin"},
    ).json()
    batch = exported["batch"]
    dry_run = client.post(
        f"/api/v1/bookkeeping/import-batches/{batch['batch_id']}/dry-run",
        json={
            "ok": True,
            "file_sha256": batch["file_sha256"],
            "voucher_count": batch["expected_count"],
            "debit_total": batch["expected_debit_total"],
            "credit_total": batch["expected_credit_total"],
        },
    )
    assert dry_run.status_code == 200

    if changed_target == "source":
        (paths.company_dir / "成本发票" / "source-1.pdf").write_bytes(b"changed-source")
    else:
        Path(batch["file_path"]).write_bytes(b"changed-xlsx")

    response = client.post(
        f"/api/v1/bookkeeping/import-batches/{batch['batch_id']}/begin",
        json={
            "file_sha256": batch["file_sha256"],
            "ledger_name": batch["ledger_name"],
            "period": batch["period"],
            "authorized_by": "tester",
            "command_id": f"begin-after-{changed_target}-change",
        },
    )

    assert response.status_code in {400, 409}
    if changed_target == "source":
        assert response.json()["error"]["code"] == "EXECUTABILITY_BLOCKED"
        assert "SOURCE_FILE_CHANGED" in {item["code"] for item in response.json()["error"]["blockers"]}
    else:
        assert "SHA256" in response.json()["detail"]
    store = load_voucher_status(paths.voucher_status_json)
    assert store.batches[batch["batch_id"]].state == "awaiting_authorization"
    assert {item.status for item in store.items.values()} == {"exported"}
