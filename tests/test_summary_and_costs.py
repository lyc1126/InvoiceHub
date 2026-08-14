import hashlib
import re
import zipfile
from pathlib import Path

import pytest

from invoice_hub.projections.cost_analysis import DETAIL_HEADERS, INVOICE_REFERENCE_HEADERS, build_cost_analysis_outputs
from invoice_hub.projections.costs import CostProjectionService
from invoice_hub.projections.summary import build_summary
from invoice_hub.storage.files import read_csv_rows, read_json_object, write_csv_rows
from invoice_hub.storage import write_csv_rows as write_summary_rows
from invoice_hub.targets import load_config, target_profile_for


def test_xml_summary_and_duplicate_marker(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    xml = Path(__file__).parent / "fixtures" / "sample_invoice.xml"
    (watch / "a.xml").write_text(xml.read_text(encoding="utf-8"), encoding="utf-8")
    (watch / "b.xml").write_text(xml.read_text(encoding="utf-8"), encoding="utf-8")
    workspace = tmp_path / "workspace"

    result = build_summary(watch, workspace)

    assert result["count"] == 2
    assert (workspace / "发票汇总.csv").exists()
    assert result["records"][1]["duplicate"] is True


def test_text_extraction_rejects_invoice_number_as_amount(monkeypatch, tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "dzfp_10000000000000000001_样本.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "电子发票\n发票号码 10000000000000000001\n价税合计(小写)\n10000000000000000001\n销售方名称 Synthetic Seller A",
    )

    record = extract_invoice_record(invoice)

    assert record.invoice_number == "10000000000000000001"
    assert record.amount == ""


def test_text_extraction_rejects_invoice_number_fragments_as_amount(monkeypatch, tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "fragmented.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "发票号码 10000000000000000001\n价税合计(小写)\n2452\n发票代码 1000000000\n税率 13%",
    )

    record = extract_invoice_record(invoice)

    assert record.amount == ""


def test_text_extraction_accepts_nearby_real_money(monkeypatch, tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "valid.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "发票号码 10000000000000000001\n价税合计(小写) ￥26,550.98\n税额合计 3054.54\n金额合计 23496.44\n税率 13%",
    )

    record = extract_invoice_record(invoice)

    assert record.amount == "26550.98"
    assert record.tax_amount == "3054.54"
    assert record.pretax_amount == "23496.44"
    assert record.tax_rate == "13%"


def test_text_extraction_prefers_unique_pdf_amount_triple_before_invoice_identity(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "dzfp_10000000000000000019_vendor.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "\n".join(
            [
                "合",
                "计",
                "价税合计（大写）",
                "（小写）",
                "¥11320.75",
                "¥679.25",
                "壹万贰仟圆整",
                "¥12000.00",
                "电子发票（增值税普通发票）",
                "发票号码：",
                "开票日期：",
                "购买方信息",
                "销售方信息",
                "10000000000000000019",
                "2026年07月31日",
                "示例购买方有限公司",
                "91320000MA00000001",
                "示例销售方有限公司",
                "91320000MA00000002",
                "¥11320.75",
                "¥11320.75",
                "*金融服务*服务费",
                "6%",
            ]
        ),
    )

    record = extract_invoice_record(invoice)

    assert record.invoice_number == "10000000000000000019"
    assert record.invoice_date == "2026-07-31"
    assert record.buyer == "示例购买方有限公司"
    assert record.seller == "示例销售方有限公司"
    assert record.amount == "12000.00"
    assert record.pretax_amount == "11320.75"
    assert record.tax_amount == "679.25"
    assert record.tax_rate == "6%"


def test_digital_party_sequence_does_not_promote_two_detail_decimals_to_header_amounts(
    monkeypatch,
    tmp_path: Path,
) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "dzfp_10000000000000000020_vendor.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "\n".join(
            [
                "电子发票（增值税普通发票）",
                "发票号码：",
                "开票日期：",
                "购买方信息",
                "销售方信息",
                "10000000000000000020",
                "2026年07月31日",
                "示例购买方有限公司",
                "91320000MA00000001",
                "示例销售方有限公司",
                "91320000MA00000002",
                "11320.75",
                "11320.75",
                "6%",
            ]
        ),
    )

    record = extract_invoice_record(invoice)

    assert record.buyer == "示例购买方有限公司"
    assert record.seller == "示例销售方有限公司"
    assert record.amount == ""
    assert record.pretax_amount == ""
    assert record.tax_amount == ""


@pytest.mark.parametrize(
    ("text", "expected"),
    [
        (
            "价税合计（小写）\n¥100.00\n¥0.00\n壹佰圆整\n¥100.00",
            {"pretax_amount": "100.00", "tax_amount": "0.00", "amount": "100.00"},
        ),
        (
            "价税合计（小写）\n¥-100.00\n¥-13.00\n负壹佰壹拾叁圆整\n¥-113.00",
            {"pretax_amount": "-100.00", "tax_amount": "-13.00", "amount": "-113.00"},
        ),
        (
            "价税合计（小写）\n¥100.00\n¥13.00\n壹佰壹拾叁圆整\n¥113.00"
            "\n明细\n¥100.00\n¥100.00",
            {"pretax_amount": "100.00", "tax_amount": "13.00", "amount": "113.00"},
        ),
    ],
)
def test_pdf_amount_triple_accepts_safe_arithmetic_edges(
    text: str,
    expected: dict[str, str],
) -> None:
    import invoice_hub.extraction.parsers as parsers

    assert parsers._extract_pdf_amount_triple(text) == expected


@pytest.mark.parametrize(
    "text",
    [
        "价税合计（小写）\n100.00\n13.00\n113.00",
        "价税合计（小写）\n¥100.00\n¥13.00\n¥120.00",
        (
            "价税合计（小写）\n¥100.00\n¥13.00\n¥113.00"
            "\n¥200.00\n¥26.00\n¥226.00"
        ),
        "价税合计（小写）\n¥100.00\n¥13.00\f¥113.00",
    ],
)
def test_pdf_amount_triple_rejects_missing_ambiguous_or_cross_page_evidence(text: str) -> None:
    import invoice_hub.extraction.parsers as parsers

    assert parsers._extract_pdf_amount_triple(text) == {}


def test_labeled_money_fallback_stays_on_same_or_adjacent_logical_line() -> None:
    import invoice_hub.extraction.parsers as parsers

    labels = ["价税合计"]
    assert parsers._first_money_near("价税合计（小写） ￥120.00", labels) == "120.00"
    assert parsers._first_money_near("价税合计（小写）\n￥120.00", labels) == "120.00"
    assert parsers._first_money_near("价税合计（小写）\n发票号码 12345678\n￥120.00", labels) == ""


def test_text_extraction_handles_digital_invoice_value_sequence(monkeypatch, tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "dzfp_00000000000000000001_vendor.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "\n".join(
            [
                "合",
                "计",
                "价税合计（大写）",
                "（小写）",
                "电子发票（增值税专用发票）",
                "发票号码：",
                "开票日期：",
                "购",
                "买",
                "方",
                "信",
                "息",
                "销",
                "售",
                "方",
                "信",
                "息",
                "00000000000000000001",
                "2026年01月15日",
                "示例采购方甲有限公司",
                "000000000000000001",
                "示例销售方甲有限公司",
                "000000000000000002",
                "¥120.00",
                "¥15.60",
                "壹佰叁拾伍圆陆角整",
                "¥135.60",
                "*黑色金属冶炼压延品*螺",
                "13%",
            ]
        ),
    )

    record = extract_invoice_record(invoice)

    assert record.invoice_number == "00000000000000000001"
    assert record.invoice_date == "2026-01-15"
    assert record.buyer == "示例采购方甲有限公司"
    assert record.seller == "示例销售方甲有限公司"
    assert record.invoice_type == "增值税专用发票"
    assert record.pretax_amount == "120.00"
    assert record.tax_amount == "15.60"
    assert record.amount == "135.60"


def test_text_extraction_keeps_short_person_seller_and_total_amount(monkeypatch, tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record
    import invoice_hub.extraction.parsers as parsers

    invoice = tmp_path / "dzfp_00000000000000000002_示例销售方乙.pdf"
    invoice.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "\n".join(
            [
                "电子发票（增值税专用发票）",
                "发票号码：",
                "开票日期：",
                "购买方信息",
                "销售方信息",
                "价税合计（大写）",
                "（小写）",
                "00000000000000000002",
                "2026年02月20日",
                "示例采购方乙有限公司",
                "000000000000000003",
                "示例销售方乙",
                "000000000000000004",
                "¥100.00",
                "¥3.00",
                "壹佰零叁圆整",
                "¥103.00",
                "示例销售方乙",
                "*黑色金属冶炼压延品*钢筋网片",
                "3%",
            ]
        ),
    )

    record = extract_invoice_record(invoice)

    assert record.invoice_number == "00000000000000000002"
    assert record.invoice_date == "2026-02-20"
    assert record.buyer == "示例采购方乙有限公司"
    assert record.seller == "示例销售方乙"
    assert record.seller != "壹佰零叁圆整"
    assert record.pretax_amount == "100.00"
    assert record.tax_amount == "3.00"
    assert record.amount == "103.00"


def _write_minimal_ofd(path: Path, values: dict[str, str]) -> None:
    refs = {
        "InvoiceNo": ("6922", values["invoice_number"]),
        "IssueDate": ("6923", values["invoice_date"]),
        "Buyer/BuyerName": ("6924", values["buyer"]),
        "Buyer/BuyerTaxID": ("6926", values.get("buyer_tax_id", "000000000000000003")),
        "Seller/SellerName": ("6927", values["seller"]),
        "Seller/SellerTaxID": ("6928", values.get("seller_tax_id", "000000000000000004")),
        "TaxExclusiveTotalAmount": ("6930", values["pretax_amount"]),
        "TaxTotalAmount": ("6932", values["tax_amount"]),
        "TaxInclusiveTotalAmount": ("6936", values["amount"]),
        "TaxScheme": ("6940", values["tax_rate"]),
    }
    custom_items = []
    for field_path, (object_id, _value) in refs.items():
        open_tags = "".join(f"<ofd:{part}>" for part in field_path.split("/"))
        close_tags = "".join(f"</ofd:{part}>" for part in reversed(field_path.split("/")))
        custom_items.append(f'{open_tags}<ofd:ObjectRef PageRef="61">{object_id}</ofd:ObjectRef>{close_tags}')
    text_objects = [
        f'<ofd:TextObject ID="{object_id}"><ofd:TextCode>{value}</ofd:TextCode></ofd:TextObject>'
        for object_id, value in refs.values()
    ]
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(
            "Doc_0/Tags/CustomTag.xml",
            '<?xml version="1.0" encoding="UTF-8"?><ofd:root xmlns:ofd="http://www.ofdspec.org/2016">'
            + "".join(custom_items)
            + "</ofd:root>",
        )
        archive.writestr(
            "Doc_0/Pages/Page_0/Content.xml",
            '<?xml version="1.0" encoding="UTF-8"?><ofd:Page xmlns:ofd="http://www.ofdspec.org/2016"><ofd:Content>'
            + "".join(text_objects)
            + "</ofd:Content></ofd:Page>",
        )


def _cost_xml_text(invoice_number: str = "10000000000000000013") -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<EInvoice>
  <EInvoiceData>
    <SellerInformation><SellerName>示例销售方有限公司</SellerName></SellerInformation>
    <BuyerInformation><BuyerName>示例购买方有限公司</BuyerName></BuyerInformation>
    <BasicInformation>
      <InvoiceType>增值税专用发票</InvoiceType>
      <BusinessType>标准电子发票</BusinessType>
      <InvoiceNumber>{invoice_number}</InvoiceNumber>
      <IssueDate>2026-05-30</IssueDate>
      <TotalAmWithoutTax>300.00</TotalAmWithoutTax>
      <TotalTaxAm>39.00</TotalTaxAm>
      <TotalTaxIncludedAmount>339.00</TotalTaxIncludedAmount>
    </BasicInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>12E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>2</Quantity>
      <UnPrice>100</UnPrice>
      <Amount>200</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>26</ComTaxAm>
      <TotalTaxIncludedAmount>226</TotalTaxIncludedAmount>
    </IssuItemInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>14E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>1</Quantity>
      <UnPrice>100</UnPrice>
      <Amount>100</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>13</ComTaxAm>
      <TotalTaxIncludedAmount>113</TotalTaxIncludedAmount>
    </IssuItemInformation>
  </EInvoiceData>
</EInvoice>
"""


def _write_cost_ofd(path: Path, invoice_number: str = "10000000000000000013") -> None:
    refs: list[tuple[str, str]] = [
        ("InvoiceNo", invoice_number),
        ("IssueDate", "2026年05月30日"),
        ("Buyer/BuyerName", "示例购买方有限公司"),
        ("Seller/SellerName", "示例销售方有限公司"),
        ("TaxExclusiveTotalAmount", "300.00"),
        ("TaxTotalAmount", "39.00"),
        ("TaxInclusiveTotalAmount", "339.00"),
        ("TaxScheme", "13%"),
        ("Item", "*材料*钢"),
        ("Item", "筋"),
        ("Item", "*材料*钢"),
        ("Item", "筋"),
        ("Specification", "12E"),
        ("Specification", "14E"),
        ("MeasurementDimension", "吨"),
        ("MeasurementDimension", "吨"),
        ("Quantity", "2"),
        ("Quantity", "1"),
        ("Price", "100"),
        ("Price", "100"),
        ("Amount", "200"),
        ("Amount", "100"),
        ("TaxScheme", "13%"),
        ("TaxScheme", "13%"),
        ("TaxAmount", "26"),
        ("TaxAmount", "13"),
    ]
    custom_items = []
    text_objects = []
    for index, (field_path, value) in enumerate(refs, start=7000):
        object_id = str(index)
        open_tags = "".join(f"<ofd:{part}>" for part in field_path.split("/"))
        close_tags = "".join(f"</ofd:{part}>" for part in reversed(field_path.split("/")))
        custom_items.append(f"{open_tags}<ofd:ObjectRef>{object_id}</ofd:ObjectRef>{close_tags}")
        text_objects.append(f'<ofd:TextObject ID="{object_id}"><ofd:TextCode>{value}</ofd:TextCode></ofd:TextObject>')
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(
            "Doc_0/Tags/CustomTag.xml",
            '<?xml version="1.0" encoding="UTF-8"?><ofd:root xmlns:ofd="http://www.ofdspec.org/2016">'
            + "".join(custom_items)
            + "</ofd:root>",
        )
        archive.writestr(
            "Doc_0/Pages/Page_0/Content.xml",
            '<?xml version="1.0" encoding="UTF-8"?><ofd:Page xmlns:ofd="http://www.ofdspec.org/2016"><ofd:Content>'
            + "".join(text_objects)
            + "</ofd:Content></ofd:Page>",
        )


def test_ofd_structured_custom_tag_fields_are_preferred(tmp_path: Path) -> None:
    from invoice_hub.extraction import extract_invoice_record

    invoice = tmp_path / "示例销售方丙有限公司_10100.ofd"
    _write_minimal_ofd(
        invoice,
        {
            "invoice_number": "00000000000000000003",
            "invoice_date": "2026年03月10日",
            "buyer": "示例采购方乙有限公司",
            "seller": "示例销售方丙有限公司",
            "pretax_amount": "100.00",
            "tax_amount": "1.00",
            "amount": "101.00",
            "tax_rate": "1%",
        },
    )

    record = extract_invoice_record(invoice)

    assert record.file_type == "ofd"
    assert record.invoice_number == "00000000000000000003"
    assert record.invoice_date == "2026-03-10"
    assert record.buyer == "示例采购方乙有限公司"
    assert record.seller == "示例销售方丙有限公司"
    assert record.pretax_amount == "100.00"
    assert record.tax_amount == "1.00"
    assert record.amount == "101.00"
    assert record.tax_rate == "1%"


def test_xml_cost_detail_rows_are_parsed(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    source = watch / "dzfp_10000000000000000013.xml"
    source.write_text(_cost_xml_text(), encoding="utf-8")

    result = build_cost_analysis_outputs(watch, watch)
    rows = read_csv_rows(watch / "成本发票明细.csv")

    assert result["source_invoice_count"] == 1
    assert result["format_counts"]["xml"] == 1
    assert result["detail_count"] == 2
    assert len(rows) == 2
    assert rows[0]["销售方"] == "示例销售方有限公司"
    assert rows[0]["内部项目名称"] == "钢筋"
    assert rows[0]["发票代码(**内文字)"] == "材料"
    assert rows[0]["规格型号"] == "12E"
    assert rows[0]["税率"] == "13%"
    assert rows[0]["价税合计"] == "226"


def test_ofd_cost_detail_rows_are_parsed_from_custom_tag(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    source = watch / "dzfp_10000000000000000013.ofd"
    _write_cost_ofd(source)

    result = build_cost_analysis_outputs(watch, watch)
    rows = read_csv_rows(watch / "成本发票明细.csv")

    assert result["source_invoice_count"] == 1
    assert result["format_counts"]["ofd"] == 1
    assert result["detail_count"] == 2
    assert len(rows) == 2
    assert rows[0]["销售方"] == "示例销售方有限公司"
    assert rows[0]["内部项目名称"] == "钢筋"
    assert rows[0]["规格型号"] == "12E"
    assert rows[0]["数量"] == "2"
    assert rows[0]["价税合计"] == "226"


def test_cost_rebuild_prefers_structured_xml_over_same_invoice_pdf(monkeypatch, tmp_path: Path) -> None:
    import invoice_hub.projections.cost_analysis as cost_analysis

    watch = tmp_path / "发票文件"
    watch.mkdir()
    invoice_number = "10000000000000000013"
    (watch / f"dzfp_{invoice_number}.xml").write_text(_cost_xml_text(invoice_number), encoding="utf-8")
    (watch / f"dzfp_{invoice_number}.pdf").write_bytes(b"%PDF")
    pdf_row = _cost_detail_row(发票号码=invoice_number, 内部项目名称="PDF项目", 源文件=f"dzfp_{invoice_number}.pdf")
    monkeypatch.setattr(cost_analysis, "analyze_pdf_costs", lambda _path, metadata=None: {"source": Path(_path).name, "rows": [pdf_row], "invoice": {}, "status": "ok", "message": ""})

    result = build_cost_analysis_outputs(watch, watch)
    rows = read_csv_rows(watch / "成本发票明细.csv")

    assert result["pdf_count"] == 1
    assert result["source_invoice_count"] == 1
    assert result["detail_count"] == 2
    assert {row["内部项目名称"] for row in rows} == {"钢筋"}
    assert all(row["源文件"].endswith(".xml") for row in rows)


def test_cost_sync_counts_same_invoice_xml_and_ofd_as_one_source(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    invoice_number = "10000000000000000013"
    (watch / f"dzfp_{invoice_number}.xml").write_text(_cost_xml_text(invoice_number), encoding="utf-8")
    _write_cost_ofd(watch / f"dzfp_{invoice_number}.ofd", invoice_number)

    build_summary(watch, workspace)
    service = CostProjectionService(watch, workspace, "target")
    service.rebuild()
    sync = service.snapshot().sync

    assert sync.source_invoice_count == 1
    assert sync.parsed_invoice_count == 1
    assert sync.checked_invoice_count == 1
    assert sync.pending_count == 0
    assert sync.sync_state == "fresh"


def test_same_invoice_family_uses_xml_parties_for_pdf(monkeypatch, tmp_path: Path) -> None:
    import invoice_hub.extraction.parsers as parsers

    watch = tmp_path / "发票文件"
    watch.mkdir()
    xml = watch / "dzfp_00000000000000000001_20260115101537.xml"
    pdf = watch / "dzfp_00000000000000000001_示例销售方甲有限公司_20260115101537.pdf"
    xml.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
<Invoice><invoiceNo>00000000000000000001</invoiceNo><sellerName>示例销售方甲有限公司</sellerName><buyerName>示例采购方甲有限公司</buyerName><totalAmount>135.60</totalAmount></Invoice>""",
        encoding="utf-8",
    )
    pdf.write_bytes(b"%PDF")
    monkeypatch.setattr(parsers, "_text_from_pdf", lambda _path: "发票号码 00000000000000000001\n价税合计(小写) 135.60")

    result = build_summary(watch, tmp_path / "workspace")
    pdf_row = next(row for row in result["records"] if row["file_type"] == "pdf")

    assert pdf_row["seller"] == "示例销售方甲有限公司"
    assert pdf_row["buyer"] == "示例采购方甲有限公司"


def test_same_invoice_family_replaces_suspicious_non_empty_parties(monkeypatch, tmp_path: Path) -> None:
    import invoice_hub.extraction.parsers as parsers

    watch = tmp_path / "发票文件"
    watch.mkdir()
    xml = watch / "dzfp_00000000000000000004.xml"
    pdf = watch / "dzfp_00000000000000000004.pdf"
    xml.write_text(
        """<?xml version="1.0" encoding="UTF-8"?>
<Invoice><invoiceNo>00000000000000000004</invoiceNo><sellerName>示例销售方甲有限公司</sellerName><buyerName>示例采购方甲有限公司</buyerName><totalAmount>103.00</totalAmount></Invoice>""",
        encoding="utf-8",
    )
    pdf.write_bytes(b"%PDF")
    monkeypatch.setattr(
        parsers,
        "_text_from_pdf",
        lambda _path: "发票号码 00000000000000000004\n购买方名称 6958\n销售方名称 肆万圆整\n价税合计(小写) ￥103.00",
    )

    result = build_summary(watch, tmp_path / "workspace")
    pdf_row = next(row for row in result["records"] if row["file_type"] == "pdf")

    assert pdf_row["seller"] == "示例销售方甲有限公司"
    assert pdf_row["buyer"] == "示例采购方甲有限公司"


def test_cost_reference_status_locks_amount_snapshot(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    detail_rows = [
        {
            "销售方": "A",
            "购买方": "B",
            "发票号码": "1",
            "开票日期": "2026-05-23",
            "备注项目名称": "",
            "内部项目名称": "项目",
            "规格型号": "规格",
            "单位": "吨",
            "数量": "10",
            "单价(除税)": "10",
            "金额(除税)": "100",
            "税率": "0.13",
            "税金": "13",
            "价税合计": "113",
            "发票代码(**内文字)": "材料",
            "源文件": str(watch / "a.pdf"),
        }
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, detail_rows)
    service = CostProjectionService(watch, workspace, "target")
    snapshot = service.snapshot()
    key = snapshot.invoice_reference[0]["key"]

    result = service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "5"}]})
    payload = read_json_object(watch / "成本开票状态.json")

    assert result["ok"] is True
    assert "invoiced_reference_total_with_tax" in payload["items"][key]
    assert payload["items"][key]["invoiced_quantity"] == "5"

    # 新增同汇总键数量后，已开快照不得跟着新总量放大。
    detail_rows.append({**detail_rows[0], "数量": "10", "金额(除税)": "100", "税金": "13", "价税合计": "113", "源文件": str(watch / "b.pdf")})
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, detail_rows)
    snapshot_after = service.snapshot()
    row = snapshot_after.invoice_reference[0]
    assert row["quantity"] == 20.0
    assert row["invoiced_quantity"] == 5.0
    assert row["invoiced_reference_total_with_tax"] > 0
    assert row["uninvoiced_quantity"] == 15.0

    stats = snapshot_after.reference_status_stats
    assert stats["inventory_total_with_tax"] == 226.0
    assert stats["invoiced_reference_total_with_tax"] > 0
    assert stats["uninvoiced_reference_total_with_tax"] > 0


def test_project_summary_includes_tax_average_unit_price(monkeypatch, tmp_path: Path) -> None:
    from openpyxl import load_workbook
    import invoice_hub.projections.cost_analysis as cost_analysis

    watch = tmp_path / "发票文件"
    watch.mkdir()
    row = _cost_detail_row(**{"数量": "2", "单价(除税)": "9999", "金额(除税)": "6842", "税率": "13%", "税金": "615.78", "价税合计": "7457.78", "源文件": str(watch / "a.pdf")})
    (watch / "a.pdf").write_bytes(b"%PDF")
    monkeypatch.setattr(cost_analysis, "analyze_pdf_costs", lambda _path, metadata=None: {"source": "a.pdf", "rows": [row], "invoice": {}, "status": "ok", "message": ""})

    build_cost_analysis_outputs(watch, watch)
    wb = load_workbook(watch / "成本发票汇总.xlsx", read_only=True, data_only=True)
    try:
        detail_headers = [cell.value for cell in next(wb["发票明细"].iter_rows(min_row=1, max_row=1))]
        detail_data_row = next(wb["发票明细"].iter_rows(min_row=2, max_row=2, values_only=True))
        headers = [cell.value for cell in next(wb["项目规格汇总"].iter_rows(min_row=1, max_row=1))]
        data_row = next(wb["项目规格汇总"].iter_rows(min_row=2, max_row=2, values_only=True))
    finally:
        wb.close()

    assert "平均单价(含税)" in detail_headers
    detail = dict(zip(detail_headers, detail_data_row))
    assert detail["单价(除税)"] == "9999"
    assert detail["平均单价(含税)"] == 3728.89
    assert "平均单价(除税)" in headers
    assert "平均单价(含税)" in headers
    assert "库存平均单价(除税)" in headers
    assert "库存平均单价(含税)" in headers
    assert "采购参考平均单价(含税)" in headers
    project = dict(zip(headers, data_row))
    assert project["平均单价(除税)"] == 3421
    assert project["平均单价(含税)"] == 3728.89
    assert project["库存平均单价(除税)"] == 3421
    assert project["库存平均单价(含税)"] == 3728.89
    assert project["采购参考平均单价(含税)"] == 3728.89

    wb = load_workbook(watch / "成本发票汇总.xlsx", read_only=True, data_only=True)
    try:
        reference_headers = [cell.value for cell in next(wb["开票参考"].iter_rows(min_row=1, max_row=1))]
        reference_data_row = next(wb["开票参考"].iter_rows(min_row=2, max_row=2, values_only=True))
    finally:
        wb.close()
    reference = dict(zip(reference_headers, reference_data_row))
    assert reference["平均单价(含税)"] == 4174.99
    assert reference["税金合计"] == 960.62
    assert reference["价税合计"] == 8349.98


def test_project_summary_uses_weighted_stock_average_and_separate_purchase_reference(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    rows = [
        _cost_detail_row(**{"数量": "2", "单价(除税)": "100", "金额(除税)": "200", "税率": "9%", "税金": "18", "价税合计": "218", "源文件": "a.pdf"}),
        _cost_detail_row(**{"数量": "1", "单价(除税)": "200", "金额(除税)": "200", "税率": "13%", "税金": "26", "价税合计": "226", "源文件": "b.pdf"}),
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)
    snapshot = CostProjectionService(watch, workspace, "target").snapshot()

    project = snapshot.project_summary[0]
    reference = snapshot.invoice_reference[0]

    assert project["平均单价(除税)"] == 133.33
    assert project["平均单价(含税)"] == 148.0
    assert project["库存平均单价(除税)"] == 133.33
    assert project["库存平均单价(含税)"] == 148.0
    assert project["采购参考平均单价(含税)"] == 167.5
    assert reference["average_unit_price"] == 133.33333333333334
    assert reference["average_unit_price_with_tax"] == 148.0
    assert reference["purchase_reference_average_unit_price_with_tax"] == 167.5
    assert reference["reference_average_unit_price"] == 144.0
    assert reference["reference_average_unit_price_with_tax"] == 162.72
    assert reference["reference_amount"] == 432.0
    assert reference["reference_tax_amount"] == 56.16
    assert reference["reference_total_with_tax"] == 488.16


def test_project_summary_tax_average_unit_price_rejects_incomplete_amount_rows(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    rows = [
        _cost_detail_row(**{"数量": "2", "金额(除税)": "200", "税金": "18", "价税合计": "218", "源文件": "a.pdf"}),
        _cost_detail_row(**{"数量": "1", "金额(除税)": "", "税金": "26", "价税合计": "26", "源文件": "b.pdf"}),
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)
    snapshot = CostProjectionService(watch, workspace, "target").snapshot()

    assert snapshot.items[0]["平均单价(含税)"] == 109.0
    assert snapshot.items[1]["平均单价(含税)"] == ""
    assert snapshot.project_summary[0]["平均单价(含税)"] == ""
    assert snapshot.invoice_reference[0]["average_unit_price_with_tax"] == 0


def test_cost_snapshot_refreshes_old_cost_schema_from_existing_detail(tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row(**{"数量": "2", "单价(除税)": "3421", "金额(除税)": "6842", "税率": "9%", "税金": "615.78", "价税合计": "7457.78"})
    legacy_detail_headers = [header for header in DETAIL_HEADERS if header != "平均单价(含税)"]
    write_csv_rows(watch / "成本发票明细.csv", legacy_detail_headers, [row])
    wb = Workbook()
    ws = wb.active
    ws.title = "发票明细"
    ws.append(legacy_detail_headers)
    ws.append([row.get(header, "") for header in legacy_detail_headers])
    ws = wb.create_sheet("项目规格汇总")
    ws.append(["销售方", "发票代码(**内文字)", "内部项目名称", "规格型号", "单位", "数量合计", "平均单价(除税)", "金额(除税)合计", "税金合计", "价税合计", "涉及发票号码"])
    ws.append(["A", "材料", "项目", "规格", "吨", 2, 3421, 6842, 615.78, 7457.78, "1"])
    ws = wb.create_sheet("开票参考")
    legacy_reference_headers = [header for header in INVOICE_REFERENCE_HEADERS if header != "平均单价(含税)"]
    ws.append(legacy_reference_headers)
    ws.append(["材料", "项目", "规格", "吨", 2, 0, 2, "未开具", 3694.68, 7389.36, 665.04, 8054.4, 0, 7389.36, 665.04, 8054.4, "8%", ""])
    ws = wb.create_sheet("发票校验")
    ws.append(["源文件", "发票号码", "明细行数", "发票金额(除税)", "解析金额(除税)", "差异(除税)", "校验状态", "说明"])
    wb.save(watch / "成本发票汇总.xlsx")

    snapshot = CostProjectionService(watch, workspace, "target").snapshot()

    assert snapshot.items[0]["平均单价(含税)"] == 3728.89
    assert snapshot.project_summary[0]["平均单价(含税)"] == 3728.89
    assert snapshot.invoice_reference[0]["reference_average_unit_price_with_tax"] == 4174.9884
    with (watch / "成本发票明细.csv").open("r", encoding="utf-8-sig", newline="") as handle:
        assert "平均单价(含税)" in handle.readline()
    refreshed = load_workbook(watch / "成本发票汇总.xlsx", read_only=True, data_only=True)
    try:
        reference_headers = [cell.value for cell in next(refreshed["开票参考"].iter_rows(min_row=1, max_row=1))]
        assert "平均单价(含税)" in reference_headers
        row_values = next(refreshed["开票参考"].iter_rows(min_row=2, max_row=2, values_only=True))
        reference = dict(zip(reference_headers, row_values))
        assert reference["平均单价(含税)"] == 4174.99
        assert reference["税金合计"] == 960.62
        assert reference["价税合计"] == 8349.98
    finally:
        refreshed.close()


def test_cost_reference_markup_rate_can_change_without_repricing_locked_snapshot(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    service = CostProjectionService(watch, workspace, "target")
    key = service.snapshot().invoice_reference[0]["key"]

    service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "5"}]})
    service.save_reference_status(
        {
            "items": [
                {
                    "key": key,
                    "invoiced_quantity": "5",
                    "reference_markup_rate_percent": "20",
                    "reference_markup_locked": True,
                }
            ]
        }
    )
    after_rate_change = service.snapshot().invoice_reference[0]

    assert after_rate_change["markup_rate"] == "20%"
    assert after_rate_change["reference_markup_rate_percent"] == "20"
    assert after_rate_change["reference_markup_locked"] is True
    assert after_rate_change["reference_total_with_tax"] == 135.6
    assert after_rate_change["invoiced_reference_total_with_tax"] == 61.02
    assert after_rate_change["uninvoiced_reference_total_with_tax"] == 74.58

    service.save_reference_status(
        {
            "items": [
                {
                    "key": key,
                    "invoiced_quantity": "6",
                    "reference_markup_rate_percent": "20",
                    "reference_markup_locked": True,
                }
            ]
        }
    )
    after_manual_change = service.snapshot().invoice_reference[0]
    payload = read_json_object(watch / "成本开票状态.json")

    assert after_manual_change["invoiced_reference_total_with_tax"] == 81.36
    assert payload["items"][key]["reference_markup_rate"] == "0.2"
    assert payload["items"][key]["reference_markup_rate_percent"] == "20"
    assert payload["items"][key]["reference_markup_locked"] is True
    assert payload["items"][key]["invoiced_reference_total_with_tax"] == "81.36"


def test_cost_reference_markup_rate_is_row_level_and_stats_sum_rows(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    rows = [
        _cost_detail_row(内部项目名称="项目A", 数量="10", **{"金额(除税)": "100", "税金": "13", "价税合计": "113"}),
        _cost_detail_row(内部项目名称="项目B", 数量="2", 源文件="b.pdf", **{"金额(除税)": "200", "税金": "26", "价税合计": "226"}),
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)
    service = CostProjectionService(watch, workspace, "target")
    before = service.snapshot().invoice_reference
    keys = {row["内部项目名称"]: row["key"] for row in before}

    service.save_reference_status(
        {
            "items": [
                {"key": keys["项目A"], "invoiced_quantity": "0", "reference_markup_rate_percent": "20", "reference_markup_locked": True},
                {"key": keys["项目B"], "invoiced_quantity": "0", "reference_markup_rate_percent": "5", "reference_markup_locked": True},
            ]
        }
    )
    snapshot = service.snapshot()
    by_name = {row["内部项目名称"]: row for row in snapshot.invoice_reference}

    assert by_name["项目A"]["markup_rate"] == "20%"
    assert by_name["项目A"]["reference_total_with_tax"] == 135.6
    assert by_name["项目B"]["markup_rate"] == "5%"
    assert by_name["项目B"]["reference_total_with_tax"] == 237.3
    assert snapshot.reference_status_stats["reference_total_with_tax"] == 372.9


def test_cost_reference_status_keeps_zero_quantity_row_markup(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    service = CostProjectionService(watch, workspace, "target")
    key = service.snapshot().invoice_reference[0]["key"]

    service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "0", "reference_markup_rate_percent": "15", "reference_markup_locked": True}]})
    payload = read_json_object(watch / "成本开票状态.json")
    reference = service.snapshot().invoice_reference[0]

    assert key in payload["items"]
    assert payload["items"][key]["invoiced_quantity"] == "0"
    assert payload["items"][key]["reference_markup_rate_percent"] == "15"
    assert reference["markup_rate"] == "15%"
    assert reference["reference_total_with_tax"] == 129.95


def test_cost_reference_status_rejects_invalid_or_over_max_quantity(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    service = CostProjectionService(watch, workspace, "target")
    key = service.snapshot().invoice_reference[0]["key"]

    with pytest.raises(ValueError, match="已开数量不是有效数字"):
        service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "1吨"}]})

    with pytest.raises(ValueError, match="已开数量不能大于数量合计"):
        service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "10.001"}]})

    with pytest.raises(ValueError, match="开票加价率不是有效数字"):
        service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "0", "reference_markup_rate_percent": "abc"}]})

    service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "10"}]})
    payload = read_json_object(watch / "成本开票状态.json")

    assert payload["items"][key]["invoiced_quantity"] == "10"
    assert payload["items"][key]["invoice_status"] == "已开具"
    assert "invoiced_reference_total_with_tax" in payload["items"][key]


def _reference_key(*values: object) -> str:
    raw = "\x1f".join(re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ").strip()).strip() for value in values)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _cost_detail_row(**overrides: object) -> dict:
    row = {
        "销售方": "A",
        "购买方": "B",
        "发票号码": "1",
        "开票日期": "2026-05-23",
        "备注项目名称": "",
        "内部项目名称": "项目",
        "规格型号": "规格",
        "单位": "吨",
        "数量": "10",
        "单价(除税)": "10",
        "金额(除税)": "100",
        "税率": "0.13",
        "税金": "13",
        "价税合计": "113",
        "发票代码(**内文字)": "材料",
        "源文件": "a.pdf",
    }
    row.update({key: str(value) for key, value in overrides.items()})
    return row


def test_cost_reference_status_reads_old_sha1_json_key(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    key = _reference_key("材料", "项目", "规格", "吨")
    (watch / "成本开票状态.json").write_text(
        """{
  "version": 1,
  "updated_at": "2026-05-21T07:37:45Z",
  "items": {
    "%s": {
      "reference_key": "%s",
      "item_code": "材料",
      "item_name": "项目",
      "spec": "规格",
      "unit": "吨",
      "invoiced_quantity": "5",
      "invoiced_reference_amount": "54",
      "invoiced_reference_tax": "7.02",
      "invoiced_reference_total_with_tax": "61.02",
      "updated_at": "2026-05-21T07:37:45Z"
    }
  }
}"""
        % (key, key),
        encoding="utf-8",
    )

    snapshot = CostProjectionService(watch, workspace, "target").snapshot()
    reference = snapshot.invoice_reference[0]

    assert reference["key"] == key
    assert reference["invoiced_quantity"] == 5.0
    assert reference["average_unit_price_with_tax"] == 11.3
    assert reference["reference_average_unit_price_with_tax"] == 12.204
    assert reference["invoiced_reference_total_with_tax"] == 61.02
    assert reference["status_updated_at"] == "2026-05-21T07:37:45Z"


def test_cost_reference_status_legacy_quantity_is_saved_with_snapshot_fields(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    key = _reference_key("材料", "项目", "规格", "吨")
    (watch / "成本开票状态.json").write_text(
        """{"version": 1, "items": {"%s": {"reference_key": "%s", "invoiced_quantity": "5", "updated_at": "2026-05-15T15:48:04Z"}}}"""
        % (key, key),
        encoding="utf-8",
    )
    service = CostProjectionService(watch, workspace, "target")

    before = service.snapshot().invoice_reference[0]
    assert before["invoiced_quantity"] == 5.0
    assert before["invoiced_reference_total_with_tax"] == 61.02

    service.save_reference_status({"items": [{"key": key, "invoiced_quantity": "5"}]})
    payload = read_json_object(watch / "成本开票状态.json")

    assert "invoiced_reference_amount" in payload["items"][key]
    assert "invoiced_reference_tax" in payload["items"][key]
    assert "invoiced_reference_total_with_tax" in payload["items"][key]
    assert "updated_at" in payload


def test_cost_reference_status_keeps_old_locked_snapshot_when_summary_grows(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    key = _reference_key("材料", "项目", "规格", "吨")
    (watch / "成本开票状态.json").write_text(
        """{"version": 1, "items": {"%s": {"reference_key": "%s", "invoiced_quantity": "5", "invoiced_reference_total_with_tax": "61.02"}}}"""
        % (key, key),
        encoding="utf-8",
    )
    row_b = _cost_detail_row(源文件="b.pdf")
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row, row_b])

    reference = CostProjectionService(watch, workspace, "target").snapshot().invoice_reference[0]

    assert reference["quantity"] == 20.0
    assert reference["invoiced_quantity"] == 5.0
    assert reference["invoiced_reference_total_with_tax"] == 61.02
    assert reference["uninvoiced_quantity"] == 15.0


def test_cost_reference_status_falls_back_to_existing_workbook_sheet(tmp_path: Path) -> None:
    from openpyxl import Workbook

    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    row = _cost_detail_row()
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [row])
    wb = Workbook()
    ws = wb.active
    ws.title = "开票参考"
    ws.append(INVOICE_REFERENCE_HEADERS)
    ws.append(["材料", "项目", "规格", "吨", 10, 5, 5, "部分开具", 10.8, 12.204, 108, 14.04, 122.04, 61.02, 54, 7.02, 61.02, "20%", "2026-05-21T07:37:45Z"])
    wb.save(watch / "成本发票汇总.xlsx")

    reference = CostProjectionService(watch, workspace, "target").snapshot().invoice_reference[0]

    assert reference["key"] == _reference_key("材料", "项目", "规格", "吨")
    assert reference["markup_rate"] == "20%"
    assert reference["reference_total_with_tax"] == 135.6
    assert reference["invoiced_quantity"] == 5.0
    assert reference["invoiced_reference_total_with_tax"] == 61.02
    assert reference["status_updated_at"] == "2026-05-21T07:37:45Z"


def test_cost_rebuild_restores_status_json_from_existing_workbook(monkeypatch, tmp_path: Path) -> None:
    from openpyxl import Workbook
    import invoice_hub.projections.cost_analysis as cost_analysis

    watch = tmp_path / "发票文件"
    watch.mkdir()
    row = _cost_detail_row(源文件=str(watch / "a.pdf"))
    (watch / "a.pdf").write_bytes(b"%PDF")
    wb = Workbook()
    ws = wb.active
    ws.title = "开票参考"
    ws.append(INVOICE_REFERENCE_HEADERS)
    ws.append(["材料", "项目", "规格", "吨", 10, 5, 5, "部分开具", 10.8, 12.204, 108, 14.04, 122.04, 61.02, 54, 7.02, 61.02, "8%", "2026-05-21T07:37:45Z"])
    wb.save(watch / "成本发票汇总.xlsx")
    monkeypatch.setattr(cost_analysis, "analyze_pdf_costs", lambda _path, metadata=None: {"source": "a.pdf", "rows": [row], "invoice": {}, "status": "ok", "message": ""})

    build_cost_analysis_outputs(watch, watch)
    payload = read_json_object(watch / "成本开票状态.json")
    key = _reference_key("材料", "项目", "规格", "吨")

    assert payload["items"][key]["invoiced_quantity"] == "5"
    assert payload["items"][key]["reference_markup_rate_percent"] == "8"
    assert payload["items"][key]["invoiced_reference_total_with_tax"] == "61.02"
    assert payload["items"][key]["updated_at"] == "2026-05-21T07:37:45Z"


def test_cost_sync_matches_summary_absolute_paths_to_detail_file_names(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    workspace = tmp_path / "workspace"
    watch.mkdir()
    workspace.mkdir()
    source = watch / "a.pdf"
    source.write_text("pdf", encoding="utf-8")
    write_summary_rows(
        workspace / "发票汇总.csv",
        ["文件名", "文件路径", "发票类型", "发票号码", "开票时间", "销售方", "购买方", "开票金额", "税率", "除税价", "税金", "重复发票", "手改状态"],
        [{"文件名": source.name, "文件路径": str(source), "发票类型": "pdf", "发票号码": "1"}],
    )
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [{"源文件": source.name, "发票号码": "1", "内部项目名称": "项目", "规格型号": "规格", "单位": "吨", "数量": "1", "金额(除税)": "1", "税金": "0.13", "价税合计": "1.13"}],
    )
    service = CostProjectionService(watch, workspace, "target")

    sync = service.snapshot().sync

    assert sync.source_invoice_count == 1
    assert sync.parsed_invoice_count == 1
    assert sync.pending_count == 0
    assert sync.sync_state == "fresh"
