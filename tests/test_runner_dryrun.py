import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from invoice_hub.bookkeeping.batches import (
    begin_import_batch,
    finalize_import_batch,
    prepare_import_batch_files,
    record_batch_dry_run,
    register_export_batch,
)
from invoice_hub.bookkeeping.import_file import load_jierui_import_facts
from invoice_hub.bookkeeping.paths import company_bookkeeping_paths
from invoice_hub.bookkeeping.status import load_voucher_status, posting_key, proposal_revision_hash, write_voucher_status
from invoice_hub.domain.models import VoucherDraft, VoucherLine, VoucherStatusItem
from invoice_hub.runners import jierui_voucher_import as packaged_runner
from scripts.tools import jierui_voucher_import as runner


COMPANY_ID = "runner-company-id"


def _facts() -> dict:
    return load_jierui_import_facts(Path(__file__).resolve().parents[1] / "docs" / "jierui" / "voucher-import-template.facts.json")


def _draft(invoice: str, *, date: str = "2026-07-06", amount: str = "100.00", tax: str = "13.00", total: str = "113.00") -> VoucherDraft:
    key = posting_key(COMPANY_ID, "purchase_recognition", invoice)
    draft = VoucherDraft(
        voucher_key=key,
        posting_key=key,
        company_id=COMPANY_ID,
        event_type="purchase_recognition",
        anchor_business_key=invoice,
        voucher_date=date,
        period=date[:7],
        lines=[
            VoucherLine(summary=f"采购 {invoice}", account_code="1405", account_name="库存商品", direction="debit", amount=amount),
            VoucherLine(summary=f"采购 {invoice}", account_code="222101", account_name="进项税额", direction="debit", amount=tax),
            VoucherLine(summary=f"采购 {invoice}", account_code="2202", account_name="应付账款", direction="credit", amount=total),
        ],
        source_invoice_nos=[invoice],
        source_rows=[f"row:{invoice}"],
        source_file_hashes={f"{invoice}.pdf": "fixture"},
        balance_ok=True,
        review_tier="auto",
        generated_at="2026-07-06T00:00:00+08:00",
        business_class="inventory_purchase",
        tax_treatment="deductible",
        tax_evidence_refs=["test"],
    )
    return draft.model_copy(update={"proposal_revision_hash": proposal_revision_hash(draft)})


def _fixture(tmp_path: Path) -> tuple[Path, object, object]:
    company = tmp_path / "公司 A"
    paths = company_bookkeeping_paths(company)
    paths.batch_dir.mkdir(parents=True)
    drafts = [_draft("invoice-1"), _draft("invoice-2", amount="200.00", tax="26.00", total="226.00")]
    items = {
        draft.posting_key: VoucherStatusItem(
            status="approved",
            snapshot=draft.model_dump(mode="json"),
            approved_revision_hash=draft.proposal_revision_hash,
            audit=[],
        )
        for draft in drafts
    }
    store = write_voucher_status(
        paths.voucher_status_json,
        items,
        company_id=COMPANY_ID,
        ledger_environment="test",
        ledger_identity_sha256="a" * 64,
        ledger_profile_sha256="b" * 64,
    )
    batch, _created = prepare_import_batch_files(
        paths,
        list(store.items.items()),
        company_id=COMPANY_ID,
        ledger_environment="test",
        ledger_identity_sha256="a" * 64,
        ledger_profile_sha256="b" * 64,
        ledger_name="测试账套",
        period="2026-07",
        facts=_facts(),
        account_table={"1405": "库存商品", "222101": "进项税额", "2202": "应付账款"},
        account_table_sha256="account-hash",
        aux_catalog_sha256="",
    )
    register_export_batch(paths.voucher_status_json, batch, expected_revision=store.revision)
    return company, paths, batch


def test_runner_dry_run_uses_immutable_manifest_and_planned_numbers(tmp_path: Path) -> None:
    company, _paths, batch = _fixture(tmp_path)

    summary = runner.dry_run(company, batch_manifest=batch.manifest_path)

    assert summary["ok"] is True
    assert summary["batch_id"] == batch.batch_id
    assert summary["file_sha256"] == batch.file_sha256
    assert summary["voucher_count"] == 2
    assert summary["expected_exported_count"] == 2
    assert summary["debit_total"] == "339.00"
    assert summary["credit_total"] == "339.00"
    assert "测试账套" in summary["target_hint"]


def test_packaged_runner_module_loads_the_shipped_script() -> None:
    loaded = packaged_runner._load_runner()

    assert loaded.build_parser().description == "捷锐凭证批次 runner（W8 仅开放 dry-run）"


def test_runner_rejects_modified_xlsx_before_content_comparison(tmp_path: Path) -> None:
    company, _paths, batch = _fixture(tmp_path)
    workbook = load_workbook(batch.file_path)
    worksheet = workbook.active
    worksheet.cell(row=2, column=7).value = "999.00"
    workbook.save(batch.file_path)

    with pytest.raises(runner.DryRunError) as excinfo:
        runner.dry_run(company, batch_manifest=batch.manifest_path)

    assert excinfo.value.summary["errors"]["file_hash"][0]["reason"] == "batch_file_changed"


def test_runner_refuses_to_guess_latest_import_file(tmp_path: Path) -> None:
    company = tmp_path / "公司 A"
    company_bookkeeping_paths(company).import_dir.mkdir(parents=True)

    with pytest.raises(ValueError, match="必须显式提供 --batch-manifest"):
        runner.dry_run(company)


def test_unknown_finalize_is_idempotent_and_only_reconcile_can_advance(tmp_path: Path) -> None:
    _company, paths, batch = _fixture(tmp_path)
    record_batch_dry_run(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "ok": True,
            "file_sha256": batch.file_sha256,
            "voucher_count": 2,
            "debit_total": batch.expected_debit_total,
            "credit_total": batch.expected_credit_total,
        },
    )
    begin_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "batch_id": batch.batch_id,
            "file_sha256": batch.file_sha256,
            "ledger_name": batch.ledger_name,
            "period": batch.period,
            "authorized_by": "tester",
            "command_id": "apply-once",
        },
    )
    unknown_payload = {"mode": "apply", "outcome": "unknown", "items": [], "evidence": {"reason": "connection_lost"}}

    unknown, first_idempotent, first_receipt = finalize_import_batch(paths.voucher_status_json, batch.batch_id, unknown_payload)
    revision_after_first = load_voucher_status(paths.voucher_status_json).revision
    replay, second_idempotent, second_receipt = finalize_import_batch(paths.voucher_status_json, batch.batch_id, unknown_payload)

    assert unknown.state == replay.state == "unknown"
    assert first_idempotent is False
    assert second_idempotent is True
    assert first_receipt == second_receipt
    assert load_voucher_status(paths.voucher_status_json).revision == revision_after_first
    assert {item.status for item in load_voucher_status(paths.voucher_status_json).items.values()} == {"import_unknown"}
    with pytest.raises(ValueError, match="不能开始导入"):
        begin_import_batch(paths.voucher_status_json, batch.batch_id, {})
    with pytest.raises(ValueError, match="只允许 reconcile_only"):
        finalize_import_batch(paths.voucher_status_json, batch.batch_id, {"mode": "apply", "outcome": "confirmed_success", "items": []})

    contradictory_items = [
        {
            "posting_key": item.posting_key,
            "observed_state": "import_failed_confirmed",
        }
        for item in batch.items
    ]
    with pytest.raises(ValueError, match="observed_state 必须为 imported"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "confirmed_success",
                "items": contradictory_items,
                "evidence": {"readback_hash": "absence-readback", "ledger_absence_confirmed": True},
            },
        )

    success_items = [
        {
            "posting_key": item.posting_key,
            "observed_state": "imported",
            "voucher_no": item.planned_voucher_no,
            "signature_hash": item.signature_hash,
        }
        for item in batch.items
    ]
    with pytest.raises(ValueError, match="逐项提交全部批次凭证"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "confirmed_success",
                "items": success_items[:-1],
                "evidence": {"readback_hash": "matched-readback"},
            },
        )
    with pytest.raises(ValueError, match="必须携带 readback_hash"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {"mode": "reconcile_only", "outcome": "confirmed_success", "items": success_items},
        )
    for contradictory_evidence in (
        {"readback_hash": "matched-readback", "commit_not_attempted": True},
        {"readback_hash": "matched-readback", "ledger_absence_confirmed": True},
    ):
        with pytest.raises(ValueError, match="不能携带未提交或账套未落账证据"):
            finalize_import_batch(
                paths.voucher_status_json,
                batch.batch_id,
                {
                    "mode": "reconcile_only",
                    "outcome": "confirmed_success",
                    "items": success_items,
                    "evidence": contradictory_evidence,
                },
            )
    invalid_signature_items = [dict(item) for item in success_items]
    invalid_signature_items[0]["signature_hash"] = "wrong-signature"
    with pytest.raises(ValueError, match="回读 signature 不匹配"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "confirmed_success",
                "items": invalid_signature_items,
                "evidence": {"readback_hash": "matched-readback"},
            },
        )
    missing_voucher_no_items = [dict(item) for item in success_items]
    missing_voucher_no_items[0]["voucher_no"] = ""
    with pytest.raises(ValueError, match="回读凭证号缺失"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "confirmed_success",
                "items": missing_voucher_no_items,
                "evidence": {"readback_hash": "matched-readback"},
            },
        )
    reconciled, _, success_receipt = finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "mode": "reconcile_only",
            "outcome": "confirmed_success",
            "items": success_items,
            "evidence": {"readback_hash": "matched-readback"},
        },
    )
    assert reconciled.state == "reconciled"
    assert success_receipt["state"] == "reconciled"
    assert {item.status for item in load_voucher_status(paths.voucher_status_json).items.values()} == {"imported"}

    current, historical_replay, historical_receipt = finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        unknown_payload,
    )
    assert historical_replay is True
    assert current.state == "reconciled"
    assert historical_receipt == first_receipt


def test_failed_before_commit_requires_precommit_or_complete_absence_evidence(tmp_path: Path) -> None:
    _company, paths, batch = _fixture(tmp_path)
    record_batch_dry_run(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "ok": True,
            "file_sha256": batch.file_sha256,
            "voucher_count": batch.expected_count,
            "debit_total": batch.expected_debit_total,
            "credit_total": batch.expected_credit_total,
        },
    )
    begin_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "batch_id": batch.batch_id,
            "file_sha256": batch.file_sha256,
            "ledger_name": batch.ledger_name,
            "period": batch.period,
            "authorized_by": "tester",
            "command_id": "apply-unknown",
        },
    )
    finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {"mode": "apply", "outcome": "unknown", "items": [], "evidence": {"reason": "connection_lost"}},
    )

    with pytest.raises(ValueError, match="完整账套回读"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {"mode": "reconcile_only", "outcome": "failed_before_commit", "items": [], "evidence": {}},
        )

    absent_items = [
        {"posting_key": item.posting_key, "observed_state": "import_failed_confirmed"}
        for item in batch.items
    ]
    with pytest.raises(ValueError, match="ledger_absence_confirmed"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "failed_before_commit",
                "items": absent_items,
                "evidence": {"readback_hash": "absence-readback"},
            },
        )

    failed, idempotent, receipt = finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "mode": "reconcile_only",
            "outcome": "failed_before_commit",
            "items": absent_items,
            "evidence": {"readback_hash": "absence-readback", "ledger_absence_confirmed": True},
        },
    )

    assert idempotent is False
    assert failed.state == receipt["state"] == "failed_before_commit"
    assert {item.status for item in load_voucher_status(paths.voucher_status_json).items.values()} == {"import_failed_confirmed"}


def test_precommit_failure_rejects_observations_and_accepts_explicit_no_commit(tmp_path: Path) -> None:
    _company, paths, batch = _fixture(tmp_path)
    record_batch_dry_run(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "ok": True,
            "file_sha256": batch.file_sha256,
            "voucher_count": batch.expected_count,
            "debit_total": batch.expected_debit_total,
            "credit_total": batch.expected_credit_total,
        },
    )
    begin_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "batch_id": batch.batch_id,
            "file_sha256": batch.file_sha256,
            "ledger_name": batch.ledger_name,
            "period": batch.period,
            "authorized_by": "tester",
            "command_id": "apply-precommit-failure",
        },
    )

    with pytest.raises(ValueError, match="未点击提交"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "apply",
                "outcome": "failed_before_commit",
                "items": [{"posting_key": batch.items[0].posting_key}],
                "evidence": {"commit_not_attempted": True},
            },
        )

    failed, _, _ = finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "mode": "apply",
            "outcome": "failed_before_commit",
            "items": [],
            "evidence": {"commit_not_attempted": True, "failure_stage": "file_selection"},
        },
    )

    assert failed.state == "failed_before_commit"
    assert {item.status for item in load_voucher_status(paths.voucher_status_json).items.values()} == {"import_failed_confirmed"}


def test_runner_rejects_manifest_drift_from_registered_batch(tmp_path: Path) -> None:
    company, _paths, batch = _fixture(tmp_path)
    manifest = Path(batch.manifest_path)
    payload = json.loads(manifest.read_text(encoding="utf-8"))
    payload["template_facts_sha256"] = "tampered"
    manifest.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")

    with pytest.raises(ValueError, match="immutable manifest"):
        runner.dry_run(company, batch_manifest=manifest)


def test_reconcile_cannot_downgrade_an_already_imported_item(tmp_path: Path) -> None:
    _company, paths, batch = _fixture(tmp_path)
    record_batch_dry_run(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "ok": True,
            "file_sha256": batch.file_sha256,
            "voucher_count": batch.expected_count,
            "debit_total": batch.expected_debit_total,
            "credit_total": batch.expected_credit_total,
        },
    )
    begin_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "batch_id": batch.batch_id,
            "file_sha256": batch.file_sha256,
            "ledger_name": batch.ledger_name,
            "period": batch.period,
            "authorized_by": "tester",
            "command_id": "apply-partial",
        },
    )
    imported = batch.items[0]
    partial, _, _ = finalize_import_batch(
        paths.voucher_status_json,
        batch.batch_id,
        {
            "mode": "apply",
            "outcome": "partial",
            "items": [
                {
                    "posting_key": imported.posting_key,
                    "observed_state": "imported",
                    "voucher_no": imported.planned_voucher_no,
                    "signature_hash": imported.signature_hash,
                }
            ],
            "evidence": {"readback_hash": "partial-readback"},
        },
    )
    assert partial.state == "partial"

    with pytest.raises(ValueError, match="终态不能"):
        finalize_import_batch(
            paths.voucher_status_json,
            batch.batch_id,
            {
                "mode": "reconcile_only",
                "outcome": "partial",
                "items": [{"posting_key": imported.posting_key, "observed_state": "import_failed_confirmed"}],
                "evidence": {"readback_hash": "later-readback", "ledger_absence_confirmed": True},
            },
        )
