import json
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from invoice_hub.api.app import create_app
from invoice_hub.projections.cost_analysis import DETAIL_HEADERS
from invoice_hub.projections.documents import rmb_uppercase
from invoice_hub.services.app_state import AppState
from invoice_hub.storage.files import read_json_object, write_csv_rows


def _cost_xml_text(invoice_number: str = "10000000000000000013") -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<EInvoice>
  <EInvoiceData>
    <SellerInformation><SellerName>开具销售方有限公司</SellerName></SellerInformation>
    <BuyerInformation><BuyerName>开具购买方有限公司</BuyerName></BuyerInformation>
    <BasicInformation>
      <InvoiceType>增值税专用发票</InvoiceType>
      <BusinessType>标准电子发票</BusinessType>
      <InvoiceNumber>{invoice_number}</InvoiceNumber>
      <IssueDate>2026-05-30</IssueDate>
      <TotalAmWithoutTax>300.00</TotalAmWithoutTax>
      <TotalTaxAm>39.00</TotalTaxAm>
      <TotalTaxIncludedAmount>339.00</TotalTaxIncludedAmount>
    </BasicInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>12E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>2</Quantity>
      <UnPrice>100</UnPrice>
      <Amount>200</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>26</ComTaxAm>
      <TotalTaxIncludedAmount>226</TotalTaxIncludedAmount>
    </IssuItemInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>12E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>1</Quantity>
      <UnPrice>120</UnPrice>
      <Amount>120</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>15.6</ComTaxAm>
      <TotalTaxIncludedAmount>135.6</TotalTaxIncludedAmount>
    </IssuItemInformation>
  </EInvoiceData>
</EInvoice>
"""


def _detail_row(**overrides: object) -> dict:
    row = {
        "销售方": "入库销售方有限公司",
        "购买方": "入库购买方有限公司",
        "发票号码": "10000000000000000013",
        "开票日期": "2026-05-30",
        "备注项目名称": "",
        "内部项目名称": "钢筋",
        "规格型号": "12E",
        "单位": "吨",
        "数量": "2",
        "单价(除税)": "100",
        "平均单价(含税)": "113",
        "金额(除税)": "200",
        "税率": "13%",
        "税金": "26",
        "价税合计": "226",
        "发票代码(**内文字)": "材料",
        "源文件": "a.xml",
    }
    row.update({key: str(value) for key, value in overrides.items()})
    return row


def _client(tmp_path: Path, monkeypatch) -> TestClient:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    return TestClient(create_app(tmp_path))


def test_documents_state_defaults_and_invoice_lists(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    outbound = tmp_path / "开具发票"
    outbound.mkdir()
    (outbound / "dzfp_10000000000000000013.xml").write_text(_cost_xml_text(), encoding="utf-8")
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [_detail_row(), _detail_row(**{"单价(除税)": "120", "金额(除税)": "120", "税金": "15.6", "价税合计": "135.6"})],
    )
    assert client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(outbound)}).json()["ok"] is True

    state = client.get("/api/v1/documents/state").json()

    assert state["defaults"]["inbound"] == {"采购员": "", "负责人": "", "仓管员": "", "制表人": ""}
    assert state["defaults"]["outbound"]["收货单位"] == ""
    assert state["outbound_invoice_dir"] == str(outbound)
    assert state["recent_outbound_invoice_dirs"][0] == str(outbound)
    assert state["inbound_invoices"][0]["invoice_number"] == "10000000000000000013"
    assert state["inbound_invoices"][0]["row_count"] == 2
    assert state["outbound_invoices"][0]["invoice_number"] == "10000000000000000013"


def test_documents_outbound_dir_serializes_project_internal_path_and_removes_recent(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    first = tmp_path / "一号开具"
    second = tmp_path / "二号开具"
    first.mkdir()
    second.mkdir()

    assert client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(first)}).json()["ok"] is True
    assert client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(second)}).json()["ok"] is True
    removed = client.post("/api/v1/documents/recent-outbound-dirs/remove", json={"outbound_invoice_dir": str(first)}).json()

    assert removed["outbound_invoice_dir"] == str(second)
    assert str(first) not in removed["recent_outbound_invoice_dirs"]
    saved = json.loads((tmp_path / "config" / "app.local.json").read_text(encoding="utf-8-sig"))
    assert saved["outbound_invoice_dir"] == "./二号开具"
    assert "./一号开具" not in saved["recent_outbound_invoice_dirs"]


def test_documents_validate_outbound_dir_is_backend_source_of_truth(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    outbound = tmp_path / "待保存开具发票"
    outbound.mkdir()
    (outbound / "invoice.xml").write_text(_cost_xml_text(), encoding="utf-8")

    validation = client.post(
        "/api/v1/documents/validate-outbound-dir",
        json={"outbound_invoice_dir": str(outbound)},
    ).json()

    assert validation["ok"] is True
    assert validation["path"] == str(outbound)
    assert validation["can_use"] is True
    assert validation["supported_count"] == 1
    assert client.get("/api/v1/documents/state").json()["outbound_invoice_dir"] == ""


def test_document_defaults_allow_blank_values(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)

    result = client.put(
        "/api/v1/documents/defaults",
        json={"inbound": {"采购员": "张三", "负责人": ""}, "outbound": {"收货单位": "", "电话": "123"}},
    ).json()

    assert result["ok"] is True
    assert result["defaults"]["inbound"]["采购员"] == "张三"
    assert result["defaults"]["inbound"]["负责人"] == ""
    assert result["defaults"]["outbound"]["收货单位"] == ""
    assert result["defaults"]["outbound"]["电话"] == "123"
    defaults_path = tmp_path / "runtime" / "local_state" / "documents" / "defaults.json"
    assert read_json_object(defaults_path)["inbound"]["采购员"] == "张三"


def test_inbound_preview_filters_invoice_without_merging_same_spec_different_price(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [
            _detail_row(),
            _detail_row(**{"单价(除税)": "120", "金额(除税)": "120", "税金": "15.6", "价税合计": "135.6"}),
            _detail_row(发票号码="10000000000000000015", **{"金额(除税)": "999", "税金": "1"}),
        ],
    )

    preview = client.get("/api/v1/documents/inbound/preview", params={"invoice_number": "10000000000000000013"}).json()

    assert preview["row_count"] == 2
    assert [row["unit_price"] for row in preview["rows"]] == ["100.00", "120.00"]
    assert preview["total_with_tax"] == "361.60"
    assert preview["supplier"] == "入库销售方有限公司"


def test_rmb_uppercase_keeps_expected_zero_positions() -> None:
    assert rmb_uppercase("343190.14") == "人民币叁拾肆万叁仟壹佰玖拾元壹角肆分"
    assert rmb_uppercase("340190.14") == "人民币叁拾肆万零壹佰玖拾元壹角肆分"
    assert rmb_uppercase("1000101.01") == "人民币壹佰万零壹佰零壹元零壹分"


def test_inbound_export_summary_matches_preview_for_five_line_invoice(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    rows = [
        _detail_row(
            发票号码="10000000000000000014",
            开票日期="2026-06-02",
            规格型号="25E",
            数量="34.561",
            **{"单价(除税)": "2929.20", "金额(除税)": "101236.20", "税金": "13160.71", "价税合计": "114396.91"},
        ),
        _detail_row(
            发票号码="10000000000000000014",
            开票日期="2026-06-02",
            规格型号="18E",
            数量="10.32",
            **{"单价(除税)": "2884.96", "金额(除税)": "29772.74", "税金": "3870.46", "价税合计": "33643.20"},
        ),
        _detail_row(
            发票号码="10000000000000000014",
            开票日期="2026-06-02",
            规格型号="22E",
            数量="14.518",
            **{"单价(除税)": "2929.20", "金额(除税)": "42526.18", "税金": "5528.40", "价税合计": "48054.58"},
        ),
        _detail_row(
            发票号码="10000000000000000014",
            开票日期="2026-06-02",
            规格型号="28E",
            数量="9.855",
            **{"单价(除税)": "2938.05", "金额(除税)": "28954.51", "税金": "3764.09", "价税合计": "32718.60"},
        ),
        _detail_row(
            发票号码="10000000000000000014",
            开票日期="2026-06-02",
            规格型号="20E",
            数量="34.765",
            **{"单价(除税)": "2911.50", "金额(除税)": "101218.45", "税金": "13158.40", "价税合计": "114376.85"},
        ),
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)

    preview = client.get("/api/v1/documents/inbound/preview", params={"invoice_number": "10000000000000000014"}).json()
    exported = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000014"}).json()

    assert preview["total_with_tax"] == "343190.14"
    assert preview["total_with_tax_upper"] == "人民币叁拾肆万叁仟壹佰玖拾元壹角肆分"
    wb = load_workbook(exported["path"], data_only=False)
    try:
        ws = wb["入库单"]
        merges = [str(item) for item in ws.merged_cells.ranges]
        assert "B16:E16" in merges
        assert "G16:J16" in merges
        assert ws["A16"].value == "合计（大写）"
        assert ws["B16"].value == preview["total_with_tax_upper"]
        assert ws["F16"].value == "合计（小写）"
        assert float(ws["G16"].value) == 343190.14
        assert ws["C16"].value is None
        assert ws["H16"].value is None
        assert ws["G16"].alignment.horizontal == "left"
    finally:
        wb.close()


def test_outbound_preview_parses_selected_invoice_details(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    outbound = tmp_path / "开具发票"
    outbound.mkdir()
    (outbound / "dzfp_10000000000000000013.xml").write_text(_cost_xml_text(), encoding="utf-8")
    client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(outbound)})

    preview = client.get("/api/v1/documents/outbound/preview", params={"invoice_number": "10000000000000000013"}).json()

    assert preview["row_count"] == 2
    assert preview["rows"][0]["item_name"] == "钢筋"
    assert preview["rows"][0]["unit_price"] == "113.00"
    assert preview["rows"][1]["unit_price"] == "135.60"
    assert preview["total_with_tax"] == "361.60"
    assert preview["total_with_tax_upper"].startswith("人民币")


def test_inbound_export_writes_excel_and_open_requires_existing_file(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [_detail_row()])

    before_open = client.post("/api/v1/documents/inbound/open", json={"invoice_number": "10000000000000000013"}).json()
    exported = client.post(
        "/api/v1/documents/inbound/export",
        json={"invoice_number": "10000000000000000013", "defaults": {"采购员": "张三", "负责人": "", "仓管员": "李四", "制表人": ""}},
    ).json()
    after_open = client.post("/api/v1/documents/inbound/open", json={"invoice_number": "10000000000000000013"}).json()

    assert before_open["ok"] is False
    assert exported["ok"] is True
    path = Path(exported["path"])
    assert path == watch / "入库单" / "入库单-10000000000000000013-2026-05-30.xlsx"
    assert path.exists()
    assert after_open["ok"] is True
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb["入库单"]
        assert ws["B3"].value == "入库销售方有限公司"
        assert ws["E3"].value == "2026-05-30"
        assert ws["G3"].value == "NO：10000000000000000013"
        assert ws["B5"].value == "钢筋"
        assert float(ws["G5"].value) == 200.0
        assert float(ws["H5"].value) == 26.0
        assert "贰佰贰拾陆元整" in ws["B16"].value
        assert float(ws["G16"].value) == 226.0
        assert ws["G16"].alignment.horizontal == "left"
        assert ws["B17"].value == "张三"
        assert ws["G17"].value == "李四"
    finally:
        wb.close()


def test_inbound_export_status_copy_location_deleted_and_occupied_messages(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [_detail_row()])

    initial_status = client.post("/api/v1/documents/inbound/export-status", json={"invoice_number": "10000000000000000013"}).json()
    exported = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013"}).json()
    exported_path = Path(exported["path"])
    status = client.post("/api/v1/documents/inbound/export-status", json={"invoice_number": "10000000000000000013"}).json()
    copy_export = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013", "mode": "copy"}).json()
    location = client.post("/api/v1/documents/inbound/open-location", json={"invoice_number": "10000000000000000013"}).json()

    assert initial_status["exists"] is False
    assert initial_status["message"] == "单据尚未导出。"
    assert status["exists"] is True
    assert status["occupied"] is False
    assert status["path"] == str(exported_path)
    assert status["folder_path"] == str(exported_path.parent)
    assert copy_export["ok"] is True
    assert copy_export["copy"] is True
    assert Path(copy_export["path"]).name == "入库单-10000000000000000013-2026-05-30-副本1.xlsx"
    assert Path(copy_export["path"]).exists()
    assert location["ok"] is True
    assert location["folder_path"] == str(exported_path.parent)

    exported_path.unlink()
    deleted_status = client.post("/api/v1/documents/inbound/export-status", json={"invoice_number": "10000000000000000013"}).json()
    deleted_copy = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013", "mode": "copy"}).json()
    deleted_open = client.post("/api/v1/documents/inbound/open-location", json={"invoice_number": "10000000000000000013"}).json()
    assert deleted_status["exists"] is False
    assert deleted_status["message"] == "单据尚未导出。"
    assert deleted_copy["ok"] is False
    assert deleted_copy["exported"] is False
    assert "文件已经被删除或尚未导出" in deleted_copy["message"]
    assert deleted_open["ok"] is False
    assert "文件已经被删除或尚未导出" in deleted_open["message"]

    recreated = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013"}).json()
    occupied_path = Path(recreated["path"])
    monkeypatch.setattr(AppState, "_document_path_occupied", staticmethod(lambda path: path == occupied_path))
    occupied_status = client.post("/api/v1/documents/inbound/export-status", json={"invoice_number": "10000000000000000013"}).json()
    occupied_export = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013"}).json()
    occupied_open = client.post("/api/v1/documents/inbound/open", json={"invoice_number": "10000000000000000013"}).json()
    assert occupied_status["exists"] is True
    assert occupied_status["occupied"] is True
    assert occupied_status["message"] == "文件被占用，请关闭后再操作。"
    assert occupied_export["ok"] is False
    assert occupied_export["occupied"] is True
    assert occupied_export["message"] == "文件被占用，请关闭后再操作。"
    assert occupied_open["ok"] is False
    assert occupied_open["occupied"] is True
    assert occupied_open["message"] == "文件被占用，请关闭后再操作。"


def test_outbound_export_writes_excel_with_tax_amount_and_uppercase(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    outbound = tmp_path / "开具发票"
    outbound.mkdir()
    (outbound / "dzfp_10000000000000000013.xml").write_text(_cost_xml_text(), encoding="utf-8")
    client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(outbound)})

    exported = client.post(
        "/api/v1/documents/outbound/export",
        json={
            "invoice_number": "10000000000000000013",
            "defaults": {"收货单位": "项目部", "地址": "南通", "电话": "123", "联系人": "王五", "编辑人": "编", "收货人": "收", "项目负责人": "负"},
        },
    ).json()

    path = Path(exported["path"])
    assert path == outbound / "出库单" / "出库单-10000000000000000013-2026-05-30.xlsx"
    wb = load_workbook(path, data_only=False)
    try:
        ws = wb["出库单"]
        assert ws["A3"].value == "收货单位：项目部"
        assert ws["C3"].value == "开单日期：2026-05-30"
        assert ws["F3"].value == "单据编号：10000000000000000013"
        assert ws["B6"].value == "钢筋"
        assert float(ws["F6"].value) == 113.0
        assert float(ws["G6"].value) == 226.0
        assert "叁佰陆拾壹元陆角" in ws["B16"].value
        assert float(ws["F16"].value) == 361.6
        assert ws["F16"].alignment.horizontal == "left"
        assert ws["B18"].value == "编"
        assert ws["E18"].value == "收"
        assert ws["H18"].value == "负"
    finally:
        wb.close()


def test_outbound_export_status_copy_and_location_are_scoped_to_outbound_dir(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    client = _client(tmp_path, monkeypatch)
    outbound = tmp_path / "开具发票"
    outbound.mkdir()
    (outbound / "dzfp_10000000000000000013.xml").write_text(_cost_xml_text(), encoding="utf-8")
    client.put("/api/v1/documents/outbound-dir", json={"outbound_invoice_dir": str(outbound)})

    exported = client.post("/api/v1/documents/outbound/export", json={"invoice_number": "10000000000000000013"}).json()
    status = client.post("/api/v1/documents/outbound/export-status", json={"invoice_number": "10000000000000000013"}).json()
    copy_export = client.post("/api/v1/documents/outbound/export", json={"invoice_number": "10000000000000000013", "mode": "copy"}).json()
    location = client.post("/api/v1/documents/outbound/open-location", json={"invoice_number": "10000000000000000013"}).json()

    exported_path = Path(exported["path"])
    assert exported_path == outbound / "出库单" / "出库单-10000000000000000013-2026-05-30.xlsx"
    assert status["exists"] is True
    assert status["folder_path"] == str(outbound / "出库单")
    assert Path(copy_export["path"]).name == "出库单-10000000000000000013-2026-05-30-副本1.xlsx"
    assert location["ok"] is True
    assert location["folder_path"] == str(outbound / "出库单")


def test_inbound_export_inserts_extra_rows_when_detail_exceeds_template(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    rows = [_detail_row(源文件=f"{index}.xml", **{"金额(除税)": "10", "税金": "1.3"}) for index in range(12)]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)

    exported = client.post("/api/v1/documents/inbound/export", json={"invoice_number": "10000000000000000013"}).json()

    wb = load_workbook(exported["path"], data_only=False)
    try:
        ws = wb["入库单"]
        assert ws.max_row == 18
        assert ws["B16"].value == "钢筋"
        assert ws["C16"].value == "12E"
        assert ws["A17"].value == "合计（大写）"
        merges = [str(item) for item in ws.merged_cells.ranges]
        assert "B17:E17" in merges
        assert "G17:J17" in merges
        assert "壹佰叁拾伍元陆角" in ws["B17"].value
        assert float(ws["G17"].value) == 135.6
        assert ws["A18"].value == "采购员"
        assert ws["B16"].border.left.style == ws["B15"].border.left.style
    finally:
        wb.close()


def test_open_document_api_does_not_accept_arbitrary_path(tmp_path: Path, monkeypatch) -> None:
    client = _client(tmp_path, monkeypatch)
    watch = tmp_path / "发票文件"
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, [_detail_row()])
    outside = tmp_path / "outside.xlsx"
    outside.write_text("not a workbook", encoding="utf-8")

    response = client.post(
        "/api/v1/documents/inbound/open",
        json={"invoice_number": "10000000000000000013", "path": str(outside)},
    ).json()

    assert response["ok"] is False
    assert str(outside) not in response.get("path", "")
