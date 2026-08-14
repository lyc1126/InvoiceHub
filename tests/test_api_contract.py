import importlib
import json
import multiprocessing
import os
import subprocess
import sys
import threading
import time
from io import BytesIO
from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient

from invoice_hub.api.app import _resolve_event_stream_cursor, create_app
from invoice_hub.monitoring.state import MonitorState
from invoice_hub.projections.cost_analysis import DETAIL_HEADERS
from invoice_hub.projections.summary import SUMMARY_HEADERS
from invoice_hub.release.build_manifest import API_CAPABILITIES, API_CONTRACT_VERSION, BOOKKEEPING_PROTOCOL_VERSION, write_build_manifest
from invoice_hub.services.app_state import AppState
from invoice_hub.services.skins import MAX_SKIN_FILE_BYTES, MAX_SKIN_FILES
from invoice_hub.storage.files import write_csv_rows
from invoice_hub.targets import target_profile_for


def _slow_background_sync_worker(
    profile_payload: dict[str, str],
    db_path: str,
    reference_markup_rate: str,
    trigger: str,
    result_sender,
) -> None:
    del db_path, reference_markup_rate
    try:
        time.sleep(1.0)
        result_sender.send(
            {
                "ok": True,
                "sync": {
                    "ok": True,
                    "trigger": trigger,
                    "target_id": profile_payload["id"],
                    "added": 0,
                    "updated": 0,
                    "deleted": 0,
                    "manual_changed": 0,
                    "rebuilt": False,
                },
            }
        )
    finally:
        result_sender.close()


def test_cli_creates_exactly_one_app_state(tmp_path: Path) -> None:
    script = r"""
import json
import sys

from invoice_hub.services.app_state import AppState

startup_calls = []
AppState.run_background_diagnostics = lambda self, trigger="startup_sync": startup_calls.append(trigger)

import uvicorn

def fake_run(app, **kwargs):
    print(json.dumps({
        "startup_calls": startup_calls,
        "root_dir": str(app.state.invoice_hub.config.root_dir),
        "config_path": str(app.state.invoice_hub.config.config_path),
    }))

uvicorn.run = fake_run
sys.argv = ["invoice-hub", "--root", sys.argv[1], "--config", sys.argv[2]]
from invoice_hub.api.main import main
raise SystemExit(main())
"""
    config_path = tmp_path / "config" / "app.local.json"
    (tmp_path / "web" / "templates").mkdir(parents=True)
    (tmp_path / "web" / "static").mkdir(parents=True)
    env = os.environ.copy()
    env["PYTHONPATH"] = str(Path(__file__).resolve().parents[1] / "src")
    result = subprocess.run(
        [sys.executable, "-c", script, str(tmp_path), str(config_path)],
        check=True,
        capture_output=True,
        text=True,
        env=env,
    )

    payload = json.loads(result.stdout.strip().splitlines()[-1])
    assert payload["startup_calls"] == ["startup_sync"]
    assert payload["root_dir"] == str(tmp_path.resolve())
    assert payload["config_path"] == str(config_path.resolve())


def test_cli_rejects_abbreviated_long_options(tmp_path: Path) -> None:
    script = r"""
import sys

from invoice_hub.services.app_state import AppState

AppState.run_background_diagnostics = lambda self, trigger="startup_sync": None

import uvicorn

uvicorn.run = lambda app, **kwargs: None

from invoice_hub.api.main import main
raise SystemExit(main())
"""
    config_path = tmp_path / "config" / "app.local.json"
    (tmp_path / "web" / "templates").mkdir(parents=True)
    (tmp_path / "web" / "static").mkdir(parents=True)
    env = os.environ.copy()
    env["PYTHONPATH"] = str(Path(__file__).resolve().parents[1] / "src")
    abbreviated_arguments = [
        ["--roo", str(tmp_path)],
        ["--root", str(tmp_path), "--conf", str(config_path)],
        ["--root", str(tmp_path), f"--confi={config_path}"],
        ["--root", str(tmp_path), "--hos", "127.0.0.1"],
        ["--root", str(tmp_path), "--por", "8766"],
    ]

    for arguments in abbreviated_arguments:
        result = subprocess.run(
            [sys.executable, "-c", script, *arguments],
            check=False,
            capture_output=True,
            text=True,
            env=env,
        )

        assert result.returncode == 2, arguments
        assert "unrecognized arguments" in result.stderr


def _make_skin_zip(skin_id: str = "custom-skin", files: dict[str, bytes | str] | None = None, manifest: dict | None = None) -> bytes:
    payload = BytesIO()
    data = {
        "id": skin_id,
        "name": "Custom Skin",
        "version": "1.0.0",
        "entry": "skin.css",
        "entrypoint": "skin.css",
    }
    data.update(manifest or {})
    with ZipFile(payload, "w") as archive:
        archive.writestr("skin.json", json.dumps(data, ensure_ascii=False))
        for name, content in (files or {"skin.css": ":root { --color-accent: #2563eb; }\n"}).items():
            archive.writestr(name, content if isinstance(content, bytes) else content.encode("utf-8"))
    return payload.getvalue()


def _cost_xml_text(invoice_number: str = "10000000000000000013") -> str:
    return f"""<?xml version="1.0" encoding="UTF-8"?>
<EInvoice>
  <EInvoiceData>
    <SellerInformation><SellerName>示例销售方有限公司</SellerName></SellerInformation>
    <BuyerInformation><BuyerName>示例购买方有限公司</BuyerName></BuyerInformation>
    <BasicInformation>
      <InvoiceType>增值税专用发票</InvoiceType>
      <BusinessType>标准电子发票</BusinessType>
      <InvoiceNumber>{invoice_number}</InvoiceNumber>
      <IssueDate>2026-05-30</IssueDate>
      <TotalAmWithoutTax>300.00</TotalAmWithoutTax>
      <TotalTaxAm>39.00</TotalTaxAm>
      <TotalTaxIncludedAmount>339.00</TotalTaxIncludedAmount>
    </BasicInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>12E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>2</Quantity>
      <UnPrice>100</UnPrice>
      <Amount>200</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>26</ComTaxAm>
      <TotalTaxIncludedAmount>226</TotalTaxIncludedAmount>
    </IssuItemInformation>
    <IssuItemInformation>
      <ItemName>*材料*钢筋</ItemName>
      <SpecMod>14E</SpecMod>
      <MeaUnits>吨</MeaUnits>
      <Quantity>1</Quantity>
      <UnPrice>100</UnPrice>
      <Amount>100</Amount>
      <TaxRate>0.13</TaxRate>
      <ComTaxAm>13</ComTaxAm>
      <TotalTaxIncludedAmount>113</TotalTaxIncludedAmount>
    </IssuItemInformation>
  </EInvoiceData>
</EInvoice>
"""


def _summary_invoice_row(source: Path, invoice_number: str = "10000000000000000001") -> dict[str, str]:
    return {
        "文件名": source.name,
        "文件路径": str(source),
        "发票类型": "增值税专用发票",
        "特定业务类型": "标准电子发票",
        "类型识别状态": "ok",
        "类型识别说明": "",
        "发票号码": invoice_number,
        "开票时间": "2026-07-06",
        "销售方": "示例销售方",
        "购买方": "示例购买方",
        "开票金额": "1175.20",
        "税率": "13%",
        "除税价": "1040.00",
        "税金": "135.20",
        "重复发票": "",
        "手改状态": "",
    }


def _cost_detail_row(
    invoice_number: str,
    source: Path | str,
    project: str,
    spec: str,
    quantity: str,
    amount: str,
    tax: str,
    total: str,
    unit: str = "吨",
    tax_rate: str = "13%",
) -> dict[str, str]:
    return {
        "销售方": "示例销售方",
        "购买方": "示例购买方",
        "发票号码": invoice_number,
        "开票日期": "2026-07-06",
        "备注项目名称": project,
        "内部项目名称": project,
        "规格型号": spec,
        "单位": unit,
        "数量": quantity,
        "单价(除税)": "",
        "平均单价(含税)": "",
        "金额(除税)": amount,
        "税率": tax_rate,
        "税金": tax,
        "价税合计": total,
        "发票代码(**内文字)": f"*材料*{project}",
        "源文件": str(source),
    }


def test_api_contract_contains_cost_paths(tmp_path: Path) -> None:
    app = create_app(tmp_path)
    client = TestClient(app)

    health = client.get("/api/v1/health").json()
    assert health["ok"] is True
    assert health["pid"] > 0
    assert health["config_path"] == str((tmp_path / "config" / "app.local.json").resolve())
    assert health["runtime_dir"] == str((tmp_path / "runtime").resolve())
    assert health["build_id"] == "development"
    assert health["api_contract_version"] == API_CONTRACT_VERSION
    assert health["bookkeeping_protocol_version"] == BOOKKEEPING_PROTOCOL_VERSION
    assert health["build_manifest_present"] is False
    assert "documents" in health["capabilities"]
    assert "documents.validate-outbound-dir" in health["capabilities"]
    assert "bookkeeping.review" in health["capabilities"]
    assert "bookkeeping.executability.v2" in health["capabilities"]
    assert "bookkeeping.import-batch.v1" in health["capabilities"]
    assert "bookkeeping.import-finalize.v1" in health["capabilities"]
    assert "bookkeeping.jierui.facts.v2" in health["capabilities"]
    assert "bookkeeping.jierui.runner.dry-run.v2" in health["capabilities"]
    assert "bookkeeping.state-cas.v1" in health["capabilities"]
    assert "bookkeeping.w9-ledger-review.v1" in health["capabilities"]
    assert "bookkeeping.mapping-resolution.v1" in health["capabilities"]
    assert "bookkeeping.targeted-recompute.v1" in health["capabilities"]
    assert "bookkeeping.migration-cas.v2" in health["capabilities"]
    assert "costs.internal-scroll" in health["capabilities"]
    assert "settings.center.v1" in health["capabilities"]
    assert "settings.preferences.v1" in health["capabilities"]
    assert "diagnostics.support-package.v1" in health["capabilities"]
    assert "invoices.classification.v1" in health["capabilities"]
    assert "invoices.file-preview.v1" in health["capabilities"]
    assert "invoices.batch-print.v1" in health["capabilities"]
    assert "invoices.rename-safe.v1" in health["capabilities"]
    assert "invoices.selection-summary.v1" in health["capabilities"]
    assert "monitor.ready-handshake.v1" in health["capabilities"]
    assert "server.shutdown-choice.v1" in health["capabilities"]
    assert "skins.zip-portable" in health["capabilities"]
    favicon = client.get("/favicon.ico")
    assert favicon.status_code == 204
    assert favicon.content == b""

    settings = client.get("/api/v1/settings").json()
    assert settings["ok"] is True
    assert "active_target_paths" in settings
    assert settings["bridge"]["sync_interval_seconds"] == 60
    assert "log_path" in settings["bridge"]
    assert "lock_path" in settings["bridge"]
    assert "stop_file_path" in settings["bridge"]
    assert "summary_xlsx_path" in settings["active_summary"]
    assert "summary_xlsx_exists" in settings["active_summary"]
    assert "output_detail_csv_exists" in settings["active_cost_analysis"]
    assert "output_summary_xlsx_exists" in settings["active_cost_analysis"]
    assert "reference_status_exists" in settings["active_cost_analysis"]
    assert "diagnostics" in settings
    assert "runtime_dir" in settings["diagnostics"]
    assert "db_path" in settings["diagnostics"]
    assert "server_state_path" in settings["diagnostics"]

    cost = client.get("/api/v1/cost-analysis").json()
    for key in (
        "watch_dir",
        "source_dir",
        "target_id",
        "output_detail_csv_path",
        "output_summary_xlsx_path",
        "reference_status_path",
        "invoice_reference",
        "reference_status_stats",
        "reference_markup_rate",
        "sync",
        "recent_watch_dirs",
    ):
        assert key in cost
    assert cost["output_detail_csv_path"].endswith("成本发票明细.csv")
    assert cost["output_summary_xlsx_path"].endswith("成本发票汇总.xlsx")
    for key in (
        "inventory_total_with_tax",
        "invoiced_reference_total_with_tax",
        "uninvoiced_reference_total_with_tax",
    ):
        assert key in cost["reference_status_stats"]


def test_preferences_api_defaults_and_persistence(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    client = TestClient(app)

    response = client.get("/api/v1/preferences")
    assert response.status_code == 200
    payload = response.json()
    assert payload["ok"] is True
    assert payload["preferences"] == {
        "cost_row_limit": 30,
        "long_path_display": "truncate-hover-scroll",
        "document_export_existing_strategy": "prompt",
        "system_shutdown_behavior": "ask",
        "ocr_candidate_dir": "",
        "auto_check_updates": True,
        "startup_surface": "desktop" if sys.platform == "darwin" else "browser",
    }
    assert payload["allowed"]["cost_row_limit"] == [30, 60, 100]
    assert payload["allowed"]["long_path_display"] == ["truncate-hover-scroll", "wrap"]
    assert payload["allowed"]["document_export_existing_strategy"] == ["copy", "open", "prompt"]
    assert payload["allowed"]["system_shutdown_behavior"] == ["ask", "keep_monitor", "stop_monitor"]
    assert payload["allowed"]["startup_surface"] == ["browser", "desktop"]
    assert payload["allowed"]["desktop_available"] is (sys.platform == "darwin")

    preferences_path = Path(payload["preferences_path"]).resolve()
    assert preferences_path == (tmp_path / "runtime" / "local_state" / "preferences.json").resolve()
    settings = client.get("/api/v1/settings").json()
    assert settings["preferences"] == payload["preferences"]
    assert settings["preferences_path"] == str(preferences_path)
    assert settings["diagnostics"]["preferences_path"] == str(preferences_path)

    update = client.put(
        "/api/v1/preferences",
        json={
            "cost_row_limit": 60,
            "long_path_display": "wrap",
            "document_export_existing_strategy": "copy",
            "system_shutdown_behavior": "stop_monitor",
            "ocr_candidate_dir": "ocr候选",
            "auto_check_updates": False,
            "startup_surface": "browser",
        },
    )
    assert update.status_code == 200
    preferences = update.json()["preferences"]
    assert preferences["cost_row_limit"] == 60
    assert preferences["long_path_display"] == "wrap"
    assert preferences["document_export_existing_strategy"] == "copy"
    assert preferences["system_shutdown_behavior"] == "stop_monitor"
    assert preferences["ocr_candidate_dir"] == str((tmp_path / "ocr候选").resolve())
    assert preferences["auto_check_updates"] is False
    assert preferences["startup_surface"] == "browser"
    assert json.loads(preferences_path.read_text(encoding="utf-8")) == preferences

    rejected = client.put(
        "/api/v1/preferences",
        headers={"Origin": "https://attacker.example"},
        json={"auto_check_updates": True},
    )
    assert rejected.status_code == 403
    assert client.get("/api/v1/preferences").json()["preferences"]["auto_check_updates"] is False

    refreshed_settings = client.get("/api/v1/settings").json()
    assert refreshed_settings["preferences"] == preferences

    invalid = client.put("/api/v1/preferences", json={"cost_row_limit": 45})
    assert invalid.status_code == 400
    assert "成本页显示行数" in invalid.json()["detail"]

    invalid_shutdown = client.put("/api/v1/preferences", json={"system_shutdown_behavior": "close_everything"})
    assert invalid_shutdown.status_code == 400
    assert "系统关闭方式" in invalid_shutdown.json()["detail"]

    app.state.invoice_hub._package_manifest["platform"] = "windows"
    unsupported_desktop = client.put("/api/v1/preferences", json={"startup_surface": "desktop"})
    assert unsupported_desktop.status_code == 422
    assert "Windows 便携版" in unsupported_desktop.json()["detail"]


def test_about_api_is_local_only_and_update_check_payload_is_strict(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    client = TestClient(app)

    def unexpected_transport(*_args, **_kwargs):
        raise AssertionError("GET /about must not access the network")

    app.state.invoice_hub._update_service.transport = unexpected_transport
    about = client.get("/api/v1/about")
    assert about.status_code == 200
    payload = about.json()
    assert payload["product"]["version"] == "0.3.0-alpha.1"
    assert payload["package"]["manifest_status"] == "missing"
    assert payload["update"]["status"] == "idle"

    assert client.post("/api/v1/update/check", json={"force": "yes"}).status_code == 400
    assert client.post("/api/v1/update/check", json={"force": True, "url": "https://example.com"}).status_code == 400

def test_server_shutdown_api_preserves_or_stops_monitor_and_finalizes_state(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    monitor = {"running": True, "pid": 4321}
    stop_calls: list[bool] = []
    scheduled_states = []

    def fake_bridge_status(self) -> dict:
        return dict(monitor)

    def fake_bridge_stop(self) -> dict:
        stop_calls.append(True)
        monitor.update({"running": False, "pid": 0})
        return {"ok": True, "running": False, "status": dict(monitor)}

    monkeypatch.setattr("invoice_hub.services.app_state.AppState.bridge_status", fake_bridge_status)
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.bridge_stop", fake_bridge_stop)

    app = create_app(tmp_path, shutdown_scheduler=scheduled_states.append)
    client = TestClient(app)
    state = app.state.invoice_hub

    invalid = client.post("/api/v1/server/shutdown", json={"shutdown_behavior": "ask", "remember": False})
    assert invalid.status_code == 400
    invalid_remember = client.post("/api/v1/server/shutdown", json={"shutdown_behavior": "keep_monitor", "remember": "yes"})
    assert invalid_remember.status_code == 400
    assert scheduled_states == []

    state.layout.server_pid.write_text("987654", encoding="utf-8")
    keep = client.post("/api/v1/server/shutdown", json={"shutdown_behavior": "keep_monitor", "remember": True})
    assert keep.status_code == 200
    keep_payload = keep.json()
    assert keep_payload["ok"] is True
    assert keep_payload["scheduled"] is True
    assert keep_payload["shutdown_behavior"] == "keep_monitor"
    assert keep_payload["monitor_was_running"] is True
    assert keep_payload["monitor_running"] is True
    assert stop_calls == []
    assert scheduled_states == [state]
    assert client.get("/api/v1/preferences").json()["preferences"]["system_shutdown_behavior"] == "keep_monitor"

    stopping_state = json.loads(state.layout.server_state.read_text(encoding="utf-8"))
    assert stopping_state["status"] == "stopping"
    assert stopping_state["shutdown_behavior"] == "keep_monitor"
    assert stopping_state["monitor_running"] is True

    duplicate = client.post("/api/v1/server/shutdown", json={"shutdown_behavior": "stop_monitor", "remember": False})
    assert duplicate.status_code == 200
    assert duplicate.json()["idempotent"] is True
    assert duplicate.json()["scheduled"] is False
    assert scheduled_states == [state]

    state.finalize_server_shutdown()
    assert not state.layout.server_pid.exists()
    stopped_state = json.loads(state.layout.server_state.read_text(encoding="utf-8"))
    assert stopped_state["status"] == "stopped"
    assert stopped_state["shutdown_behavior"] == "keep_monitor"
    assert stopped_state["monitor_running"] is True
    assert "pid" not in stopped_state

    monitor.update({"running": True, "pid": 8765})
    stop_root = tmp_path / "stop-case"
    stop_app = create_app(stop_root, shutdown_scheduler=scheduled_states.append)
    stop_client = TestClient(stop_app)
    stop_state = stop_app.state.invoice_hub
    stop_response = stop_client.post("/api/v1/server/shutdown", json={"shutdown_behavior": "stop_monitor", "remember": False})
    assert stop_response.status_code == 200
    stop_payload = stop_response.json()
    assert stop_payload["scheduled"] is True
    assert stop_payload["shutdown_behavior"] == "stop_monitor"
    assert stop_payload["monitor_was_running"] is True
    assert stop_payload["monitor_running"] is False
    assert len(stop_calls) == 1
    assert scheduled_states[-1] is stop_state
    assert stop_client.get("/api/v1/preferences").json()["preferences"]["system_shutdown_behavior"] == "ask"
    stop_state.finalize_server_shutdown()
    assert json.loads(stop_state.layout.server_state.read_text(encoding="utf-8"))["status"] == "stopped"


def test_server_shutdown_does_not_close_webui_when_monitor_remains_running(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    monitor_status = {"running": True, "pid": 4321}
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.bridge_status", lambda self: dict(monitor_status))
    monkeypatch.setattr(
        "invoice_hub.services.app_state.AppState.bridge_stop",
        lambda self: {"ok": True, "running": False, "status": dict(monitor_status)},
    )
    scheduled_states = []
    app = create_app(tmp_path, shutdown_scheduler=scheduled_states.append)
    client = TestClient(app)

    response = client.post(
        "/api/v1/server/shutdown",
        json={"shutdown_behavior": "stop_monitor", "remember": True},
    )

    assert response.status_code == 500
    assert "监控未能停止" in response.json()["detail"]
    assert scheduled_states == []
    assert client.get("/api/v1/preferences").json()["preferences"]["system_shutdown_behavior"] == "ask"
    events = app.state.invoice_hub.repo.list_events_after(0, limit=20)
    assert any(item["event_type"] == "server.shutdown_failed" for item in events)


def test_diagnostics_apis_export_support_package(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    client = TestClient(app)

    summary = client.get("/api/v1/diagnostics/summary")
    assert summary.status_code == 200
    summary_payload = summary.json()
    assert summary_payload["ok"] is True
    assert summary_payload["target_id"]
    assert summary_payload["paths"]["watch_dir"]
    assert summary_payload["paths"]["runtime_dir"] == str(tmp_path / "runtime")
    assert "monitor" in summary_payload
    assert "products" in summary_payload
    assert summary_payload["safety"]["contains_source_invoices"] is False
    assert summary_payload["safety"]["contains_projection_files"] is False
    assert "config/app.local.json" in summary_payload["release_warning"]
    assert "InvoiceHub 诊断摘要" in summary_payload["text"]

    health = client.get("/api/v1/diagnostics/config-health")
    assert health.status_code == 200
    health_payload = health.json()
    assert health_payload["ok"] is True
    assert health_payload["overall"] in {"ok", "warning", "danger"}
    check_keys = {item["key"] for item in health_payload["checks"]}
    assert {"watch_dir", "runtime_dir", "skin_recovery", "release_config"}.issubset(check_keys)

    exported = client.post("/api/v1/diagnostics/support-package")
    assert exported.status_code == 200
    package_payload = exported.json()
    package_path = Path(package_payload["package_path"]).resolve()
    support_dir = (tmp_path / "runtime" / "local_state" / "support_packages").resolve()
    assert package_path.parent == support_dir
    assert package_path.exists()
    assert package_payload["manifest"]["contains_source_invoices"] is False
    assert package_payload["manifest"]["contains_projection_files"] is False

    with ZipFile(package_path) as archive:
        names = set(archive.namelist())
        assert {"manifest.json", "diagnostic_summary.json", "diagnostic_summary.txt", "config_health.json", "events_tail.json"}.issubset(names)
        assert any(name.startswith("logs/") and name.endswith(".txt") for name in names)
        manifest = json.loads(archive.read("manifest.json").decode("utf-8"))
        assert manifest["contains_source_invoices"] is False
        assert manifest["contains_projection_files"] is False
        assert not any(name.lower().endswith((".pdf", ".ofd", ".xml")) for name in names)
        forbidden_projection_names = {"发票汇总.csv", "发票汇总.xlsx", "成本发票明细.csv", "成本发票汇总.xlsx", "成本开票状态.json"}
        assert not any(any(forbidden in name for forbidden in forbidden_projection_names) for name in names)

def test_events_stream_cursor_defaults_to_latest_and_keeps_explicit_replay(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    history = state.append_event("history.changed", {"phase": "old"})

    assert _resolve_event_stream_cursor(state, None, None) == int(history["seq"])
    assert _resolve_event_stream_cursor(state, 0, None) == 0
    assert _resolve_event_stream_cursor(state, int(history["seq"]) - 1, None) == int(history["seq"]) - 1


def test_events_stream_last_event_id_reads_only_following_events(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    first = state.append_event("history.first", {})
    second = state.append_event("history.second", {})
    cursor = _resolve_event_stream_cursor(state, None, str(second["seq"]))
    future = state.append_event("future.after-reconnect", {})

    events = state.wait_events(cursor)
    assert [event["event_type"] for event in events] == ["future.after-reconnect"]
    assert int(events[0]["seq"]) == int(future["seq"]) > int(second["seq"]) > int(first["seq"])


def test_health_reads_packaged_build_manifest(tmp_path: Path) -> None:
    (tmp_path / "src").mkdir()
    (tmp_path / "web").mkdir()
    (tmp_path / "scripts" / "tools").mkdir(parents=True)
    (tmp_path / "docs" / "jierui").mkdir(parents=True)
    (tmp_path / "src" / "a.py").write_text("x = 1\n", encoding="utf-8")
    (tmp_path / "web" / "index.html").write_text("ok\n", encoding="utf-8")
    (tmp_path / "scripts" / "tools" / "jierui_voucher_import.py").write_text("MODE = 'dry-run'\n", encoding="utf-8")
    (tmp_path / "docs" / "jierui" / "voucher-import-template.facts.json").write_text("{}\n", encoding="utf-8")
    (tmp_path / "pyproject.toml").write_text("[project]\n", encoding="utf-8")
    manifest = write_build_manifest(
        tmp_path,
        tmp_path / "invoice-hub-build.json",
        source_commit="e38126b",
        built_at="2026-06-18T00:00:00Z",
    )

    health = TestClient(create_app(tmp_path)).get("/api/v1/health").json()

    assert health["build_manifest_present"] is True
    assert health["build_id"] == manifest["build_id"]
    assert health["bookkeeping_protocol_version"] == manifest["bookkeeping_protocol_version"] == BOOKKEEPING_PROTOCOL_VERSION
    assert health["capabilities"] == manifest["capabilities"] == list(API_CAPABILITIES)
    assert health["source_commit"] == "e38126b"
    assert health["built_at"] == "2026-06-18T00:00:00Z"


def test_health_uses_capabilities_from_present_build_manifest(tmp_path: Path) -> None:
    (tmp_path / "src").mkdir()
    (tmp_path / "web").mkdir()
    (tmp_path / "scripts" / "tools").mkdir(parents=True)
    (tmp_path / "docs" / "jierui").mkdir(parents=True)
    (tmp_path / "src" / "a.py").write_text("x = 1\n", encoding="utf-8")
    (tmp_path / "web" / "index.html").write_text("ok\n", encoding="utf-8")
    (tmp_path / "scripts" / "tools" / "jierui_voucher_import.py").write_text("MODE = 'dry-run'\n", encoding="utf-8")
    (tmp_path / "docs" / "jierui" / "voucher-import-template.facts.json").write_text("{}\n", encoding="utf-8")
    (tmp_path / "pyproject.toml").write_text("[project]\n", encoding="utf-8")
    manifest_path = tmp_path / "invoice-hub-build.json"
    write_build_manifest(tmp_path, manifest_path, source_commit="e38126b", built_at="2026-06-18T00:00:00Z")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    manifest["capabilities"] = ["manifest-only.v1"]
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")

    health = TestClient(create_app(tmp_path)).get("/api/v1/health").json()

    assert health["build_manifest_present"] is True
    assert health["capabilities"] == ["manifest-only.v1"]


def test_business_dossier_infers_company_folder_and_opens_known_links(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    # Windows forbids ':' in directory names; this is an explicit synthetic
    # Chinese/percent-containing company-name fixture.
    company = tmp_path / "测试公司-一般纳税人13%"
    cost = company / "成本发票"
    bank = company / "银行流水"
    deduction = company / "进项抵扣明细"
    issued = company / "开具发票"
    for folder in (cost, bank, deduction, issued):
        folder.mkdir(parents=True)
    (cost / "dzfp_10000000000000000001.pdf").write_text("pdf", encoding="utf-8")
    (cost / "成本发票明细.csv").write_text("", encoding="utf-8")
    (cost / "成本发票汇总.xlsx").write_text("", encoding="utf-8")
    (bank / "测试公司4月.pdf").write_text("pdf", encoding="utf-8")
    (deduction / "用途确认信息.xlsx").write_text("", encoding="utf-8")
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps(
            {
                "host": "127.0.0.1",
                "port": 8766,
                "watch_dir": str(cost),
                "runtime_dir": "./runtime",
                "reference_markup_rate": "0.08",
                "release_capabilities": {"local_ocr": False},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    client = TestClient(create_app(tmp_path))

    dossier = client.get("/api/v1/business-dossier").json()

    assert dossier["ok"] is True
    assert dossier["business_dir"] == str(company)
    assert dossier["watch_dir"] == str(cost)
    assert dossier["stats"]["invoice_files"] == 2
    assert dossier["stats"]["bank_flow_files"] == 1
    assert dossier["stats"]["deduction_files"] == 1
    keys = {item["key"] for item in dossier["links"]}
    assert {"business_dir", "watch_dir", "cost_invoice_dir", "bank_flow_dir", "input_deduction_dir", "issued_invoice_dir", "cost_summary_xlsx"} <= keys
    assert "已连通公司资料夹" in dossier["summary"]

    opened = client.post("/api/v1/business-dossier/open", json={"key": "bank_flow_dir"}).json()
    assert opened["ok"] is True
    assert opened["opened"] is True
    assert opened["path"] == str(bank)

    blocked = client.post("/api/v1/business-dossier/open", json={"path": str(tmp_path.parent)}).json()
    assert blocked["ok"] is False
    assert "当前业务资料夹" in blocked["message"]


def test_business_dossier_scan_is_bounded_and_marks_partial_counts(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    monkeypatch.setattr("invoice_hub.services.app_state.BUSINESS_DOSSIER_SCAN_MAX_ENTRIES", 3)
    monkeypatch.setattr("invoice_hub.services.app_state.BUSINESS_DOSSIER_SCAN_MAX_SECONDS", 30.0)
    company = tmp_path / "大型公司资料夹"
    cost = company / "成本发票"
    cost.mkdir(parents=True)
    for index in range(12):
        (cost / f"invoice-{index}.pdf").write_text("pdf", encoding="utf-8")
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps(
            {
                "host": "127.0.0.1",
                "port": 8766,
                "watch_dir": str(cost),
                "runtime_dir": "./runtime",
                "reference_markup_rate": "0.08",
                "release_capabilities": {"local_ocr": False},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )

    dossier = TestClient(create_app(tmp_path)).get("/api/v1/business-dossier").json()

    assert dossier["scan"] == {
        "complete": False,
        "truncated": True,
        "reason": "entry_limit",
        "scanned_entries": 3,
        "unreadable_directories": 0,
        "unreadable_entries": 0,
    }
    assert dossier["stats"]["total_files"] <= 2
    assert "已快速统计至少" in dossier["summary"]
    assert all(link["file_count_complete"] is False for link in dossier["links"] if link["is_dir"])


def test_business_dossier_iterator_oserror_returns_partial_lower_bound(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    company = tmp_path / "可部分读取的公司资料夹"
    cost = company / "成本发票"
    other = company / "其他资料"
    cost.mkdir(parents=True)
    other.mkdir()
    (cost / "invoice-a.pdf").write_text("pdf", encoding="utf-8")
    (other / "bank.xlsx").write_text("sheet", encoding="utf-8")
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps(
            {
                "host": "127.0.0.1",
                "port": 8766,
                "watch_dir": str(cost),
                "runtime_dir": "./runtime",
                "reference_markup_rate": "0.08",
                "release_capabilities": {"local_ocr": False},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    original_scandir = os.scandir

    class InterruptedScandir:
        def __init__(self, entries) -> None:
            self.entries = entries
            self.failed = False

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc_value, traceback) -> bool:
            self.close()
            return False

        def __iter__(self):
            return self

        def __next__(self):
            if self.failed:
                raise OSError("directory iterator interrupted")
            self.failed = True
            return next(self.entries)

        def close(self) -> None:
            self.entries.close()

    def scandir_with_interrupted_root(path):
        entries = original_scandir(path)
        if Path(path) == company:
            return InterruptedScandir(entries)
        return entries

    monkeypatch.setattr("invoice_hub.services.app_state.os.scandir", scandir_with_interrupted_root)

    response = TestClient(create_app(tmp_path)).get("/api/v1/business-dossier")

    assert response.status_code == 200
    dossier = response.json()
    assert dossier["scan"]["complete"] is False
    assert dossier["scan"]["truncated"] is False
    assert dossier["scan"]["reason"] == "unreadable_entries"
    assert dossier["scan"]["unreadable_directories"] == 1
    assert "至少" in dossier["summary"]
    assert "统计不完整" in dossier["summary"]
    assert all(link["file_count_complete"] is False for link in dossier["links"] if link["is_dir"])


def test_business_dossier_open_runs_in_threadpool(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    company = tmp_path / "公司资料夹"
    cost = company / "成本发票"
    cost.mkdir(parents=True)
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps(
            {
                "host": "127.0.0.1",
                "port": 8766,
                "watch_dir": str(cost),
                "runtime_dir": "./runtime",
                "reference_markup_rate": "0.08",
                "release_capabilities": {"local_ocr": False},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    api_module = importlib.import_module("invoice_hub.api.app")
    calls: list[tuple[str, tuple[object, ...]]] = []

    async def capture_threadpool(function, *args, **kwargs):
        assert not kwargs
        calls.append((function.__name__, args))
        return function(*args)

    monkeypatch.setattr(api_module, "run_in_threadpool", capture_threadpool)
    response = TestClient(create_app(tmp_path)).post("/api/v1/business-dossier/open", json={"key": "business_dir"})

    assert response.status_code == 200
    assert response.json()["opened"] is True
    assert calls == [("open_business_dossier", ({"key": "business_dir"},))]


def test_skin_api_defaults_to_no_skin_and_enables_builtin(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    client = TestClient(create_app(tmp_path))

    skins = client.get("/api/v1/skins").json()

    assert skins["ok"] is True
    assert skins["enabled_skin_id"] is None
    assert skins["active_skin"] is None
    assert skins["default_skin_id"] is None
    assert skins["skins"]
    assert all(item["read_only"] is True for item in skins["skins"] if item["builtin"])
    builtins = {item["id"]: item for item in skins["skins"] if item["builtin"]}
    assert set(builtins) == {"animal-island", "ink-pulse"}
    assert builtins["ink-pulse"]["version"] == "1.3.0"
    builtin = builtins["animal-island"]

    css = client.get(builtin["stylesheet_url"])
    assert css.status_code == 200
    assert "text/css" in css.headers["content-type"]
    assert css.headers.get("cache-control") == "public, max-age=31536000, immutable"
    assert b"@import" not in css.content

    ink_css = client.get(builtins["ink-pulse"]["stylesheet_url"])
    assert ink_css.status_code == 200
    assert b"--ink-lime" in ink_css.content
    assert b"@import" not in ink_css.content

    versioned_static = client.get("/static/css/app.css?v=contract")
    assert versioned_static.status_code == 200
    assert versioned_static.headers.get("cache-control") == "public, max-age=31536000, immutable"
    unversioned_static = client.get("/static/css/app.css")
    assert unversioned_static.status_code == 200
    assert unversioned_static.headers.get("cache-control") != "public, max-age=31536000, immutable"

    enabled = client.post(f"/api/v1/skins/{builtin['id']}/enable").json()
    assert enabled["enabled_skin_id"] == builtin["id"]
    assert enabled["active_skin"]["read_only"] is True

    reset = client.post("/api/v1/skins/reset").json()
    assert reset["enabled_skin_id"] is None
    assert reset["active_skin"] is None


def test_skin_import_stores_files_under_runtime_local_state(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    client = TestClient(app)
    watch_dir = Path(app.state.invoice_hub.active_profile.watch_dir).resolve()

    response = client.post(
        "/api/v1/skins/import",
        content=_make_skin_zip(files={"skin.css": ".app { color: #123456; background-image: url(assets/logo.png); }", "assets/logo.png": b"\x89PNG\r\n\x1a\n"}),
        headers={"content-type": "application/zip"},
    )

    assert response.status_code == 200
    imported = response.json()["skin"]
    storage_dir = Path(imported["storage_dir"]).resolve()
    assert storage_dir.is_relative_to((tmp_path / "runtime" / "local_state" / "skins" / "imported").resolve())
    assert not storage_dir.is_relative_to(watch_dir)
    assert (storage_dir / "skin.json").is_file()
    assert (storage_dir / "skin.css").is_file()

    skins = client.get("/api/v1/skins").json()
    assert any(item["id"] == "custom-skin" and item["imported"] and not item["read_only"] for item in skins["skins"])

    css = client.get("/api/v1/skins/custom-skin/files/skin.css")
    assert css.status_code == 200
    assert b"#123456" in css.content
    traversal = client.get("/api/v1/skins/custom-skin/files/%2e%2e/skin.css")
    assert traversal.status_code == 400


def test_skin_import_accepts_wrapped_zip_with_safe_metadata(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    client = TestClient(create_app(tmp_path))
    payload = BytesIO()
    with ZipFile(payload, "w") as archive:
        archive.writestr("wrapped-portable/skin.json", json.dumps({
            "id": "wrapped-portable",
            "name": "Wrapped Portable",
            "version": "1.2.1",
            "entry": "skin.css",
        }, ensure_ascii=False))
        archive.writestr("wrapped-portable/skin.css", ":root { --accent: #111111; background-image: url(textures/pulse.png); }\n")
        archive.writestr("wrapped-portable/textures/pulse.png", b"\x89PNG\r\n\x1a\n")
        archive.writestr("wrapped-portable/asset-sources.json", "{}")
        archive.writestr("wrapped-portable/fonts/OFL-WrappedPortable.txt", "Open Font License")
        archive.writestr("__MACOSX/._wrapped-portable", b"mac")
        archive.writestr("__MACOSX/wrapped-portable/._skin.css", b"mac")

    response = client.post(
        "/api/v1/skins/import",
        content=payload.getvalue(),
        headers={"content-type": "application/zip"},
    )

    assert response.status_code == 200
    imported = response.json()["skin"]
    storage_dir = Path(imported["storage_dir"])
    assert (storage_dir / "skin.json").is_file()
    assert (storage_dir / "skin.css").is_file()
    assert (storage_dir / "textures" / "pulse.png").is_file()
    assert (storage_dir / "asset-sources.json").is_file()
    assert (storage_dir / "fonts" / "OFL-WrappedPortable.txt").is_file()
    assert not (storage_dir / "__MACOSX").exists()


def test_skin_import_rejects_unsafe_zip_content(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    client = TestClient(create_app(tmp_path))
    cases = [
        {"skin.css": "@import url('other.css');"},
        {"skin.css": ".x { background: url(https://example.test/a.png); }"},
        {"skin.css": ".x { background: url(data:image/png;base64,AAAA); }"},
        {"skin.css": ".x { background: url(javascript:alert(1)); }"},
        {"skin.css": ".x {}", "script.js": "alert(1);"},
        {"skin.css": ".x {}", "../escape.css": ".x {}"},
        {"skin.css": ".x {}", "notes.json": "{}"},
    ]
    for index, files in enumerate(cases):
        response = client.post(
            "/api/v1/skins/import",
            content=_make_skin_zip(skin_id=f"bad-skin-{index}", files=files),
            headers={"content-type": "application/zip"},
        )
        assert response.status_code == 400

    too_many = {f"assets/{index}.png": b"p" for index in range(MAX_SKIN_FILES + 1)}
    too_many["skin.css"] = ".x {}"
    assert client.post("/api/v1/skins/import", content=_make_skin_zip("too-many", too_many), headers={"content-type": "application/zip"}).status_code == 400

    too_large = {"skin.css": ".x {}", "assets/large.png": b"0" * (MAX_SKIN_FILE_BYTES + 1)}
    assert client.post("/api/v1/skins/import", content=_make_skin_zip("too-large", too_large), headers={"content-type": "application/zip"}).status_code == 400

    builtin_collision = client.post("/api/v1/skins/import", content=_make_skin_zip("animal-island"), headers={"content-type": "application/zip"})
    assert builtin_collision.status_code == 409

    multipart = client.post(
        "/api/v1/skins/import",
        content=_make_skin_zip("multipart-skin"),
        headers={"content-type": "multipart/form-data; boundary=abc"},
    )
    assert multipart.status_code == 415

    json_upload = client.post("/api/v1/skins/import", json={"zip_base64": ""})
    assert json_upload.status_code == 415


def test_skin_replace_overwrites_imported_skin(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    client = TestClient(create_app(tmp_path))

    created = client.post("/api/v1/skins/import", content=_make_skin_zip("replaceable", {"skin.css": ".x { color: #111111; }"}), headers={"content-type": "application/zip"})
    assert created.status_code == 200
    duplicate = client.post("/api/v1/skins/import", content=_make_skin_zip("replaceable", {"skin.css": ".x { color: #222222; }"}), headers={"content-type": "application/zip"})
    assert duplicate.status_code == 409

    replaced = client.post("/api/v1/skins/replace", content=_make_skin_zip("replaceable", {"skin.css": ".x { color: #333333; }"}), headers={"content-type": "application/zip"})

    assert replaced.status_code == 200
    assert replaced.json()["replaced"] is True
    assert replaced.json()["enabled_skin_id"] == "replaceable"
    assert replaced.json()["active_skin"]["id"] == "replaceable"
    css = client.get("/api/v1/skins/replaceable/files/skin.css")
    assert css.status_code == 200
    assert b"#333333" in css.content

    mismatch = client.post(
        "/api/v1/skins/other-skin/replace",
        content=_make_skin_zip("replaceable", {"skin.css": ".x { color: #444444; }"}),
        headers={"content-type": "application/zip"},
    )
    assert mismatch.status_code == 400


def test_skin_page_injects_active_skin_and_no_skin_bypasses(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    client = TestClient(create_app(tmp_path))
    assert client.post("/api/v1/skins/animal-island/enable").status_code == 200

    page = client.get("/").text
    assert 'id="activeSkinStylesheet"' in page
    assert "/api/v1/skins/animal-island/files/skin.css" in page

    bypass = client.get("/?no_skin=1").text
    assert 'id="activeSkinStylesheet"' not in bypass

    backend = client.get("/backend").text
    assert 'id="activeSkinStylesheet"' not in backend


def test_reference_status_api_rejects_quantity_over_total(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [
            {
                "销售方": "A",
                "购买方": "B",
                "发票号码": "1",
                "开票日期": "2026-05-24",
                "内部项目名称": "项目",
                "规格型号": "规格",
                "单位": "吨",
                "数量": "10",
                "单价(除税)": "10",
                "金额(除税)": "100",
                "税率": "9%",
                "税金": "9",
                "价税合计": "109",
                "发票代码(**内文字)": "材料",
                "源文件": "a.pdf",
            }
        ],
    )
    app = create_app(tmp_path)
    client = TestClient(app)
    reference = client.get("/api/v1/cost-analysis").json()["invoice_reference"][0]
    assert reference["average_unit_price"] == 10.0
    assert reference["average_unit_price_with_tax"] == 10.9

    response = client.post(
        "/api/v1/cost-analysis/reference-status",
        json={"items": [{"key": reference["key"], "invoiced_quantity": "10.001"}]},
    )

    assert response.status_code == 400
    assert "已开数量不能大于数量合计" in response.json()["detail"]


def test_reference_status_api_accepts_row_markup_rate(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    app = create_app(tmp_path)
    client = TestClient(app)
    deadline = time.time() + 3
    while time.time() < deadline and app.state.invoice_hub._background_status in {"initializing", "running"}:
        time.sleep(0.05)
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [
            {
                "销售方": "A",
                "购买方": "B",
                "发票号码": "1",
                "开票日期": "2026-05-24",
                "内部项目名称": "项目",
                "规格型号": "规格",
                "单位": "吨",
                "数量": "10",
                "单价(除税)": "10",
                "金额(除税)": "100",
                "税率": "0.13",
                "税金": "13",
                "价税合计": "113",
                "发票代码(**内文字)": "材料",
                "源文件": "a.pdf",
            }
        ],
    )

    reference = client.get("/api/v1/cost-analysis").json()["invoice_reference"][0]
    response = client.post(
        "/api/v1/cost-analysis/reference-status",
        json={
            "items": [
                {
                    "key": reference["key"],
                    "invoiced_quantity": "0",
                    "reference_markup_rate_percent": "20",
                    "reference_markup_locked": True,
                }
            ]
        },
    ).json()
    assert response["ok"] is True

    cost = client.get("/api/v1/cost-analysis").json()
    assert cost["items"][0]["平均单价(含税)"] == 11.3
    assert cost["invoice_reference"][0]["average_unit_price_with_tax"] == 11.3
    assert cost["invoice_reference"][0]["reference_average_unit_price_with_tax"] == 13.56
    assert cost["invoice_reference"][0]["reference_total_with_tax"] == 135.6
    assert cost["invoice_reference"][0]["markup_rate"] == "20%"
    assert cost["invoice_reference"][0]["reference_markup_rate_percent"] == "20"
    assert cost["invoice_reference"][0]["reference_markup_locked"] is True

    saved = json.loads((watch / "成本开票状态.json").read_text(encoding="utf-8-sig"))
    assert saved["items"][reference["key"]]["reference_markup_rate"] == "0.2"
    assert saved["items"][reference["key"]]["reference_markup_locked"] is True

    invalid = client.post(
        "/api/v1/cost-analysis/reference-status",
        json={"items": [{"key": reference["key"], "invoiced_quantity": "0", "reference_markup_rate_percent": "abc"}]},
    )
    assert invalid.status_code == 400
    assert "开票加价率不是有效数字" in invalid.json()["detail"]


def test_reference_status_api_waits_for_profile_lock_without_blocking_health(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [_cost_detail_row("10000000000000000001", "cost.pdf", "项目", "规格", "2", "200", "26", "226")],
    )
    reference = state.cost_snapshot()["invoice_reference"][0]

    holder_state = MonitorState(state.active_profile, state.layout.db_path)
    lock_entered = threading.Event()
    release_lock = threading.Event()
    save_entered = threading.Event()
    save_finished = threading.Event()
    health_finished = threading.Event()
    responses = {}
    original_save = state.save_cost_reference_status

    def hold_lock() -> None:
        with holder_state.sync_write_lock():
            lock_entered.set()
            release_lock.wait(timeout=10)

    def observed_save(payload: dict) -> dict:
        save_entered.set()
        return original_save(payload)

    def request_save(client: TestClient) -> None:
        try:
            responses["save"] = client.post(
                "/api/v1/cost-analysis/reference-status",
                json={"items": [{"key": reference["key"], "invoiced_quantity": "1"}]},
            )
        finally:
            save_finished.set()

    def request_health(client: TestClient) -> None:
        try:
            responses["health"] = client.get("/api/v1/health")
        finally:
            health_finished.set()

    monkeypatch.setattr(state, "save_cost_reference_status", observed_save)
    holder = threading.Thread(target=hold_lock)
    with TestClient(app) as client:
        holder.start()
        assert lock_entered.wait(timeout=2)
        save_thread = threading.Thread(target=request_save, args=(client,))
        health_thread = threading.Thread(target=request_health, args=(client,))
        save_thread.start()
        assert save_entered.wait(timeout=2)
        health_thread.start()
        try:
            health_responded_during_save = health_finished.wait(timeout=0.5)
            save_was_still_waiting = not save_finished.is_set()
        finally:
            release_lock.set()
            holder.join(timeout=2)
            save_thread.join(timeout=5)
            health_thread.join(timeout=2)

    assert health_responded_during_save
    assert save_was_still_waiting
    assert responses["health"].status_code == 200
    assert responses["save"].status_code == 200
    saved = json.loads((watch / "成本开票状态.json").read_text(encoding="utf-8-sig"))
    assert saved["items"][reference["key"]]["invoiced_quantity"] == "1"
    snapshot = responses["save"].json()
    assert snapshot["ok"] is True

    from openpyxl import load_workbook

    workbook = load_workbook(watch / "成本发票汇总.xlsx", read_only=True, data_only=True)
    try:
        sheet = workbook["开票参考"]
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        values = next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        assert float(values[headers.index("已开数量")]) == 1.0
    finally:
        workbook.close()


def test_cost_analysis_project_summary_contains_tax_average_unit_price(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    write_csv_rows(
        watch / "成本发票明细.csv",
        DETAIL_HEADERS,
        [
            {
                "销售方": "A",
                "购买方": "B",
                "发票号码": "1",
                "开票日期": "2026-05-24",
                "内部项目名称": "项目",
                "规格型号": "规格",
                "单位": "吨",
                "数量": "10",
                "单价(除税)": "9999",
                "金额(除税)": "34210",
                "税率": "13%",
                "税金": "3078.9",
                "价税合计": "37288.9",
                "发票代码(**内文字)": "材料",
                "源文件": "a.pdf",
            }
        ],
    )
    app = create_app(tmp_path)
    client = TestClient(app)

    project = client.get("/api/v1/cost-analysis").json()["project_summary"][0]

    assert project["平均单价(除税)"] == 3421.0
    assert project["平均单价(含税)"] == 3728.89
    assert project["库存平均单价(除税)"] == 3421.0
    assert project["库存平均单价(含税)"] == 3728.89
    assert project["采购参考平均单价(含税)"] == 3728.89


def test_cost_analysis_project_summary_splits_stock_and_purchase_reference_averages(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    rows = [
        {
            "销售方": "A",
            "购买方": "B",
            "发票号码": "1",
            "开票日期": "2026-05-24",
            "内部项目名称": "项目",
            "规格型号": "规格",
            "单位": "吨",
            "数量": "2",
            "单价(除税)": "100",
            "金额(除税)": "200",
            "税率": "9%",
            "税金": "18",
            "价税合计": "218",
            "发票代码(**内文字)": "材料",
            "源文件": "a.pdf",
        },
        {
            "销售方": "A",
            "购买方": "B",
            "发票号码": "2",
            "开票日期": "2026-05-25",
            "内部项目名称": "项目",
            "规格型号": "规格",
            "单位": "吨",
            "数量": "1",
            "单价(除税)": "200",
            "金额(除税)": "200",
            "税率": "13%",
            "税金": "26",
            "价税合计": "226",
            "发票代码(**内文字)": "材料",
            "源文件": "b.pdf",
        },
    ]
    write_csv_rows(watch / "成本发票明细.csv", DETAIL_HEADERS, rows)
    app = create_app(tmp_path)
    client = TestClient(app)

    cost = client.get("/api/v1/cost-analysis").json()

    assert cost["items"][0]["平均单价(含税)"] == 109.0
    assert cost["items"][1]["平均单价(含税)"] == 226.0
    assert cost["project_summary"][0]["平均单价(除税)"] == 133.33
    assert cost["project_summary"][0]["平均单价(含税)"] == 148.0
    assert cost["project_summary"][0]["库存平均单价(除税)"] == 133.33
    assert cost["project_summary"][0]["库存平均单价(含税)"] == 148.0
    assert cost["project_summary"][0]["采购参考平均单价(含税)"] == 167.5
    assert cost["invoice_reference"][0]["average_unit_price"] == 133.33333333333334
    assert cost["invoice_reference"][0]["average_unit_price_with_tax"] == 148.0
    assert cost["invoice_reference"][0]["purchase_reference_average_unit_price_with_tax"] == 167.5
    assert cost["invoice_reference"][0]["reference_average_unit_price"] == 144.0
    assert cost["invoice_reference"][0]["reference_average_unit_price_with_tax"] == 162.72
    assert cost["invoice_reference"][0]["reference_amount"] == 432.0
    assert cost["invoice_reference"][0]["reference_total_with_tax"] == 488.16


def test_cost_analysis_api_refreshes_old_cost_schema(monkeypatch, tmp_path: Path) -> None:
    from openpyxl import Workbook, load_workbook

    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    config_path = tmp_path / "config" / "app.local.json"
    config_path.parent.mkdir()
    config_path.write_text(
        json.dumps(
            {
                "host": "127.0.0.1",
                "port": 8766,
                "watch_dir": "./发票文件",
                "runtime_dir": "./runtime",
                "reference_markup_rate": "0.08",
                "release_capabilities": {"local_ocr": False},
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    app = create_app(tmp_path)
    client = TestClient(app)
    watch = Path(app.state.invoice_hub.active_profile.watch_dir)
    legacy_headers = [header for header in DETAIL_HEADERS if header != "平均单价(含税)"]
    write_csv_rows(
        watch / "成本发票明细.csv",
        legacy_headers,
        [
            {
                "销售方": "A",
                "购买方": "B",
                "发票号码": "1",
                "开票日期": "2026-05-24",
                "内部项目名称": "项目",
                "规格型号": "规格",
                "单位": "吨",
                "数量": "2",
                "单价(除税)": "9999",
                "金额(除税)": "6842",
                "税率": "13%",
                "税金": "615.78",
                "价税合计": "7457.78",
                "发票代码(**内文字)": "材料",
                "源文件": "a.pdf",
            }
        ],
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "开票参考"
    ws.append([header for header in ("发票代码(**内文字)", "内部项目名称", "规格型号", "单位", "数量合计", "已开数量", "未开数量", "开票状态", "平均单价(除税)", "金额(除税)合计", "税金合计", "价税合计", "已开参考价税合计", "未开参考金额(除税)", "未开参考税金", "未开参考价税合计", "参考加价率", "状态更新时间")])
    ws.append(["材料", "项目", "规格", "吨", 2, 0, 2, "未开具", 3694.68, 7389.36, 665.04, 8054.4, 0, 7389.36, 665.04, 8054.4, "8%", ""])
    wb.save(watch / "成本发票汇总.xlsx")
    cost = client.get("/api/v1/cost-analysis").json()

    assert cost["items"][0]["平均单价(含税)"] == 3728.89
    assert cost["project_summary"][0]["平均单价(含税)"] == 3728.89
    assert round(cost["invoice_reference"][0]["reference_average_unit_price_with_tax"], 2) == 4174.99
    assert round(cost["invoice_reference"][0]["reference_tax_amount"], 2) == 960.62
    assert round(cost["invoice_reference"][0]["reference_total_with_tax"], 2) == 8349.98
    refreshed = load_workbook(watch / "成本发票汇总.xlsx", read_only=True, data_only=True)
    try:
        headers = [cell.value for cell in next(refreshed["开票参考"].iter_rows(min_row=1, max_row=1))]
        assert "平均单价(含税)" in headers
    finally:
        refreshed.close()


def test_cost_analysis_schema_refresh_waits_for_profile_lock_without_blocking_health(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    legacy_headers = [header for header in DETAIL_HEADERS if header != "平均单价(含税)"]
    write_csv_rows(
        watch / "成本发票明细.csv",
        legacy_headers,
        [_cost_detail_row("10000000000000000001", "legacy.pdf", "项目", "规格", "2", "200", "26", "226")],
    )

    holder_state = MonitorState(state.active_profile, state.layout.db_path)
    lock_entered = threading.Event()
    release_lock = threading.Event()
    snapshot_entered = threading.Event()
    snapshot_finished = threading.Event()
    health_finished = threading.Event()
    responses = {}
    original_snapshot = state.cost_snapshot

    def hold_lock() -> None:
        with holder_state.sync_write_lock():
            lock_entered.set()
            release_lock.wait(timeout=10)

    def observed_snapshot() -> dict:
        snapshot_entered.set()
        return original_snapshot()

    def request_snapshot(client: TestClient) -> None:
        try:
            responses["snapshot"] = client.get("/api/v1/cost-analysis")
        finally:
            snapshot_finished.set()

    def request_health(client: TestClient) -> None:
        try:
            responses["health"] = client.get("/api/v1/health")
        finally:
            health_finished.set()

    monkeypatch.setattr(state, "cost_snapshot", observed_snapshot)
    holder = threading.Thread(target=hold_lock)
    with TestClient(app) as client:
        holder.start()
        assert lock_entered.wait(timeout=2)
        snapshot_thread = threading.Thread(target=request_snapshot, args=(client,))
        health_thread = threading.Thread(target=request_health, args=(client,))
        snapshot_thread.start()
        assert snapshot_entered.wait(timeout=2)
        health_thread.start()
        try:
            health_responded_during_refresh = health_finished.wait(timeout=0.5)
            refresh_was_still_waiting = not snapshot_finished.is_set()
        finally:
            release_lock.set()
            holder.join(timeout=2)
            snapshot_thread.join(timeout=5)
            health_thread.join(timeout=2)

    assert health_responded_during_refresh
    assert refresh_was_still_waiting
    assert responses["health"].status_code == 200
    assert responses["snapshot"].status_code == 200
    assert responses["snapshot"].json()["items"][0]["平均单价(含税)"] == 113.0
    headers = (watch / "成本发票明细.csv").read_text(encoding="utf-8-sig").splitlines()[0].split(",")
    assert "平均单价(含税)" in headers
    assert (watch / "成本发票汇总.xlsx").exists()


def test_bridge_rebuild_and_events(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    (watch / "sample.xml").write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    result = client.post("/api/v1/bridge/rebuild").json()
    assert result["ok"] is True
    assert client.get("/api/v1/invoices").json()["count"] == 1
    task = client.get(f"/api/v1/tasks/{result['task_id']}").json()
    assert task["status"] == "success"


def test_bridge_rebuild_reports_empty_archive_only_directory(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    (watch / "成本发票明细.csv").write_text("", encoding="utf-8")
    (watch / "synthetic-cost-invoices.rar").write_text("archive placeholder", encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    settings = client.get("/api/v1/settings").json()
    assert settings["path_validation"]["can_monitor"] is True
    assert settings["path_validation"]["has_supported_files"] is False
    assert settings["path_validation"]["supported_count"] == 0
    assert settings["path_validation"]["archive_count"] == 1
    assert "请先解压" in settings["path_validation"]["summary"]

    result = client.post("/api/v1/bridge/rebuild").json()

    assert result["ok"] is True
    assert result["detail"]["summary"]["count"] == 0
    assert "结果为 0 条" in result["message"]
    assert "请先解压" in result["message"]


def test_bridge_rebuild_builds_xml_cost_analysis_details(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    source = watch / "dzfp_10000000000000000013.xml"
    source.write_text(_cost_xml_text(), encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    result = client.post("/api/v1/bridge/rebuild").json()

    assert result["ok"] is True
    invoices = client.get("/api/v1/invoices").json()
    assert invoices["count"] == 1
    invoice = invoices["items"][0]
    assert invoice["file_format"] == "xml"
    assert invoice["invoice_type"] == "增值税专用发票"
    assert invoice["business_type"] == "标准电子发票"
    assert invoice["classification_status"] == "ok"
    assert invoice["classification_issue"] == ""
    assert client.get("/api/v1/invoices", params={"invoice_type": "增值税专用发票"}).json()["count"] == 1
    assert client.get("/api/v1/invoices", params={"business_type": "标准电子发票"}).json()["count"] == 1
    assert client.get("/api/v1/invoices", params={"classification_status": "ok"}).json()["count"] == 1
    detail = client.get("/api/v1/invoices/0").json()
    assert detail["invoice"]["invoice_number"] == "10000000000000000013"
    assert detail["invoice"]["invoice_type"] == "增值税专用发票"
    assert detail["invoice"]["business_type"] == "标准电子发票"
    assert detail["invoice"]["classification_status"] == "ok"
    assert detail["invoice"]["classification_issue"] == ""
    cost = client.get("/api/v1/cost-analysis").json()
    assert cost["detail_count"] == 2
    assert cost["items"][0]["内部项目名称"] == "钢筋"
    assert cost["checks"][0]["发票大类"] == "增值税专用发票"
    assert cost["checks"][0]["特定业务类型"] == "标准电子发票"
    assert cost["checks"][0]["类型识别状态"] == "ok"
    assert cost["checks"][0]["类型识别说明"] == ""
    assert cost["checks"][0]["校验状态"] == "通过"
    assert cost["sync"]["source_invoice_count"] == 1
    assert cost["sync"]["parsed_invoice_count"] == 1
    assert cost["sync"]["review_count"] == 0
    assert cost["sync"]["sync_state"] == "fresh"


def test_server_startup_background_sync_builds_invoice_summary(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    (watch / "sample.xml").write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    deadline = time.time() + 8
    payload = client.get("/api/v1/invoices").json()
    while time.time() < deadline and payload["count"] != 1:
        time.sleep(0.2)
        payload = client.get("/api/v1/invoices").json()

    assert payload["count"] == 1
    assert payload["snapshot"]["source_exists"] is True
    assert Path(payload["snapshot"]["source_path"]).exists()
    health = client.get("/api/v1/health").json()
    assert health["background_status"] in {"ready", "running"}


def test_background_startup_sync_uses_spawn_worker_and_clears_current_profile_cache(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    calls: dict[str, object] = {}

    class ResultPipe:
        def __init__(self) -> None:
            self.payload = None
            self.ready = threading.Event()

        def send(self, payload) -> None:
            self.payload = payload
            self.ready.set()

        def poll(self, timeout: float) -> bool:
            return self.ready.wait(timeout)

        def recv(self):
            if self.payload is None:
                raise EOFError
            return self.payload

        def close(self) -> None:
            return None

    class FakeProcess:
        exitcode = 0

        def __init__(self, target, args, name, daemon) -> None:
            self.target = target
            self.args = args
            self.name = name
            self.daemon = daemon

        def start(self) -> None:
            self.target(*self.args)

        def join(self, timeout: float | None = None) -> None:
            return None

        def is_alive(self) -> bool:
            return False

        def terminate(self) -> None:
            return None

        def close(self) -> None:
            return None

    class SpawnContext:
        def Pipe(self, duplex: bool) -> tuple[ResultPipe, ResultPipe]:
            assert duplex is False
            pipe = ResultPipe()
            return pipe, pipe

        def Process(self, **kwargs) -> FakeProcess:
            calls["process"] = kwargs
            return FakeProcess(**kwargs)

    def fake_sync_worker(profile_payload, db_path, reference_markup_rate, trigger, result_sender) -> None:
        calls["profile_payload"] = profile_payload
        calls["db_path"] = db_path
        calls["reference_markup_rate"] = reference_markup_rate
        calls["trigger"] = trigger
        result_sender.send({"ok": True, "sync": {"ok": True, "target_id": profile_payload["id"]}})

    context = SpawnContext()
    monkeypatch.setattr("invoice_hub.services.app_state.multiprocessing.get_context", lambda method: calls.setdefault("context", method) and context)
    monkeypatch.setattr("invoice_hub.services.app_state._run_background_sync_process", fake_sync_worker)
    state._invoice_cache_key = (1, 1)
    state._invoice_cache_rows = [{"文件名": "stale.csv"}]

    state.run_background_diagnostics("startup_sync")
    deadline = time.monotonic() + 2
    while time.monotonic() < deadline and state._background_status != "ready":
        time.sleep(0.01)

    process = calls["process"]
    assert calls["context"] == "spawn"
    assert process["target"] is fake_sync_worker
    assert process["name"] == "invoice-hub-background-sync"
    assert process["daemon"] is True
    assert calls["profile_payload"] == state.active_profile.model_dump()
    assert calls["db_path"] == str(state.layout.db_path)
    assert calls["trigger"] == "startup_sync"
    assert state._background_status == "ready"
    assert state._invoice_cache_key is None
    assert state._invoice_cache_rows == []


def test_background_sync_completion_does_not_invalidate_new_active_profile_cache(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    started = threading.Event()
    release = threading.Event()
    notifications: list[tuple[str, dict[str, int]]] = []
    captured_target_id = state.active_profile.id

    def capture_current_test_notification(monitor_state, trigger, counts) -> None:
        if monitor_state.profile.id == captured_target_id:
            notifications.append((trigger, counts))

    monkeypatch.setattr(
        MonitorState,
        "notify_invoice_change",
        capture_current_test_notification,
    )

    class ResultPipe:
        def __init__(self) -> None:
            self.payload = None
            self.ready = threading.Event()

        def send(self, payload) -> None:
            self.payload = payload
            self.ready.set()

        def poll(self, timeout: float) -> bool:
            return self.ready.wait(timeout)

        def recv(self):
            if self.payload is None:
                raise EOFError
            return self.payload

        def close(self) -> None:
            return None

    class FakeProcess:
        def __init__(self, target, args, name, daemon) -> None:
            self.target = target
            self.args = args
            self.name = name
            self.daemon = daemon
            self.thread: threading.Thread | None = None
            self.exitcode: int | None = None

        def start(self) -> None:
            self.thread = threading.Thread(target=self.target, args=self.args)
            self.thread.start()

        def join(self, timeout: float | None = None) -> None:
            assert self.thread is not None
            self.thread.join(timeout)
            if not self.thread.is_alive():
                self.exitcode = 0

        def is_alive(self) -> bool:
            return bool(self.thread and self.thread.is_alive())

        def terminate(self) -> None:
            release.set()

        def close(self) -> None:
            return None

    class SpawnContext:
        def Pipe(self, duplex: bool) -> tuple[ResultPipe, ResultPipe]:
            assert duplex is False
            pipe = ResultPipe()
            return pipe, pipe

        def Process(self, **kwargs) -> FakeProcess:
            return FakeProcess(**kwargs)

    def blocked_sync_worker(profile_payload, db_path, reference_markup_rate, trigger, result_sender) -> None:
        started.set()
        assert release.wait(timeout=5)
        result_sender.send(
            {
                "ok": True,
                "sync": {
                    "ok": True,
                    "target_id": profile_payload["id"],
                    "added": 1,
                    "updated": 0,
                    "deleted": 0,
                    "blocked": 0,
                    "manual_changed": 1,
                    "rebuilt": True,
                },
            }
        )

    monkeypatch.setattr("invoice_hub.services.app_state.multiprocessing.get_context", lambda method: SpawnContext())
    monkeypatch.setattr("invoice_hub.services.app_state._run_background_sync_process", blocked_sync_worker)

    state.run_background_diagnostics("startup_sync")
    assert started.wait(timeout=2)
    next_watch = tmp_path / "new-active-profile"
    next_watch.mkdir()
    next_profile = target_profile_for(state.config, next_watch)
    with state._lock:
        state._active_profile = next_profile
        state._invoice_cache_key = (9, 9)
        state._invoice_cache_rows = [{"文件名": "new-profile.csv"}]

    try:
        release.set()
        deadline = time.monotonic() + 2
        events = []
        while time.monotonic() < deadline:
            events = state.wait_events(0)
            if any(event["event_type"] == "server.background_stale" for event in events):
                break
            time.sleep(0.01)
    finally:
        release.set()

    stale_events = [event for event in events if event["event_type"] == "server.background_stale"]
    assert stale_events
    assert stale_events[-1]["payload"]["captured_target_id"] != stale_events[-1]["payload"]["active_target_id"]
    ordinary_sync_events = {
        "invoice.changed",
        "cost_analysis.updated",
        "monitor.sync_completed",
        "manual_edit.synced",
        "monitor.heartbeat",
        "monitor.sync_failed",
    }
    assert not ordinary_sync_events.intersection(event["event_type"] for event in state.wait_events(0))
    assert notifications == []
    assert state._background_status == "running"
    assert state._invoice_cache_key == (9, 9)
    assert state._invoice_cache_rows == [{"文件名": "new-profile.csv"}]


def test_background_spawn_worker_keeps_fastapi_responsive_while_sync_runs(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    real_get_context = multiprocessing.get_context

    class SlowSpawnContext:
        def __init__(self) -> None:
            self._delegate = real_get_context("spawn")

        def Pipe(self, duplex: bool):
            return self._delegate.Pipe(duplex=duplex)

        def Process(self, **kwargs):
            kwargs["target"] = _slow_background_sync_worker
            return self._delegate.Process(**kwargs)

    def get_slow_spawn_context(method: str):
        assert method == "spawn"
        return SlowSpawnContext()

    monkeypatch.setattr("invoice_hub.services.app_state.multiprocessing.get_context", get_slow_spawn_context)
    client = TestClient(app)
    state.run_background_diagnostics("startup_sync")

    process = None
    deadline = time.monotonic() + 5
    while time.monotonic() < deadline:
        with state._lock:
            process = state._background_process
        if process is not None and process.is_alive():
            break
        time.sleep(0.01)

    assert process is not None
    assert process.is_alive()
    request_started = time.monotonic()
    assert client.get("/").status_code == 200
    health = client.get("/api/v1/health")
    assert health.status_code == 200
    assert health.json()["background_status"] == "running"
    assert time.monotonic() - request_started < 0.75

    while time.monotonic() < deadline and state._background_status != "ready":
        time.sleep(0.02)
    assert state._background_status == "ready"


def test_new_background_generation_retires_prior_worker_with_bounded_kill(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub

    class PriorProcess:
        def __init__(self) -> None:
            self.alive = True
            self.terminated = 0
            self.killed = 0
            self.join_timeouts: list[float | None] = []
            self.closed = False
            self.kill_called = threading.Event()

        def is_alive(self) -> bool:
            return self.alive

        def terminate(self) -> None:
            self.terminated += 1

        def kill(self) -> None:
            self.killed += 1
            self.alive = False
            self.kill_called.set()

        def join(self, timeout: float | None = None) -> None:
            self.join_timeouts.append(timeout)

        def close(self) -> None:
            self.closed = True

    class ResultPipe:
        def __init__(self) -> None:
            self.payload = None

        def send(self, payload) -> None:
            self.payload = payload

        def poll(self, timeout: float) -> bool:
            return self.payload is not None

        def recv(self):
            if self.payload is None:
                raise EOFError
            return self.payload

        def close(self) -> None:
            return None

    class FastProcess:
        exitcode = 0

        def __init__(self, target, args, name, daemon) -> None:
            self.target = target
            self.args = args
            self.name = name
            self.daemon = daemon

        def start(self) -> None:
            self.target(*self.args)

        def join(self, timeout: float | None = None) -> None:
            return None

        def is_alive(self) -> bool:
            return False

        def close(self) -> None:
            return None

    class SpawnContext:
        def Pipe(self, duplex: bool) -> tuple[ResultPipe, ResultPipe]:
            assert duplex is False
            pipe = ResultPipe()
            return pipe, pipe

        def Process(self, **kwargs) -> FastProcess:
            return FastProcess(**kwargs)

    def fast_sync_worker(profile_payload, db_path, reference_markup_rate, trigger, result_sender) -> None:
        result_sender.send({"ok": True, "sync": {"ok": True, "target_id": profile_payload["id"], "rebuilt": False}})

    prior = PriorProcess()
    with state._lock:
        state._background_process = prior
        state._background_process_generation = 7
    monkeypatch.setattr("invoice_hub.services.app_state.multiprocessing.get_context", lambda method: SpawnContext())
    monkeypatch.setattr("invoice_hub.services.app_state._run_background_sync_process", fast_sync_worker)

    started = time.monotonic()
    state.run_background_diagnostics("startup_sync")
    assert time.monotonic() - started < 0.5
    assert prior.kill_called.wait(timeout=2)
    assert prior.terminated == 1
    assert prior.killed == 1
    assert prior.closed is True
    assert prior.join_timeouts
    assert all(timeout is not None and timeout <= 2.0 for timeout in prior.join_timeouts)


def test_background_retire_timeout_keeps_pid_snapshot_after_process_close(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub

    class ProcessClosedByPeer:
        def __init__(self) -> None:
            self.alive_checks = 0
            self.closed = False

        @property
        def pid(self) -> int:
            if self.closed:
                raise ValueError("process object is closed")
            return 991

        def is_alive(self) -> bool:
            self.alive_checks += 1
            if self.alive_checks >= 3:
                self.closed = True
            return True

        def terminate(self) -> None:
            return None

        def join(self, timeout: float | None = None) -> None:
            return None

    process = ProcessClosedByPeer()
    state._retire_background_process_async(process, generation=4, reason="test_closed_process")

    deadline = time.monotonic() + 2
    events = []
    while time.monotonic() < deadline:
        events = state.wait_events(0)
        if any(event["event_type"] == "server.background_worker_retire_timeout" for event in events):
            break
        time.sleep(0.01)

    matching = [event for event in events if event["event_type"] == "server.background_worker_retire_timeout"]
    assert matching
    assert matching[-1]["payload"] == {
        "generation": 4,
        "reason": "test_closed_process",
        "pid": 991,
    }


def test_background_worker_rejecting_kill_stays_managed_without_sync_failure(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    monkeypatch.setattr("invoice_hub.services.app_state.BACKGROUND_SYNC_MAX_SECONDS", 0.01)
    monkeypatch.setattr("invoice_hub.services.app_state.BACKGROUND_SYNC_TERMINATE_JOIN_SECONDS", 0.0)
    monkeypatch.setattr("invoice_hub.services.app_state.BACKGROUND_SYNC_KILL_JOIN_SECONDS", 0.0)

    class ResultPipe:
        def __init__(self) -> None:
            self.payload = None

        def send(self, payload) -> None:
            self.payload = payload

        def poll(self, timeout: float) -> bool:
            return self.payload is not None

        def recv(self):
            if self.payload is None:
                raise EOFError
            return self.payload

        def close(self) -> None:
            return None

    class RefusingProcess:
        exitcode = None
        pid = 991

        def __init__(self, target, args, name, daemon) -> None:
            self.target = target
            self.args = args
            self.name = name
            self.daemon = daemon
            self.terminate_calls = 0
            self.kill_calls = 0

        def start(self) -> None:
            return None

        def is_alive(self) -> bool:
            return True

        def terminate(self) -> None:
            self.terminate_calls += 1

        def kill(self) -> None:
            self.kill_calls += 1
            raise OSError("kill rejected")

        def join(self, timeout: float | None = None) -> None:
            return None

        def close(self) -> None:
            raise AssertionError("an alive refusing worker must remain managed")

    class FastProcess:
        exitcode = 0

        def __init__(self, target, args, name, daemon) -> None:
            self.target = target
            self.args = args
            self.name = name
            self.daemon = daemon

        def start(self) -> None:
            self.target(*self.args)

        def is_alive(self) -> bool:
            return False

        def join(self, timeout: float | None = None) -> None:
            return None

        def close(self) -> None:
            return None

    created: list[object] = []

    class SpawnContext:
        def Pipe(self, duplex: bool) -> tuple[ResultPipe, ResultPipe]:
            assert duplex is False
            pipe = ResultPipe()
            return pipe, pipe

        def Process(self, **kwargs):
            process = RefusingProcess(**kwargs) if not created else FastProcess(**kwargs)
            created.append(process)
            return process

    def fast_sync_worker(profile_payload, db_path, reference_markup_rate, trigger, result_sender) -> None:
        result_sender.send({"ok": True, "sync": {"ok": True, "target_id": profile_payload["id"], "rebuilt": False}})

    monkeypatch.setattr("invoice_hub.services.app_state.multiprocessing.get_context", lambda method: SpawnContext())
    monkeypatch.setattr("invoice_hub.services.app_state._run_background_sync_process", fast_sync_worker)
    state.run_background_diagnostics("startup_sync")

    deadline = time.monotonic() + 2
    while time.monotonic() < deadline:
        events = state.wait_events(0)
        if (
            any(event["event_type"] == "server.background_worker_retire_timeout" for event in events)
            and created
            and isinstance(created[0], RefusingProcess)
            and created[0].kill_calls >= 2
        ):
            break
        time.sleep(0.01)
    event_types = {event["event_type"] for event in state.wait_events(0)}
    refusing = created[0]
    assert isinstance(refusing, RefusingProcess)
    assert "server.background_worker_retire_timeout" in event_types
    assert not {"monitor.sync_failed", "server.background_failed", "server.background_ready"}.intersection(event_types)
    assert state._background_status == "running"
    assert id(refusing) in state._retiring_background_processes

    before_retry = refusing.kill_calls
    state.run_background_diagnostics("startup_sync")
    retry_deadline = time.monotonic() + 2
    while time.monotonic() < retry_deadline and refusing.kill_calls <= before_retry:
        time.sleep(0.01)
    assert refusing.kill_calls > before_retry


def test_bridge_status_contract_exposes_monitor_lifecycle(tmp_path: Path) -> None:
    app = create_app(tmp_path)
    client = TestClient(app)

    status = client.get("/api/v1/bridge/status").json()

    for key in (
        "running",
        "ready",
        "observer_active",
        "pid",
        "lock_exists",
        "lock_path",
        "stop_file_exists",
        "stop_file_path",
        "watch_dir",
        "workspace_dir",
        "log_path",
        "sync_interval_seconds",
        "last_sync_at",
        "last_event_at",
        "last_heartbeat_at",
        "last_trigger",
        "reason",
    ):
        assert key in status
    assert status["sync_interval_seconds"] == 60



def test_bridge_open_runtime_paths_use_platform_open(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv('INVOICE_HUB_DISABLE_OPEN', '1')
    monkeypatch.setattr(
        "invoice_hub.services.app_state.AppState.run_background_diagnostics",
        lambda self, trigger="startup_sync": None,
    )
    app = create_app(tmp_path)
    client = TestClient(app)

    log_missing = client.post('/api/v1/bridge/open-log').json()

    assert log_missing['opened'] is True
    assert log_missing['ok'] is False
    assert Path(log_missing['folder_path']).exists()
    assert log_missing['file_path'].endswith('文件变化监控日志.txt')

    log_path = Path(log_missing['file_path'])
    log_path.write_text('STARTUP_SYNC ok', encoding='utf-8')
    log_existing = client.post('/api/v1/bridge/open-log').json()

    assert log_existing['ok'] is True
    assert log_existing['opened'] is True
    assert log_existing['file_name'] == '文件变化监控日志.txt'

    runtime = client.post('/api/v1/bridge/open-runtime-dir').json()

    assert runtime['ok'] is True
    assert runtime['opened'] is True
    assert Path(runtime['folder_path']).exists()

def test_manual_fields_are_persisted_to_active_summary(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setenv("INVOICE_HUB_DISABLE_OPEN", "1")
    watch = tmp_path / "发票文件"
    watch.mkdir()
    source = watch / "sample.xml"
    source.write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)
    client.post("/api/v1/bridge/rebuild")

    patch = client.patch(
        "/api/v1/invoices/0/manual-fields",
        json={"fields": {"销售方": "手工销售方", "开票金额": "123.45", "发票号码": "999"}},
    ).json()
    assert patch["ok"] is True
    detail = client.get("/api/v1/invoices/0").json()
    assert detail["invoice"]["seller"] == "手工销售方"
    assert detail["invoice"]["amount"] == "123.45"
    assert detail["invoice"]["invoice_number"] == "999"
    invoices = client.get("/api/v1/invoices").json()
    assert invoices["items"][0]["file_format"] == "xml"
    assert invoices["items"][0]["file_type"] == "xml"
    assert invoices["items"][0]["invoice_type"] != "xml"
    for key in (
        "buyer",
        "invoice_type",
        "business_type",
        "classification_status",
        "classification_issue",
        "file_format",
        "tax_amount",
        "pretax_amount",
        "tax_rate",
        "source_exists",
        "source_size_bytes",
        "source_modified_at",
    ):
        assert key in detail["invoice"]
    opened = client.post("/api/v1/invoices/0/open-file").json()
    assert opened["ok"] is True
    assert opened["file_name"] == "sample.xml"
    location = client.post("/api/v1/invoices/0/open-location").json()
    assert location["ok"] is True
    assert location["file_name"] == "sample.xml"
    assert location["folder_path"] == str(watch)


def test_manual_fields_api_waits_for_current_profile_lock_without_blocking_health(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    source = watch / "sample.xml"
    source.write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    assert state.bridge_rebuild()["ok"] is True

    holder_state = MonitorState(state.active_profile, state.layout.db_path)
    lock_entered = threading.Event()
    release_lock = threading.Event()
    update_entered = threading.Event()
    update_finished = threading.Event()
    health_finished = threading.Event()
    responses = {}
    original_update = state.update_manual_fields

    def hold_lock() -> None:
        with holder_state.sync_write_lock():
            lock_entered.set()
            release_lock.wait(timeout=10)

    def observed_update(invoice_key: str, payload: dict) -> dict:
        update_entered.set()
        return original_update(invoice_key, payload)

    def request_update(client: TestClient) -> None:
        try:
            responses["update"] = client.patch(
                "/api/v1/invoices/0/manual-fields",
                json={"fields": {"销售方": "并发手工销售方"}},
            )
        finally:
            update_finished.set()

    def request_health(client: TestClient) -> None:
        try:
            responses["health"] = client.get("/api/v1/health")
        finally:
            health_finished.set()

    monkeypatch.setattr(state, "update_manual_fields", observed_update)
    holder = threading.Thread(target=hold_lock)
    with TestClient(app) as client:
        holder.start()
        assert lock_entered.wait(timeout=2)
        update_thread = threading.Thread(target=request_update, args=(client,))
        health_thread = threading.Thread(target=request_health, args=(client,))
        update_thread.start()
        assert update_entered.wait(timeout=2)
        health_thread.start()
        try:
            health_responded_during_update = health_finished.wait(timeout=0.5)
            update_was_still_waiting = not update_finished.is_set()
        finally:
            release_lock.set()
            holder.join(timeout=2)
            update_thread.join(timeout=5)
            health_thread.join(timeout=2)

    assert health_responded_during_update
    assert update_was_still_waiting
    assert not holder.is_alive()
    assert not update_thread.is_alive()
    assert responses["health"].status_code == 200
    assert responses["update"].status_code == 200
    assert responses["update"].json()["ok"] is True
    assert state.invoice_detail("0")["invoice"]["seller"] == "并发手工销售方"


def test_manual_fields_suppresses_stale_profile_event_after_old_profile_write(tmp_path: Path, monkeypatch) -> None:
    with monkeypatch.context() as startup_patch:
        startup_patch.setattr(AppState, "run_background_diagnostics", lambda self, trigger="startup_sync": None)
        app = create_app(tmp_path)
    state = app.state.invoice_hub
    first_profile = state.active_profile.model_copy(deep=True)
    source = Path(first_profile.watch_dir) / "sample.xml"
    source.write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    assert state.bridge_rebuild()["ok"] is True
    next_watch_dir = tmp_path / "切换后的发票文件"
    next_watch_dir.mkdir()
    next_profile = target_profile_for(state.config, next_watch_dir)
    event_seq = state.repo.event_bounds()["max_seq"]
    app_state_module = importlib.import_module("invoice_hub.services.app_state")
    original_write_summary_xlsx = app_state_module.write_summary_xlsx

    def switch_profile_after_old_write(path: Path, rows: list[dict[str, str]]) -> None:
        original_write_summary_xlsx(path, rows)
        with state._lock:
            state._active_profile = next_profile

    monkeypatch.setattr(app_state_module, "write_summary_xlsx", switch_profile_after_old_write)

    result = state.update_manual_fields("0", {"fields": {"销售方": "旧目录手工销售方"}})

    assert result["ok"] is True
    assert state.active_profile.id == next_profile.id
    overrides = json.loads((Path(first_profile.state_dir) / "manual_overrides.json").read_text(encoding="utf-8"))
    assert overrides["items"][str(source)]["fields"]["销售方"] == "旧目录手工销售方"
    events = state.repo.list_events_after(event_seq)
    assert all(event["event_type"] != "invoice.manual_fields_updated" for event in events)


def test_invoice_detail_returns_cost_breakdown_by_invoice_number(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    source = watch / "10000000000000000001.xml"
    source.write_text("<invoice />", encoding="utf-8")
    write_csv_rows(Path(state.active_profile.workspace_dir) / "发票汇总.csv", SUMMARY_HEADERS, [_summary_invoice_row(source)])
    write_csv_rows(
        state.cost_service().detail_csv,
        DETAIL_HEADERS,
        [
            _cost_detail_row("10000000000000000001", source, "螺纹钢", "HRB400E 12", "2", "200", "26", "226"),
            _cost_detail_row("10000000000000000001", source, "螺纹钢", "HRB400E 12", "3", "360", "46.8", "406.8"),
            _cost_detail_row("10000000000000000001", source, "螺纹钢", "HRB500E 16", "1", "80", "10.4", "90.4"),
            _cost_detail_row("10000000000000000001", source, "盘螺", "HPB300 8", "4", "400", "52", "452"),
            _cost_detail_row("00000000000000000000", "other.xml", "螺纹钢", "HRB400E 12", "9", "900", "117", "1017"),
        ],
    )
    client = TestClient(app)

    breakdown = client.get("/api/v1/invoices/0").json()["cost_breakdown"]

    assert breakdown["available"] is True
    assert breakdown["match_strategy"] == "invoice_number"
    assert breakdown["detail_count"] == 4
    projects = {project["project_name"]: project for project in breakdown["projects"]}
    assert projects["螺纹钢"]["quantity_total"] == 6.0
    assert projects["螺纹钢"]["amount_pretax_total"] == 640.0
    assert projects["螺纹钢"]["total_with_tax"] == 723.2
    assert projects["盘螺"]["quantity_total"] == 4.0
    spec = {item["specification"]: item for item in projects["螺纹钢"]["specs"]}["HRB400E 12"]
    assert spec["quantity_total"] == 5.0
    assert spec["arithmetic_average_unit_price_pretax"] == 110.0
    assert spec["arithmetic_average_unit_price_with_tax"] == 124.3
    assert spec["weighted_average_unit_price_pretax"] == 112.0
    assert spec["weighted_average_unit_price_with_tax"] == 126.56


def test_invoice_detail_cost_breakdown_falls_back_to_source_file(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    source = watch / "manual-number.xml"
    source.write_text("<invoice />", encoding="utf-8")
    write_csv_rows(Path(state.active_profile.workspace_dir) / "发票汇总.csv", SUMMARY_HEADERS, [_summary_invoice_row(source, invoice_number="999")])
    write_csv_rows(
        state.cost_service().detail_csv,
        DETAIL_HEADERS,
        [_cost_detail_row("10000000000000000001", source.name, "盘螺", "HPB300 10", "1.5", "150", "19.5", "169.5")],
    )
    client = TestClient(app)

    breakdown = client.get("/api/v1/invoices/0").json()["cost_breakdown"]

    assert breakdown["available"] is True
    assert breakdown["match_strategy"] == "source_file"
    assert breakdown["invoice_number"] == "999"
    assert breakdown["detail_count"] == 1
    assert breakdown["projects"][0]["project_name"] == "盘螺"
    assert breakdown["projects"][0]["quantity_total"] == 1.5


def test_invoice_detail_cost_breakdown_empty_when_cost_detail_missing(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    source = Path(state.active_profile.watch_dir) / "no-cost.xml"
    source.write_text("<invoice />", encoding="utf-8")
    write_csv_rows(Path(state.active_profile.workspace_dir) / "发票汇总.csv", SUMMARY_HEADERS, [_summary_invoice_row(source)])
    client = TestClient(app)

    breakdown = client.get("/api/v1/invoices/0").json()["cost_breakdown"]

    assert breakdown == {
        "available": False,
        "match_strategy": "none",
        "invoice_number": "10000000000000000001",
        "source_file": "no-cost.xml",
        "detail_count": 0,
        "projects": [],
    }


def test_invoice_selection_summary_deduplicates_conflicts_and_splits_tax_rates(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    sources = [
        watch / "family-a.xml",
        watch / "family-a.pdf",
        watch / "family-b.xml",
        watch / "family-b.ofd",
    ]
    for source in sources:
        source.write_text("<invoice />", encoding="utf-8")

    family_a_number = "10000000000000000001"
    family_b_number = "10000000000000000002"
    rows = [
        _summary_invoice_row(sources[0], family_a_number),
        _summary_invoice_row(sources[1], family_a_number),
        _summary_invoice_row(sources[2], family_b_number),
        _summary_invoice_row(sources[3], family_b_number),
    ]
    for row in rows[:2]:
        row.update({"除税价": "-100.00", "税金": "-13.00", "开票金额": "-113.00"})
    rows[2].update({"除税价": "200.00", "税金": "", "开票金额": "226.00"})
    rows[3].update({"除税价": "201.00", "税金": "", "开票金额": "226.00"})
    write_csv_rows(Path(state.active_profile.workspace_dir) / "发票汇总.csv", SUMMARY_HEADERS, rows)

    cost_rows = [
        _cost_detail_row(family_a_number, sources[0], "钢筋", "HRB400E 12", "2", "200", "26", "226", tax_rate="0.13"),
        _cost_detail_row(family_a_number, sources[1], "钢筋", "HRB400E 12", "1", "130", "16.9", "146.9", tax_rate="13%"),
        _cost_detail_row(family_b_number, sources[2], "钢筋", "HRB400E 12", "1", "100", "9", "109", tax_rate="9%"),
        _cost_detail_row(family_b_number, sources[3], "钢筋", "HRB400E 14", "1", "80", "0", "80", tax_rate=""),
    ]
    write_csv_rows(state.cost_service().detail_csv, DETAIL_HEADERS, cost_rows)
    client = TestClient(app)
    invoice_items = client.get("/api/v1/invoices").json()["items"]

    response = client.post(
        "/api/v1/invoices/selection-summary",
        json={
            "items": [
                {"invoice_key": item["invoice_key"], "source_path": item["source_path"]}
                for item in invoice_items
            ]
        },
    )

    assert response.status_code == 200
    payload = response.json()
    assert payload["selection"] == {
        "record_count": 4,
        "invoice_count": 2,
        "collapsed_record_count": 2,
    }
    assert payload["totals"]["pretax_amount"] == {
        "value": -100.0,
        "valid_invoice_count": 1,
        "missing_invoice_count": 0,
        "conflict_invoice_count": 1,
    }
    assert payload["totals"]["tax_amount"] == {
        "value": -13.0,
        "valid_invoice_count": 1,
        "missing_invoice_count": 1,
        "conflict_invoice_count": 0,
    }
    assert payload["totals"]["total_with_tax"] == {
        "value": 113.0,
        "valid_invoice_count": 2,
        "missing_invoice_count": 0,
        "conflict_invoice_count": 0,
    }

    breakdown = payload["cost_breakdown"]
    assert breakdown["available"] is True
    assert breakdown["matched_invoice_count"] == 2
    assert breakdown["unmatched_invoice_count"] == 0
    assert breakdown["detail_count"] == 4
    assert breakdown["match_strategy_counts"] == {"invoice_number": 2, "source_file": 0}
    assert [project["display_tax_rate"] for project in breakdown["projects"]] == ["13%", "9%", "税率未识别"]
    normalized_project = breakdown["projects"][0]
    assert normalized_project["display_project_name"] == "钢筋"
    assert normalized_project["quantity_total"] == 3.0
    assert normalized_project["amount_pretax_total"] == 330.0
    assert normalized_project["total_with_tax"] == 372.9
    assert normalized_project["specs"][0]["arithmetic_average_unit_price_pretax"] == 115.0
    assert normalized_project["specs"][0]["weighted_average_unit_price_pretax"] == 110.0


def test_invoice_selection_summary_does_not_block_health_while_sync_work_is_running(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    selection_started = threading.Event()
    release_selection = threading.Event()
    selection_finished = threading.Event()
    health_finished = threading.Event()
    responses = {}

    def blocking_selection_summary(payload: dict) -> dict:
        selection_started.set()
        if not release_selection.wait(timeout=5):
            raise TimeoutError("test did not release selection summary")
        return {"ok": True, "items": payload.get("items", [])}

    monkeypatch.setattr(state, "invoice_selection_summary", blocking_selection_summary)

    with TestClient(app) as client:
        def request_selection_summary() -> None:
            try:
                responses["selection"] = client.post(
                    "/api/v1/invoices/selection-summary",
                    json={"items": [{"invoice_key": "0", "source_path": "mock.xml"}]},
                )
            finally:
                selection_finished.set()

        def request_health() -> None:
            try:
                responses["health"] = client.get("/api/v1/health")
            finally:
                health_finished.set()

        selection_thread = threading.Thread(target=request_selection_summary)
        health_thread = threading.Thread(target=request_health)
        selection_thread.start()
        assert selection_started.wait(timeout=2)
        health_thread.start()
        try:
            health_responded_during_selection = health_finished.wait(timeout=0.5)
            selection_was_still_running = not selection_finished.is_set()
        finally:
            release_selection.set()
            selection_thread.join(timeout=2)
            health_thread.join(timeout=2)

    assert health_responded_during_selection
    assert selection_was_still_running
    assert responses["health"].status_code == 200
    assert responses["health"].json()["ok"] is True
    assert responses["selection"].status_code == 200


def test_invoice_selection_summary_uses_filename_and_source_fallbacks(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    watch = Path(state.active_profile.watch_dir)
    filename_number = "10000000000000000003"
    filename_sources = [
        watch / f"{filename_number}.pdf",
        watch / f"copy-{filename_number}.ofd",
    ]
    manual_source = watch / "manual-number.xml"
    missing_source = watch / "no-cost.xml"
    for source in [*filename_sources, manual_source, missing_source]:
        source.write_text("<invoice />", encoding="utf-8")

    filename_rows = [
        _summary_invoice_row(filename_sources[0], invoice_number=filename_number),
        _summary_invoice_row(filename_sources[1], invoice_number=""),
    ]
    manual_row = _summary_invoice_row(manual_source, invoice_number="999")
    missing_row = _summary_invoice_row(missing_source, invoice_number="")
    write_csv_rows(
        Path(state.active_profile.workspace_dir) / "发票汇总.csv",
        SUMMARY_HEADERS,
        [*filename_rows, manual_row, missing_row],
    )
    write_csv_rows(
        state.cost_service().detail_csv,
        DETAIL_HEADERS,
        [
            _cost_detail_row(filename_number, filename_sources[0].name, "水泥", "P.O 42.5", "2", "100", "13", "113"),
            _cost_detail_row("10000000000000000012", manual_source.name, "砂石", "中砂", "3", "90", "11.7", "101.7"),
        ],
    )
    client = TestClient(app)
    invoice_items = client.get("/api/v1/invoices").json()["items"]
    selected = [
        {"invoice_key": item["invoice_key"], "source_path": item["source_path"]}
        for item in invoice_items
    ]

    response = client.post("/api/v1/invoices/selection-summary", json={"items": selected})

    assert response.status_code == 200
    payload = response.json()
    assert payload["selection"] == {
        "record_count": 4,
        "invoice_count": 3,
        "collapsed_record_count": 1,
    }
    breakdown = payload["cost_breakdown"]
    assert breakdown["matched_invoice_count"] == 2
    assert breakdown["unmatched_invoice_count"] == 1
    assert breakdown["detail_count"] == 2
    assert breakdown["match_strategy_counts"] == {"invoice_number": 1, "source_file": 1}
    assert {project["project_name"] for project in breakdown["projects"]} == {"水泥", "砂石"}

    only_missing = client.post(
        "/api/v1/invoices/selection-summary",
        json={"items": [selected[-1]]},
    ).json()
    assert only_missing["cost_breakdown"] == {
        "available": False,
        "matched_invoice_count": 0,
        "unmatched_invoice_count": 1,
        "detail_count": 0,
        "match_strategy_counts": {"invoice_number": 0, "source_file": 0},
        "projects": [],
    }


def test_invoice_selection_summary_rejects_invalid_and_stale_requests(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr("invoice_hub.services.app_state.AppState.run_background_diagnostics", lambda self, trigger="startup_sync": None)
    app = create_app(tmp_path)
    state = app.state.invoice_hub
    source = Path(state.active_profile.watch_dir) / "selected.xml"
    source.write_text("<invoice />", encoding="utf-8")
    write_csv_rows(
        Path(state.active_profile.workspace_dir) / "发票汇总.csv",
        SUMMARY_HEADERS,
        [_summary_invoice_row(source)],
    )
    client = TestClient(app)
    item = client.get("/api/v1/invoices").json()["items"][0]
    selected = {"invoice_key": item["invoice_key"], "source_path": item["source_path"]}

    malformed_json = client.post(
        "/api/v1/invoices/selection-summary",
        content=b"{",
        headers={"Content-Type": "application/json"},
    )
    assert malformed_json.status_code == 400
    assert "合法 JSON" in malformed_json.json()["detail"]

    for invalid_payload in (
        [],
        {},
        {"items": []},
        {"items": ["not-an-object"]},
        {"items": [{"invoice_key": item["invoice_key"], "source_path": ""}]},
        {"items": [selected, selected]},
    ):
        response = client.post("/api/v1/invoices/selection-summary", json=invalid_payload)
        assert response.status_code == 400

    stale_path = client.post(
        "/api/v1/invoices/selection-summary",
        json={"items": [{**selected, "source_path": str(source.with_name("different.xml"))}]},
    )
    assert stale_path.status_code == 409
    assert "已过期" in stale_path.json()["detail"]

    stale_key = client.post(
        "/api/v1/invoices/selection-summary",
        json={"items": [{"invoice_key": "9999", "source_path": item["source_path"]}]},
    )
    assert stale_key.status_code == 409


def test_native_picker_contract_can_be_mocked(tmp_path: Path, monkeypatch) -> None:
    selected = tmp_path / "外部 发票目录"
    selected.mkdir()
    monkeypatch.setenv("INVOICE_HUB_DIALOG_MOCK_PATH", str(selected))
    app = create_app(tmp_path)
    client = TestClient(app)

    payload = client.post("/api/v1/settings/pick-watch-dir").json()
    assert payload["ok"] is True
    assert payload["selected"] is True
    assert payload["validation"]["can_monitor"] is True


def test_settings_rename_invoice_files_keeps_manual_fields_and_rebuilds(tmp_path: Path) -> None:
    watch = tmp_path / "\u53d1\u7968\u6587\u4ef6"
    watch.mkdir()
    source = watch / "sample.xml"
    source.write_text((Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8"), encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    assert client.post("/api/v1/bridge/rebuild").json()["ok"] is True
    patch = client.patch(
        "/api/v1/invoices/0/manual-fields",
        json={"fields": {"\u9500\u552e\u65b9": "Manual Seller", "\u5f00\u7968\u91d1\u989d": "123.45"}},
    ).json()
    assert patch["ok"] is True

    result = client.post("/api/v1/settings/rename-invoice-files", json={}).json()

    assert result["ok"] is True
    assert result["format"] == "YY-MM-DD_\u9500\u552e\u65b9&\u8d2d\u4e70\u65b9_\u91d1\u989d\u5143.ext"
    assert result["supported_extensions"] == ["ofd", "pdf", "xml"]
    assert result["renamed"] == 1
    assert result["manual_overrides_migrated"] == 1
    renamed = next(item for item in result["files"] if item["status"] == "renamed")
    renamed_path = Path(renamed["new_path"])
    assert renamed_path.exists()
    assert not source.exists()
    assert "_Manual Seller&" in renamed_path.name
    assert renamed_path.name.endswith("_123.45\u5143.xml")

    invoices = client.get("/api/v1/invoices").json()
    assert invoices["count"] == 1
    assert invoices["items"][0]["source_file"] == renamed_path.name
    assert invoices["items"][0]["seller"] == "Manual Seller"
    assert invoices["items"][0]["amount"] == "123.45"

    overrides = json.loads((Path(app.state.invoice_hub.active_profile.state_dir) / "manual_overrides.json").read_text(encoding="utf-8"))
    assert str(source) not in overrides["items"]
    assert str(renamed_path) in overrides["items"]


def test_settings_rename_invoice_files_skips_duplicate_target_names(tmp_path: Path) -> None:
    watch = tmp_path / "\u53d1\u7968\u6587\u4ef6"
    watch.mkdir()
    fixture = (Path(__file__).parent / "fixtures" / "sample_invoice.xml").read_text(encoding="utf-8")
    first = watch / "first.xml"
    second = watch / "second.xml"
    first.write_text(fixture, encoding="utf-8")
    second.write_text(fixture, encoding="utf-8")
    app = create_app(tmp_path)
    client = TestClient(app)

    result = client.post("/api/v1/settings/rename-invoice-files", json={}).json()

    assert result["ok"] is True
    assert result["renamed"] == 0
    assert result["skipped"] == 2
    assert result["skipped_by_reason"]["duplicate_target_name"] == 2
    assert first.exists()
    assert second.exists()


def test_settings_update_persists_watch_dir(tmp_path: Path) -> None:
    selected = tmp_path / "外部 发票目录"
    selected.mkdir()
    app = create_app(tmp_path)
    client = TestClient(app)

    payload = client.put("/api/v1/settings", json={"watch_dir": str(selected)}).json()
    assert payload["ok"] is True
    assert payload["watch_dir"] == str(selected)

    restarted = TestClient(create_app(tmp_path)).get("/api/v1/settings").json()
    assert restarted["watch_dir"] == str(selected)
    assert restarted["recent_watch_dirs"][0] == str(selected)


def test_settings_update_serializes_project_internal_watch_dir_as_relative(tmp_path: Path) -> None:
    selected = tmp_path / "发票文件"
    selected.mkdir()
    app = create_app(tmp_path)
    client = TestClient(app)

    payload = client.put("/api/v1/settings", json={"watch_dir": str(selected)}).json()

    assert payload["ok"] is True
    assert payload["watch_dir"] == str(selected)
    saved = json.loads((tmp_path / "config" / "app.local.json").read_text(encoding="utf-8-sig"))
    assert saved["watch_dir"] == "./发票文件"
    assert str(tmp_path) not in saved["watch_dir"]


def test_settings_recent_watch_dir_remove_persists_and_keeps_current(tmp_path: Path) -> None:
    first = tmp_path / "一号 发票目录"
    second = tmp_path / "二号 发票目录"
    current = tmp_path / "当前 发票目录"
    for path in (first, second, current):
        path.mkdir()
    app = create_app(tmp_path)
    client = TestClient(app)

    for path in (first, second, current):
        assert client.put("/api/v1/settings", json={"watch_dir": str(path)}).json()["ok"] is True

    before = client.get("/api/v1/settings").json()
    assert str(current) == before["watch_dir"]
    assert str(second) in before["recent_watch_dirs"]

    removed = client.post("/api/v1/settings/recent-watch-dirs/remove", json={"watch_dir": str(second)}).json()
    assert removed["ok"] is True
    assert removed["watch_dir"] == str(current)
    assert str(current) in removed["recent_watch_dirs"]
    assert str(second) not in removed["recent_watch_dirs"]

    saved = json.loads((tmp_path / "config" / "app.local.json").read_text(encoding="utf-8-sig"))
    assert "./二号 发票目录" not in saved["recent_watch_dirs"]
    assert "./当前 发票目录" in saved["recent_watch_dirs"]

    after_current_remove = client.post("/api/v1/settings/recent-watch-dirs/remove", json={"watch_dir": str(current)}).json()
    assert after_current_remove["watch_dir"] == str(current)
    assert after_current_remove["recent_watch_dirs"][0] == str(current)

    restarted = TestClient(create_app(tmp_path)).get("/api/v1/settings").json()
    assert restarted["watch_dir"] == str(current)
    assert str(second) not in restarted["recent_watch_dirs"]


def test_consistency_report_groups_same_invoice_formats(tmp_path: Path) -> None:
    watch = tmp_path / "发票文件"
    watch.mkdir()
    workspace = tmp_path / "运行状态" / "targets" / "manual" / "workspace"
    workspace.mkdir(parents=True)
    source_pdf = watch / "dzfp_10000000000000000001_vendor.pdf"
    source_xml = watch / "dzfp_10000000000000000001.xml"
    source_pdf.write_text("pdf", encoding="utf-8")
    source_xml.write_text("xml", encoding="utf-8")
    rows = [
        {
            "文件名": source_pdf.name,
            "文件路径": str(source_pdf),
            "发票类型": "增值税专用发票",
            "特定业务类型": "建筑服务",
            "类型识别状态": "ok",
            "类型识别说明": "",
            "发票号码": "10000000000000000001",
            "开票时间": "2026-04-01",
            "销售方": "供应商A",
            "购买方": "购买方A",
            "开票金额": "100.00",
            "税率": "13%",
            "除税价": "88.50",
            "税金": "11.50",
            "重复发票": "",
            "手改状态": "",
        },
        {
            "文件名": source_xml.name,
            "文件路径": str(source_xml),
            "发票类型": "增值税普通发票",
            "特定业务类型": "标准电子发票",
            "类型识别状态": "ok",
            "类型识别说明": "",
            "发票号码": "10000000000000000001",
            "开票时间": "2026-04-01",
            "销售方": "供应商A",
            "购买方": "购买方A",
            "开票金额": "101.00",
            "税率": "13%",
            "除税价": "89.38",
            "税金": "11.62",
            "重复发票": "",
            "手改状态": "",
        },
    ]
    from invoice_hub.storage import write_csv_rows
    from invoice_hub.projections.summary import SUMMARY_HEADERS

    write_csv_rows(workspace / "发票汇总.csv", SUMMARY_HEADERS, rows)
    app = create_app(tmp_path)
    app.state.invoice_hub._active_profile = app.state.invoice_hub.active_profile.model_copy(update={"workspace_dir": str(workspace)})
    client = TestClient(app)

    report = client.get("/api/v1/consistency-report").json()
    assert report["stats"]["total_groups"] == 1
    assert report["stats"]["inconsistent_groups"] == 1
    assert report["groups"][0]["formats"] == ["pdf", "xml"]
    mismatch_fields = [item["field"] for item in report["groups"][0]["mismatch_fields"]]
    assert "发票大类" in mismatch_fields
    assert "特定业务类型" in mismatch_fields
    assert "开票金额" in mismatch_fields

    detail = client.get("/api/v1/invoices/0").json()
    assert detail["consistency"]["pair_key"] == "invoice:10000000000000000001"
    assert detail["consistency"]["consistent"] is False
