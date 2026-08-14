from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook

from invoice_hub.domain import InvoiceRecord
from invoice_hub.extraction import apply_invoice_family_corrections, extract_invoice_record, supported_invoice_files
from invoice_hub.storage.files import write_csv_rows

SUMMARY_HEADERS = [
    "文件名",
    "文件路径",
    "发票类型",
    "特定业务类型",
    "类型识别状态",
    "类型识别说明",
    "发票号码",
    "开票时间",
    "销售方",
    "购买方",
    "开票金额",
    "税率",
    "除税价",
    "税金",
    "重复发票",
    "手改状态",
]


def _row(record: InvoiceRecord) -> dict[str, str]:
    return {
        "文件名": record.source_file,
        "文件路径": record.source_path,
        "发票类型": record.invoice_type,
        "特定业务类型": record.business_type,
        "类型识别状态": record.classification_status,
        "类型识别说明": record.classification_issue,
        "发票号码": record.invoice_number,
        "开票时间": record.invoice_date,
        "销售方": record.seller,
        "购买方": record.buyer,
        "开票金额": record.amount,
        "税率": record.tax_rate,
        "除税价": record.pretax_amount,
        "税金": record.tax_amount,
        "重复发票": record.duplicate_label,
        "手改状态": "",
    }


def summary_schema_needs_refresh(csv_path: Path, xlsx_path: Path) -> bool:
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)
    if not csv_path.exists() or not xlsx_path.exists():
        return True
    try:
        with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
            csv_headers = [str(value or "").strip() for value in next(csv.reader(handle), [])]
    except (OSError, csv.Error):
        return True
    if any(header not in csv_headers for header in SUMMARY_HEADERS):
        return True
    workbook = None
    try:
        workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
        sheet = workbook.active
        xlsx_headers = [
            str(cell or "").strip()
            for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        ]
    except Exception:
        return True
    finally:
        if workbook is not None:
            workbook.close()
    return any(header not in xlsx_headers for header in SUMMARY_HEADERS)


def mark_duplicates(records: list[InvoiceRecord]) -> list[InvoiceRecord]:
    seen: defaultdict[str, int] = defaultdict(int)
    marked: list[InvoiceRecord] = []
    for record in records:
        number = (record.invoice_number or "").strip()
        duplicate = bool(number and seen[number] > 0)
        seen[number] += 1
        marked.append(record.model_copy(update={"duplicate": duplicate, "duplicate_label": "重复发票" if duplicate else ""}))
    return marked


def write_summary_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "发票汇总"
    ws.append(SUMMARY_HEADERS)
    for row in rows:
        ws.append([row.get(header, "") for header in SUMMARY_HEADERS])
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(max_len + 2, 10), 50)
    wb.save(path)


def build_summary(watch_dir: Path, workspace_dir: Path) -> dict:
    files = supported_invoice_files(watch_dir)
    records = [extract_invoice_record(path) for path in files]
    records = apply_invoice_family_corrections(records, files)
    records = mark_duplicates(records)
    rows = [_row(record) for record in records]
    csv_path = workspace_dir / "发票汇总.csv"
    xlsx_path = workspace_dir / "发票汇总.xlsx"
    write_csv_rows(csv_path, SUMMARY_HEADERS, rows)
    write_summary_xlsx(xlsx_path, rows)
    return {
        "ok": True,
        "count": len(rows),
        "summary_csv": str(csv_path),
        "summary_xlsx": str(xlsx_path),
        "records": [record.model_dump() for record in records],
    }
