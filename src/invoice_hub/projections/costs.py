from __future__ import annotations

import hashlib
import os
import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from invoice_hub.domain import CostAnalysisSnapshot, CostSyncStatus
from invoice_hub.domain.models import utc_now_text
from invoice_hub.projections.cost_analysis import (
    CHECK_HEADERS,
    COST_DETAIL_CSV_NAME,
    COST_REFERENCE_STATUS_NAME,
    COST_SUMMARY_XLSX_NAME,
    DEFAULT_REFERENCE_MARKUP_RATE,
    DETAIL_HEADERS,
    INVOICE_REFERENCE_HEADERS,
    PROJECT_SPEC_SUMMARY_HEADERS,
    REFERENCE_OUTPUT_TAX_RATE,
    average_unit_price,
    detail_rows_with_tax_average,
    build_cost_analysis_outputs,
    normalize_reference_markup_percent,
    normalize_reference_markup_rate,
    project_spec_summary,
    reference_output_average_unit_price_with_tax,
    reference_markup_display,
    reference_markup_multiplier,
    reference_markup_percent_text,
    average_unit_price_with_tax,
    purchase_reference_average_unit_price_with_tax,
    _write_workbook,
)
from invoice_hub.storage.files import atomic_write_json, read_csv_rows, read_json_object, write_csv_rows

COST_SOURCE_EXTS = {".pdf", ".ofd", ".xml"}


def _decimal(value: object) -> Decimal:
    text = str(value or "").replace(",", "").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def _user_quantity(value: object) -> Decimal:
    text = str(value or "").strip()
    if not text:
        return Decimal("0")
    if not re.fullmatch(r"(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)", text):
        raise ValueError(f"已开数量不是有效数字: {value}")
    try:
        quantity = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"已开数量不是有效数字: {value}") from exc
    if not quantity.is_finite():
        raise ValueError(f"已开数量不是有效数字: {value}")
    return quantity


def _decimal_text(value: Decimal, places: str = "0.001") -> str:
    rounded = value.quantize(Decimal(places))
    text = format(rounded, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def _float_number(value: Decimal, places: str = "0.01") -> float:
    return float(value.quantize(Decimal(places)))


def _reference_key_values(*values: object) -> list[str]:
    return [re.sub(r"\s+", " ", str(value or "").replace("\u00a0", " ").strip()).strip() for value in values]


def _legacy_plain_reference_key(*values: object) -> str:
    return "|".join(str(value or "").strip() for value in values)


def _reference_key_from_values(*values: object) -> str:
    raw = "\x1f".join(_reference_key_values(*values))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _reference_key(row: dict[str, Any]) -> str:
    return _reference_key_from_values(
        row.get("发票代码(**内文字)", ""),
        row.get("内部项目名称", ""),
        row.get("规格型号", ""),
        row.get("单位", ""),
    )


def _reference_aliases(row: dict[str, Any]) -> list[str]:
    values = [
        row.get("发票代码(**内文字)", ""),
        row.get("内部项目名称", ""),
        row.get("规格型号", ""),
        row.get("单位", ""),
    ]
    aliases = [
        _reference_key_from_values(*values),
        _legacy_plain_reference_key(*values),
        "|".join(_reference_key_values(*values)),
    ]
    result = []
    for alias in aliases:
        if alias and alias not in result:
            result.append(alias)
    return result


def _normalize_status_item(raw_key: str, item: dict[str, Any]) -> dict[str, Any]:
    normalized = dict(item)
    normalized_key = str(normalized.get("reference_key") or normalized.get("key") or raw_key or "").strip()
    if normalized_key:
        normalized.setdefault("reference_key", normalized_key)
    field_aliases = {
        "locked_reference_amount": "invoiced_reference_amount",
        "locked_reference_tax_amount": "invoiced_reference_tax",
        "locked_reference_total_with_tax": "invoiced_reference_total_with_tax",
    }
    for old_field, new_field in field_aliases.items():
        if str(normalized.get(new_field) or "").strip() == "" and str(normalized.get(old_field) or "").strip() != "":
            normalized[new_field] = normalized[old_field]
    if str(normalized.get("reference_markup_rate") or "").strip() == "" and str(normalized.get("markup_rate") or "").strip() != "":
        normalized["reference_markup_rate"] = normalized["markup_rate"]
    return normalized


def _read_reference_status(path: Path) -> dict[str, dict]:
    payload = _read_reference_status_payload(path)
    items = payload.get("items", payload)
    if not isinstance(items, dict):
        return {}
    normalized: dict[str, dict] = {}
    for key, item in items.items():
        if not isinstance(item, dict):
            continue
        raw_key = str(key or "").strip()
        normalized_key = str(item.get("reference_key") or item.get("key") or raw_key).strip()
        if not normalized_key:
            continue
        normalized[normalized_key] = _normalize_status_item(normalized_key, item)
    return normalized


def _read_reference_status_payload(path: Path) -> dict:
    payload = read_json_object(path, {})
    return payload if isinstance(payload, dict) else {}


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value or "").strip().casefold() in {"1", "true", "yes", "y", "locked", "已锁定", "锁定"}


def _has_text(value: object) -> bool:
    return str(value or "").strip() != ""


def _row_markup_meta(status: dict[str, Any], fallback_rate: object, fallback_locked: bool = False) -> dict[str, Any]:
    rate = normalize_reference_markup_rate(fallback_rate)
    if _has_text(status.get("reference_markup_rate_percent")):
        rate = normalize_reference_markup_percent(status.get("reference_markup_rate_percent"))
    elif _has_text(status.get("reference_markup_rate")):
        rate = normalize_reference_markup_rate(status.get("reference_markup_rate"), fallback_rate)
    elif _has_text(status.get("markup_rate")):
        rate = normalize_reference_markup_rate(status.get("markup_rate"), fallback_rate)
    locked = _truthy(status.get("reference_markup_locked", fallback_locked))
    return {
        "rate": rate,
        "display": reference_markup_display(rate),
        "percent": reference_markup_percent_text(rate),
        "locked": locked,
    }


def _write_reference_status(
    path: Path,
    items: dict[str, dict],
    reference_markup_rate: object = DEFAULT_REFERENCE_MARKUP_RATE,
    reference_markup_locked: bool = False,
) -> None:
    markup_rate = normalize_reference_markup_rate(reference_markup_rate)
    normalized = {key: dict(value) for key, value in sorted(items.items()) if key and isinstance(value, dict)}
    atomic_write_json(
        path,
        {
            "version": 1,
            "updated_at": utc_now_text(),
            "reference_markup_rate": _decimal_text(markup_rate, "0.000001"),
            "reference_markup_rate_percent": reference_markup_percent_text(markup_rate),
            "reference_markup_locked": bool(reference_markup_locked),
            "items": normalized,
        },
    )


def _valid_invoice_number(value: object) -> str:
    match = re.search(r"(?<!\d)(\d{8,20})(?!\d)", str(value or ""))
    return match.group(1) if match else ""


def _cost_source_groups_from_summary(summary_csv: Path) -> dict[str, set[str]]:
    groups: dict[str, set[str]] = {}
    for row in read_csv_rows(summary_csv):
        source = row.get("文件路径") or row.get("source_path") or row.get("源文件") or ""
        if Path(str(source)).suffix.lower() not in COST_SOURCE_EXTS:
            continue
        invoice_number = _valid_invoice_number(row.get("发票号码") or row.get("invoice_number"))
        group_key = f"invoice:{invoice_number}" if invoice_number else f"source:{Path(str(source))}"
        keys = groups.setdefault(group_key, set())
        keys.update(_source_keys(str(source)))
        if invoice_number:
            keys.add(f"invoice:{invoice_number}")
    return groups


def _source_keys(value: str) -> set[str]:
    text = str(value or "").strip()
    if not text:
        return set()
    path = Path(text)
    keys = {str(path)}
    if path.name:
        keys.add(path.name.casefold())
    return keys


def _sources_from_rows(rows: list[dict]) -> set[str]:
    sources = set()
    for row in rows:
        source = row.get("源文件") or row.get("source_file") or ""
        if source:
            sources.add(str(Path(str(source))))
    return sources


def _keys_from_rows(rows: list[dict]) -> set[str]:
    keys: set[str] = set()
    for row in rows:
        source = row.get("源文件") or row.get("source_file") or row.get("源文件名") or ""
        keys.update(_source_keys(str(source)))
        invoice_number = _valid_invoice_number(row.get("发票号码") or row.get("invoice_number"))
        if invoice_number:
            keys.add(f"invoice:{invoice_number}")
    return keys


class CostProjectionService:
    def __init__(self, watch_dir: Path, workspace_dir: Path, target_id: str, reference_markup_rate: str = "0.08"):
        self.watch_dir = Path(watch_dir)
        self.workspace_dir = Path(workspace_dir)
        self.target_id = target_id
        self.reference_markup_rate = reference_markup_rate
        self.detail_csv = self.watch_dir / COST_DETAIL_CSV_NAME
        self.summary_xlsx = self.watch_dir / COST_SUMMARY_XLSX_NAME
        self.status_json = self.watch_dir / COST_REFERENCE_STATUS_NAME
        self.summary_csv = self.workspace_dir / "发票汇总.csv"

    def rebuild(self) -> dict:
        result = build_cost_analysis_outputs(
            self.watch_dir,
            self.watch_dir,
            invoice_metadata=self._summary_metadata(),
            reference_markup_rate=self._markup_meta()["rate"],
        )
        return result

    def _markup_meta(self) -> dict[str, Any]:
        payload = _read_reference_status_payload(self.status_json)
        rate = normalize_reference_markup_rate(payload.get("reference_markup_rate", self.reference_markup_rate))
        return {
            "rate": rate,
            "display": reference_markup_display(rate),
            "percent": reference_markup_percent_text(rate),
            "locked": _truthy(payload.get("reference_markup_locked", False)),
        }

    def _summary_metadata(self) -> list[dict]:
        rows = []
        for row in read_csv_rows(self.summary_csv):
            rows.append(
                {
                    "源文件": row.get("文件路径", ""),
                    "销售方": row.get("销售方", ""),
                    "购买方": row.get("购买方", ""),
                    "发票号码": row.get("发票号码", ""),
                    "开票日期": row.get("开票时间", ""),
                    "开票金额": row.get("开票金额", ""),
                    "除税价": row.get("除税价", ""),
                    "税金": row.get("税金", ""),
                    "发票大类": row.get("发票类型", ""),
                    "特定业务类型": row.get("特定业务类型", ""),
                    "类型识别状态": row.get("类型识别状态", ""),
                    "类型识别说明": row.get("类型识别说明", ""),
                }
            )
        return rows

    def _refresh_workbook_from_current_detail(self, reference_markup_rate: object) -> None:
        if not self.detail_csv.exists():
            return
        detail_rows = read_csv_rows(self.detail_csv)
        checks = self._read_workbook_sheet("发票校验", CHECK_HEADERS)
        _write_workbook(
            self.summary_xlsx,
            detail_rows,
            checks,
            status_items=_read_reference_status(self.status_json),
            reference_markup_rate=reference_markup_rate,
        )

    def _read_csv_headers(self) -> list[str]:
        if not self.detail_csv.exists():
            return []
        try:
            with self.detail_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                first_line = handle.readline().strip("\r\n")
        except OSError:
            return []
        if not first_line:
            return []
        return [item.strip() for item in first_line.split(",")]

    def _workbook_headers(self, sheet_name: str) -> list[str]:
        if not self.summary_xlsx.exists():
            return []
        wb = None
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.summary_xlsx, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
            return [str(cell or "").strip() for cell in row]
        except Exception:
            return []
        finally:
            if wb is not None:
                wb.close()

    def needs_schema_refresh(self) -> bool:
        if self.detail_csv.exists() and any(header not in self._read_csv_headers() for header in DETAIL_HEADERS):
            return True
        if self.summary_xlsx.exists():
            required = {
                "发票明细": DETAIL_HEADERS,
                "项目规格汇总": PROJECT_SPEC_SUMMARY_HEADERS,
                "开票参考": INVOICE_REFERENCE_HEADERS,
                "发票校验": CHECK_HEADERS,
            }
            for sheet_name, headers in required.items():
                actual = self._workbook_headers(sheet_name)
                if any(header not in actual for header in headers):
                    return True
        return False

    def refresh_schema_from_current_detail(self) -> bool:
        if not self.detail_csv.exists():
            return False
        detail_rows = read_csv_rows(self.detail_csv)
        markup = self._markup_meta()
        checks = self._read_workbook_sheet("发票校验", CHECK_HEADERS)
        status_items = _read_reference_status(self.status_json)
        workbook_status = self._read_reference_status_from_workbook()
        for key, workbook_item in workbook_status.items():
            if key not in status_items:
                status_items[key] = workbook_item
                continue
            for field, value in workbook_item.items():
                if str(status_items[key].get(field) or "").strip() == "" and str(value or "").strip() != "":
                    status_items[key][field] = value

        write_csv_rows(self.detail_csv, DETAIL_HEADERS, detail_rows_with_tax_average(detail_rows))
        _write_workbook(
            self.summary_xlsx,
            detail_rows,
            checks,
            status_items=status_items,
            reference_markup_rate=markup["rate"],
        )
        return True

    def snapshot(self, from_cache: bool = False) -> CostAnalysisSnapshot:
        if self.needs_schema_refresh():
            try:
                self.refresh_schema_from_current_detail()
            except OSError:
                pass
        detail_rows = detail_rows_with_tax_average(read_csv_rows(self.detail_csv))
        checks = self._read_workbook_sheet("发票校验", CHECK_HEADERS)
        project_summary = self._project_summary_rows()
        markup = self._markup_meta()
        invoice_reference = self._build_reference_rows(detail_rows)
        sync = self._sync_status(detail_rows, checks)
        mismatch_count = sum(1 for row in checks if str(row.get("校验状态") or "").strip() != "通过")
        return CostAnalysisSnapshot(
            watch_dir=str(self.watch_dir),
            source_dir=str(self.watch_dir),
            target_id=self.target_id,
            output_detail_csv_path=str(self.detail_csv),
            output_summary_xlsx_path=str(self.summary_xlsx),
            reference_status_path=str(self.status_json),
            reference_status_exists=self.status_json.exists(),
            reference_markup_rate=markup["display"],
            detail_count=len(detail_rows),
            check_count=len(checks),
            mismatch_count=mismatch_count,
            from_cache=from_cache,
            items=detail_rows,
            project_summary=project_summary,
            invoice_reference=invoice_reference,
            checks=checks,
            reference_status_stats=self._reference_stats(invoice_reference),
            sync=sync,
        )

    def _project_summary_rows(self) -> list[dict]:
        detail_rows = read_csv_rows(self.detail_csv)
        if detail_rows:
            return project_spec_summary(detail_rows)
        return self._read_workbook_sheet("项目规格汇总", PROJECT_SPEC_SUMMARY_HEADERS)

    def save_reference_status(self, payload: dict) -> dict:
        current = _read_reference_status(self.status_json)
        fallback_markup = self._markup_meta()
        workbook_status = self._read_reference_status_from_workbook()
        reference_rows = self._build_reference_rows(read_csv_rows(self.detail_csv), current)
        by_key = {row["key"]: row for row in reference_rows}
        for item in payload.get("items", []):
            key = str(item.get("key") or "").strip()
            if not key or key not in by_key:
                continue
            row = by_key[key]
            row_with_aliases = dict(row)
            row_with_aliases["aliases"] = _reference_aliases(row)
            existing = self._status_for_reference(row_with_aliases, current, workbook_status)
            existing_markup = _row_markup_meta(existing, fallback_markup["rate"], fallback_markup["locked"])
            quantity = _decimal(row.get("quantity"))
            if "invoiced_quantity" in item:
                invoiced_quantity = _user_quantity(item.get("invoiced_quantity"))
                quantity_changed = _decimal(existing.get("invoiced_quantity")) != invoiced_quantity
            else:
                invoiced_quantity = _decimal(existing.get("invoiced_quantity"))
                quantity_changed = False
            if invoiced_quantity < 0:
                raise ValueError("已开数量不能小于 0")
            if invoiced_quantity > quantity:
                raise ValueError(f"已开数量不能大于数量合计: {_decimal_text(quantity, '0.001')}")
            if _has_text(item.get("reference_markup_rate_percent")):
                row_rate = normalize_reference_markup_percent(item.get("reference_markup_rate_percent"))
            elif _has_text(item.get("markup_rate")):
                row_rate = normalize_reference_markup_percent(item.get("markup_rate"))
            elif _has_text(item.get("reference_markup_rate")):
                row_rate = normalize_reference_markup_rate(item.get("reference_markup_rate"), fallback_markup["rate"])
            else:
                row_rate = existing_markup["rate"]
            if "reference_markup_locked" in item:
                row_locked = _truthy(item.get("reference_markup_locked"))
            elif "locked" in item:
                row_locked = _truthy(item.get("locked"))
            else:
                row_locked = existing_markup["locked"]
            row_multiplier = reference_markup_multiplier(row_rate)
            reference_average_unit_price = _decimal(row.get("average_unit_price")) * row_multiplier
            reference_amount = reference_average_unit_price * quantity
            reference_tax = reference_amount * REFERENCE_OUTPUT_TAX_RATE
            reference_total = reference_amount + reference_tax
            ratio = Decimal("0") if quantity == 0 else invoiced_quantity / quantity
            status = "已开具" if invoiced_quantity >= quantity and quantity > 0 else ("未开具" if invoiced_quantity == 0 else "部分开具")
            for alias in _reference_aliases(row):
                if alias != key:
                    current.pop(alias, None)

            def locked_part(field: str, current_total: Decimal) -> Decimal:
                if invoiced_quantity <= 0:
                    return Decimal("0")
                if quantity_changed or not _has_text(existing.get(field)):
                    return current_total * ratio
                return max(_decimal(existing.get(field)), Decimal("0"))

            keep_item = (
                invoiced_quantity > 0
                or row_rate != fallback_markup["rate"]
                or row_locked != fallback_markup["locked"]
            )
            if not keep_item:
                current.pop(key, None)
                continue
            current[key] = {
                "reference_key": key,
                "item_code": row.get("发票代码(**内文字)", ""),
                "item_name": row.get("内部项目名称", ""),
                "spec": row.get("规格型号", ""),
                "unit": row.get("单位", ""),
                "invoiced_quantity": str(invoiced_quantity),
                "quantity_total": _decimal_text(quantity, "0.001"),
                "reference_amount_total": _decimal_text(reference_amount, "0.000001"),
                "reference_tax_total": _decimal_text(reference_tax, "0.000001"),
                "reference_total_with_tax": _decimal_text(reference_total, "0.000001"),
                "reference_markup_rate": _decimal_text(row_rate, "0.000001"),
                "reference_markup_rate_percent": reference_markup_percent_text(row_rate),
                "reference_markup_locked": row_locked,
                "invoiced_reference_amount": _decimal_text(locked_part("invoiced_reference_amount", reference_amount), "0.000001"),
                "invoiced_reference_tax": _decimal_text(locked_part("invoiced_reference_tax", reference_tax), "0.000001"),
                "invoiced_reference_total_with_tax": _decimal_text(locked_part("invoiced_reference_total_with_tax", reference_total), "0.000001"),
                "invoice_status": status,
                "updated_at": utc_now_text(),
            }
        _write_reference_status(
            self.status_json,
            current,
            reference_markup_rate=fallback_markup["rate"],
            reference_markup_locked=fallback_markup["locked"],
        )
        self._refresh_workbook_from_current_detail(fallback_markup["rate"])
        return {"ok": True, "reference_status_path": str(self.status_json), "workbook_updated": self.summary_xlsx.exists()}

    def _read_workbook_sheet(self, sheet_name: str, headers: list[str]) -> list[dict]:
        if not self.summary_xlsx.exists():
            return []
        wb = None
        try:
            from openpyxl import load_workbook

            wb = load_workbook(self.summary_xlsx, data_only=True, read_only=True)
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
        except Exception:
            return []
        finally:
            if wb is not None:
                wb.close()
        if not rows:
            return []
        header_row = [str(cell or "") for cell in rows[0]]
        effective_headers = header_row if any(header_row) else headers
        result = []
        for values in rows[1:]:
            result.append(
                {
                    effective_headers[i]: values[i] if i < len(values) and values[i] is not None else ""
                    for i in range(len(effective_headers))
                }
            )
        return result

    def _build_reference_rows(self, detail_rows: list[dict], status_items: dict[str, dict] | None = None) -> list[dict]:
        status_items = status_items if status_items is not None else _read_reference_status(self.status_json)
        workbook_status = self._read_reference_status_from_workbook()
        fallback_markup = self._markup_meta()
        grouped: dict[str, dict] = {}
        for row in detail_rows:
            key = _reference_key(row)
            if key not in grouped:
                grouped[key] = {
                    "key": key,
                    "发票代码(**内文字)": row.get("发票代码(**内文字)", ""),
                    "内部项目名称": row.get("内部项目名称", ""),
                    "规格型号": row.get("规格型号", ""),
                    "单位": row.get("单位", ""),
                    "quantity": Decimal("0"),
                    "amount": Decimal("0"),
                    "tax_amount": Decimal("0"),
                    "total_with_tax": Decimal("0"),
                    "source_rows": [],
                    "aliases": _reference_aliases(row),
                }
            grouped[key]["source_rows"].append(row)
            grouped[key]["quantity"] += _decimal(row.get("数量"))
            grouped[key]["amount"] += _decimal(row.get("金额(除税)"))
            grouped[key]["tax_amount"] += _decimal(row.get("税金"))
            grouped[key]["total_with_tax"] += _decimal(row.get("价税合计"))
        result = []
        for key, row in grouped.items():
            quantity = row["quantity"]
            status = self._status_for_reference(row, status_items, workbook_status)
            markup = _row_markup_meta(status, fallback_markup["rate"], fallback_markup["locked"])
            markup_multiplier = reference_markup_multiplier(markup["rate"])
            avg_price = average_unit_price(row.get("source_rows") or [])
            avg_price_with_tax = average_unit_price_with_tax(row.get("source_rows") or [])
            purchase_avg_price_with_tax = purchase_reference_average_unit_price_with_tax(row.get("source_rows") or [])
            reference_avg_price_with_tax = reference_output_average_unit_price_with_tax(avg_price, markup["rate"])
            reference_base_amount = avg_price * quantity if avg_price is not None and quantity > 0 else Decimal("0")
            reference_amount = reference_base_amount * markup_multiplier
            reference_tax = reference_amount * REFERENCE_OUTPUT_TAX_RATE
            reference_total = reference_amount + reference_tax
            invoiced_quantity = _decimal(status.get("invoiced_quantity"))
            if invoiced_quantity > quantity:
                invoiced_quantity = quantity
            locked_amount = self._locked_reference_part(status, "invoiced_reference_amount", reference_amount, quantity, invoiced_quantity)
            locked_tax = self._locked_reference_part(status, "invoiced_reference_tax", reference_tax, quantity, invoiced_quantity)
            locked_total = self._locked_reference_part(status, "invoiced_reference_total_with_tax", reference_total, quantity, invoiced_quantity)
            invoice_status = status.get("invoice_status") or ("已开具" if quantity > 0 and invoiced_quantity >= quantity else ("未开具" if invoiced_quantity == 0 else "部分开具"))
            result.append(
                {
                    **{k: v for k, v in row.items() if not isinstance(v, Decimal) and k not in {"aliases", "source_rows"}},
                    "quantity": float(quantity),
                    "average_unit_price": float(avg_price) if avg_price is not None else 0,
                    "average_unit_price_with_tax": float(avg_price_with_tax) if avg_price_with_tax is not None else 0,
                    "purchase_reference_average_unit_price_with_tax": float(purchase_avg_price_with_tax) if purchase_avg_price_with_tax is not None else 0,
                    "amount": float(row["amount"]),
                    "tax_amount": float(row["tax_amount"]),
                    "total_with_tax": float(row["total_with_tax"]),
                    "markup_rate": markup["display"],
                    "reference_markup_rate": _decimal_text(markup["rate"], "0.000001"),
                    "reference_markup_rate_percent": markup["percent"],
                    "reference_markup_locked": markup["locked"],
                    "reference_average_unit_price": float(reference_amount / quantity) if quantity else 0,
                    "reference_average_unit_price_with_tax": float(reference_avg_price_with_tax) if reference_avg_price_with_tax is not None else 0,
                    "reference_amount": float(reference_amount),
                    "reference_tax_amount": float(reference_tax),
                    "reference_total_with_tax": float(reference_total),
                    "invoiced_quantity": float(invoiced_quantity),
                    "uninvoiced_quantity": float(max(Decimal("0"), quantity - invoiced_quantity)),
                    "invoice_status": invoice_status,
                    "invoiced_reference_amount": float(locked_amount),
                    "invoiced_reference_tax_amount": float(locked_tax),
                    "invoiced_reference_total_with_tax": float(locked_total),
                    "uninvoiced_reference_amount": float(max(Decimal("0"), reference_amount - locked_amount)),
                    "uninvoiced_reference_tax": float(max(Decimal("0"), reference_tax - locked_tax)),
                    "uninvoiced_reference_total_with_tax": float(max(Decimal("0"), reference_total - locked_total)),
                    "status_updated_at": str(status.get("updated_at") or ""),
                }
            )
        return sorted(result, key=lambda item: item["key"])

    def _status_for_reference(self, row: dict, status_items: dict[str, dict], workbook_status: dict[str, dict]) -> dict:
        status: dict[str, Any] = {}
        for alias in row.get("aliases") or [row.get("key", "")]:
            if alias in status_items:
                status = dict(status_items[alias])
                break
        workbook_item = workbook_status.get(row.get("key", ""))
        if workbook_item:
            if not status:
                status = dict(workbook_item)
            else:
                for field in (
                    "invoiced_reference_amount",
                    "invoiced_reference_tax",
                    "invoiced_reference_total_with_tax",
                    "reference_markup_rate",
                    "reference_markup_rate_percent",
                    "reference_markup_locked",
                    "updated_at",
                ):
                    if str(status.get(field) or "").strip() == "" and str(workbook_item.get(field) or "").strip() != "":
                        status[field] = workbook_item[field]
        if status:
            status = _normalize_status_item(str(row.get("key") or ""), status)
        return status

    def _locked_reference_part(
        self,
        status: dict,
        field: str,
        current_total: Decimal,
        quantity: Decimal,
        invoiced_quantity: Decimal,
    ) -> Decimal:
        if invoiced_quantity <= 0:
            return Decimal("0")
        value = _decimal(status.get(field))
        if value == 0 and str(status.get(field) or "").strip() == "":
            value = Decimal("0") if quantity <= 0 else current_total * invoiced_quantity / quantity
        if value < 0:
            return Decimal("0")
        return value

    def _read_reference_status_from_workbook(self) -> dict[str, dict]:
        rows = self._read_workbook_sheet("开票参考", INVOICE_REFERENCE_HEADERS)
        result: dict[str, dict] = {}
        for row in rows:
            key = _reference_key_from_values(
                row.get("发票代码(**内文字)", ""),
                row.get("内部项目名称", ""),
                row.get("规格型号", ""),
                row.get("单位", ""),
            )
            if not key:
                continue
            invoiced_quantity = _decimal(row.get("已开数量"))
            markup_text = str(row.get("参考加价率") or "").strip()
            if invoiced_quantity <= 0 and not markup_text:
                continue
            markup_rate = normalize_reference_markup_rate(markup_text, self.reference_markup_rate)
            quantity = _decimal(row.get("数量合计"))
            reference_total = _decimal(row.get("价税合计"))
            locked_total = _decimal(row.get("已开参考价税合计"))
            ratio = Decimal("0") if reference_total <= 0 else locked_total / reference_total
            result[key] = {
                "reference_key": key,
                "item_code": str(row.get("发票代码(**内文字)") or ""),
                "item_name": str(row.get("内部项目名称") or ""),
                "spec": str(row.get("规格型号") or ""),
                "unit": str(row.get("单位") or ""),
                "invoiced_quantity": _decimal_text(invoiced_quantity, "0.001"),
                "quantity_total": _decimal_text(quantity, "0.001"),
                "reference_total_with_tax": _decimal_text(reference_total, "0.000001"),
                "reference_markup_rate": _decimal_text(markup_rate, "0.000001"),
                "reference_markup_rate_percent": reference_markup_percent_text(markup_rate),
                "reference_markup_locked": False,
                "invoiced_reference_amount": _decimal_text(_decimal(row.get("金额(除税)合计")) * ratio, "0.000001"),
                "invoiced_reference_tax": _decimal_text(_decimal(row.get("税金合计")) * ratio, "0.000001"),
                "invoiced_reference_total_with_tax": _decimal_text(locked_total, "0.000001"),
                "updated_at": str(row.get("状态更新时间") or ""),
            }
        return result

    def _reference_stats(self, rows: list[dict]) -> dict:
        invoiced_count = sum(1 for row in rows if row.get("invoice_status") == "已开具")
        partial_count = sum(1 for row in rows if row.get("invoice_status") == "部分开具")
        not_invoiced_count = sum(1 for row in rows if row.get("invoice_status") == "未开具")
        inventory_total_with_tax = sum(float(row.get("total_with_tax") or 0) for row in rows)
        return {
            "row_count": len(rows),
            "total": len(rows),
            "not_invoiced": not_invoiced_count,
            "partial": partial_count,
            "invoiced": invoiced_count,
            "inventory_total_with_tax": inventory_total_with_tax,
            "project_total_with_tax": inventory_total_with_tax,
            "reference_total_with_tax": sum(float(row.get("reference_total_with_tax") or 0) for row in rows),
            "invoiced_reference_total_with_tax": sum(float(row.get("invoiced_reference_total_with_tax") or 0) for row in rows),
            "uninvoiced_reference_amount": sum(float(row.get("uninvoiced_reference_amount") or 0) for row in rows),
            "uninvoiced_reference_total_with_tax": sum(float(row.get("uninvoiced_reference_total_with_tax") or 0) for row in rows),
        }

    def _sync_status(self, detail_rows: list[dict], checks: list[dict]) -> CostSyncStatus:
        source_groups = _cost_source_groups_from_summary(self.summary_csv)
        parsed_keys = _keys_from_rows(detail_rows)
        checked_keys = _keys_from_rows(checks)
        parsed = _sources_from_rows(detail_rows)
        checked = _sources_from_rows(checks)
        review_checks = [row for row in checks if str(row.get("校验状态") or "").strip() != "通过"]
        review_keys = _keys_from_rows(review_checks)
        review_sources = _sources_from_rows(review_checks)
        if not source_groups and not detail_rows and not checks:
            return CostSyncStatus(sync_state="empty")
        if not self.detail_csv.exists() and not self.summary_xlsx.exists():
            return CostSyncStatus(source_invoice_count=len(source_groups), sync_state="not_generated", pending_count=len(source_groups))
        if source_groups:
            pending = {source for source, keys in source_groups.items() if not keys.intersection(parsed_keys)}
            checked_groups = {source for source, keys in source_groups.items() if keys.intersection(checked_keys)}
            parsed_groups = {source for source, keys in source_groups.items() if keys.intersection(parsed_keys)}
            review_groups = {source for source, keys in source_groups.items() if keys.intersection(review_keys)}
            not_parsed = checked_groups - parsed_groups
            source_count = len(source_groups)
            parsed_count = len(parsed_groups)
            checked_count = len(checked_groups)
            review_count = len(review_groups)
        else:
            pending = set()
            not_parsed = checked - parsed
            source_count = len(parsed or checked)
            parsed_count = len(parsed)
            checked_count = len(checked)
            review_count = len(review_sources)
        state = "fresh"
        if pending:
            state = "pending"
        if not_parsed or review_count:
            state = "needs_review"
        return CostSyncStatus(
            source_invoice_count=source_count,
            parsed_invoice_count=parsed_count,
            checked_invoice_count=checked_count,
            missing_count=len(pending),
            pending_count=len(pending),
            not_parsed_count=len(not_parsed),
            review_count=review_count,
            sync_state=state,
        )
