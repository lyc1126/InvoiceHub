from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import xml.etree.ElementTree as ET
import zipfile
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from invoice_hub.extraction import extract_invoice_record, normalize_classification_text


COST_DETAIL_CSV_NAME = "成本发票明细.csv"
COST_SUMMARY_XLSX_NAME = "成本发票汇总.xlsx"
COST_REFERENCE_STATUS_NAME = "成本开票状态.json"
DEFAULT_REFERENCE_MARKUP_RATE = Decimal("0.08")
REFERENCE_OUTPUT_TAX_RATE = Decimal("0.13")

DETAIL_HEADERS = [
    "销售方",
    "购买方",
    "发票号码",
    "开票日期",
    "备注项目名称",
    "内部项目名称",
    "规格型号",
    "单位",
    "数量",
    "单价(除税)",
    "平均单价(含税)",
    "金额(除税)",
    "税率",
    "税金",
    "价税合计",
    "发票代码(**内文字)",
    "源文件",
]

SELLER_SUMMARY_HEADERS = [
    "销售方",
    "发票张数",
    "明细行数",
    "数量合计",
    "金额(除税)合计",
    "税金合计",
    "价税合计",
]

PROJECT_SPEC_SUMMARY_HEADERS = [
    "销售方",
    "发票代码(**内文字)",
    "内部项目名称",
    "规格型号",
    "单位",
    "数量合计",
    "平均单价(除税)",
    "平均单价(含税)",
    "库存平均单价(除税)",
    "库存平均单价(含税)",
    "采购参考平均单价(含税)",
    "金额(除税)合计",
    "税金合计",
    "价税合计",
    "涉及发票号码",
]

INVOICE_REFERENCE_HEADERS = [
    "发票代码(**内文字)",
    "内部项目名称",
    "规格型号",
    "单位",
    "数量合计",
    "已开数量",
    "未开数量",
    "开票状态",
    "平均单价(除税)",
    "平均单价(含税)",
    "金额(除税)合计",
    "税金合计",
    "价税合计",
    "已开参考价税合计",
    "未开参考金额(除税)",
    "未开参考税金",
    "未开参考价税合计",
    "参考加价率",
    "状态更新时间",
]

CHECK_HEADERS = [
    "源文件",
    "发票大类",
    "特定业务类型",
    "类型识别状态",
    "类型识别说明",
    "销售方",
    "发票号码",
    "开票日期",
    "明细行数",
    "发票金额(除税)",
    "解析金额(除税)",
    "差异(除税)",
    "发票税金",
    "解析税金",
    "差异(税金)",
    "价税合计",
    "校验状态",
    "说明",
]

NUMERIC_DETAIL_FIELDS = {"数量", "单价(除税)", "平均单价(含税)", "金额(除税)", "税金", "价税合计"}
RUNTIME_DIR_PREFIXES = ("temp_ofd_extract_", "__pycache__")
COST_SOURCE_EXTS = {".pdf", ".ofd", ".xml"}


@dataclass
class WordBox:
    x0: float
    y0: float
    x1: float
    y1: float
    text: str


@dataclass
class HeaderColumn:
    x0: float
    x1: float
    key: str
    text: str


def _clean_text(value: object) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def _local_name(tag: str) -> str:
    return str(tag).rsplit("}", 1)[-1].lower()


def _valid_invoice_number(value: object) -> str:
    text = _clean_text(value)
    match = re.search(r"(?<!\d)(\d{8,20})(?!\d)", text)
    return match.group(1) if match else ""


def _utc_now_text() -> str:
    return datetime.now(UTC).replace(microsecond=0).isoformat().replace("+00:00", "Z")


def _reference_key_values(*values: object) -> list[str]:
    return [re.sub(r"\s+", " ", _clean_text(value)).strip() for value in values]


def _reference_key(*values: object) -> str:
    raw = "\x1f".join(_reference_key_values(*values))
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()


def _legacy_plain_reference_key(*values: object) -> str:
    return "|".join(_clean_text(value) for value in values)


def _canonical_path(path_like: object) -> str:
    text = _clean_text(path_like)
    if not text:
        return ""
    try:
        return str(Path(text).resolve())
    except OSError:
        return str(Path(text))


def _to_decimal(value: object) -> Decimal | None:
    text = _clean_text(value)
    if not text:
        return None
    text = text.replace(",", "").replace("¥", "").replace("￥", "").replace("%", "").strip()
    if not text:
        return None
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def _money_text(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    return _decimal_text(number, "0.01")


def _quantity_text(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    return _decimal_text(number, "0.001")


def _unit_price_text(value: object) -> str:
    number = _to_decimal(value)
    if number is None:
        return ""
    return _decimal_text(number, "0.000001")


def _tax_rate_text(value: object) -> str:
    text = _clean_text(value).replace("％", "%")
    if not text:
        return ""
    if text.endswith("%"):
        number = _to_decimal(text[:-1])
        if number is None:
            return ""
        return f"{_decimal_text(number, '0.001')}%"
    number = _to_decimal(text)
    if number is None:
        return ""
    if Decimal("0") <= number <= Decimal("1"):
        number *= Decimal("100")
    return f"{_decimal_text(number, '0.001')}%"


def _invoice_metadata_from_record(path: Path) -> dict[str, str]:
    try:
        record = extract_invoice_record(path)
    except Exception:
        return {}
    return {
        "销售方": record.seller,
        "购买方": record.buyer,
        "发票号码": record.invoice_number,
        "开票日期": record.invoice_date,
        "发票金额(除税)": record.pretax_amount,
        "发票税金": record.tax_amount,
        "价税合计": record.amount,
        "税率": record.tax_rate,
        "发票大类": record.invoice_type,
        "特定业务类型": record.business_type,
        "类型识别状态": record.classification_status,
        "类型识别说明": record.classification_issue,
    }


def _invoice_base_metadata(path: Path, metadata: dict | None = None) -> dict:
    metadata = metadata if isinstance(metadata, dict) else {}
    fallback = _invoice_metadata_from_record(path)
    return {
        "销售方": _metadata_value(metadata, "销售方", "seller") or fallback.get("销售方", ""),
        "购买方": _metadata_value(metadata, "购买方", "buyer") or fallback.get("购买方", ""),
        "发票号码": _metadata_value(metadata, "发票号码", "invoice_number") or fallback.get("发票号码", ""),
        "开票日期": _normalize_date(_metadata_value(metadata, "开票日期", "开票时间", "invoice_date") or fallback.get("开票日期", "")),
        "备注项目名称": _metadata_value(metadata, "备注项目名称", "项目名称"),
        "发票金额(除税)": _metadata_value(metadata, "发票金额(除税)", "除税价", "pretax_amount") or fallback.get("发票金额(除税)", ""),
        "发票税金": _metadata_value(metadata, "发票税金", "税金", "tax_amount") or fallback.get("发票税金", ""),
        "价税合计": _metadata_value(metadata, "价税合计", "开票金额", "amount") or fallback.get("价税合计", ""),
        "税率": _metadata_value(metadata, "税率", "tax_rate") or fallback.get("税率", ""),
        "发票大类": _metadata_value(metadata, "发票大类", "发票类型", "invoice_type") or fallback.get("发票大类", ""),
        "特定业务类型": _metadata_value(metadata, "特定业务类型", "business_type") or fallback.get("特定业务类型", ""),
        "类型识别状态": _metadata_value(metadata, "类型识别状态", "classification_status") or fallback.get("类型识别状态", ""),
        "类型识别说明": _metadata_value(metadata, "类型识别说明", "classification_issue") or fallback.get("类型识别说明", ""),
    }


def detail_average_unit_price_with_tax(row: dict) -> Decimal | None:
    amount = _to_decimal(row.get("金额(除税)"))
    tax_amount = _to_decimal(row.get("税金"))
    quantity = _to_decimal(row.get("数量"))
    if amount is None or tax_amount is None or quantity is None or quantity <= 0:
        return None
    if amount < 0 or tax_amount < 0:
        return None
    return (amount + tax_amount) / quantity


def detail_average_unit_price(row: dict) -> Decimal | None:
    amount = _to_decimal(row.get("金额(除税)"))
    quantity = _to_decimal(row.get("数量"))
    if amount is None or quantity is None or quantity <= 0:
        return None
    if amount < 0:
        return None
    return amount / quantity


def detail_rows_with_tax_average(rows: list[dict]) -> list[dict]:
    result = []
    for row in rows:
        enriched = dict(row)
        enriched["平均单价(含税)"] = _number(detail_average_unit_price_with_tax(row))
        result.append(enriched)
    return result


def average_unit_price(rows: list[dict]) -> Decimal | None:
    return stock_average_unit_price(rows)


def average_unit_price_with_tax(rows: list[dict]) -> Decimal | None:
    return stock_average_unit_price_with_tax(rows)


def stock_average_unit_price(rows: list[dict]) -> Decimal | None:
    total_amount = Decimal("0")
    total_quantity = Decimal("0")
    seen = False
    for row in rows:
        amount = _to_decimal(row.get("金额(除税)"))
        quantity = _to_decimal(row.get("数量"))
        if amount is None or quantity is None or quantity <= 0 or amount < 0:
            return None
        total_amount += amount
        total_quantity += quantity
        seen = True
    if not seen or total_quantity <= 0:
        return None
    return total_amount / total_quantity


def stock_average_unit_price_with_tax(rows: list[dict]) -> Decimal | None:
    total_with_tax = Decimal("0")
    total_quantity = Decimal("0")
    seen = False
    for row in rows:
        amount = _to_decimal(row.get("金额(除税)"))
        tax_amount = _to_decimal(row.get("税金"))
        quantity = _to_decimal(row.get("数量"))
        if amount is None or tax_amount is None or quantity is None or quantity <= 0:
            return None
        if amount < 0 or tax_amount < 0:
            return None
        total_with_tax += amount + tax_amount
        total_quantity += quantity
        seen = True
    if not seen or total_quantity <= 0:
        return None
    return total_with_tax / total_quantity


def purchase_reference_average_unit_price(rows: list[dict]) -> Decimal | None:
    total = Decimal("0")
    count = 0
    for row in rows:
        unit_price = detail_average_unit_price(row)
        if unit_price is None:
            return None
        total += unit_price
        count += 1
    if count <= 0:
        return None
    return total / Decimal(count)


def purchase_reference_average_unit_price_with_tax(rows: list[dict]) -> Decimal | None:
    total = Decimal("0")
    count = 0
    for row in rows:
        unit_price_with_tax = detail_average_unit_price_with_tax(row)
        if unit_price_with_tax is None:
            return None
        total += unit_price_with_tax
        count += 1
    if count <= 0:
        return None
    return total / Decimal(count)


def reference_output_average_unit_price_with_tax(avg_price: Decimal | None, markup_rate: object) -> Decimal | None:
    if avg_price is None:
        return None
    reference_average_unit_price = avg_price * reference_markup_multiplier(markup_rate)
    return reference_average_unit_price * (Decimal("1") + REFERENCE_OUTPUT_TAX_RATE)


def _round(value: Decimal | None, places: str = "0.01") -> Decimal | None:
    if value is None:
        return None
    return value.quantize(Decimal(places), rounding=ROUND_HALF_UP)


def _number(value: Decimal | None, places: str = "0.01") -> float | str:
    rounded = _round(value, places)
    if rounded is None:
        return ""
    return float(rounded)


def _source_match_tokens(*values: object) -> set[str]:
    tokens: set[str] = set()
    for value in values:
        text = _clean_text(value)
        if not text:
            continue
        normalized = text.replace("\\", "/").casefold()
        tokens.add(normalized)
        tokens.add(normalized.rstrip("/").rsplit("/", 1)[-1])
        canonical = _canonical_path(text).replace("\\", "/").casefold()
        if canonical:
            tokens.add(canonical)
            tokens.add(canonical.rstrip("/").rsplit("/", 1)[-1])
    return {token for token in tokens if token}


def _weighted_unit_price_with_tax_from_total(rows: list[dict]) -> Decimal | None:
    total_with_tax = Decimal("0")
    total_quantity = Decimal("0")
    seen = False
    for row in rows:
        row_total = _to_decimal(row.get("价税合计"))
        quantity = _to_decimal(row.get("数量"))
        if row_total is None or quantity is None or quantity <= 0 or row_total < 0:
            return None
        total_with_tax += row_total
        total_quantity += quantity
        seen = True
    if not seen or total_quantity <= 0:
        return None
    return total_with_tax / total_quantity


def _match_invoice_cost_rows(
    rows: list[dict],
    invoice_numbers: Iterable[object] = (),
    source_values: Iterable[object] = (),
) -> tuple[list[dict], str]:
    numbers = {_clean_text(value) for value in invoice_numbers if _clean_text(value)}
    by_number = [row for row in rows if _clean_text(row.get("发票号码")) in numbers]
    if by_number:
        return by_number, "invoice_number"

    source_tokens = _source_match_tokens(*source_values)
    if source_tokens:
        by_source = [row for row in rows if _source_match_tokens(row.get("源文件")) & source_tokens]
        if by_source:
            return by_source, "source_file"
    return [], "none"


def _cost_project_breakdown(project_name: str, project_rows: list[dict]) -> dict:
    specs: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for row in project_rows:
        key = (
            _clean_text(row.get("内部项目名称")),
            _clean_text(row.get("规格型号")),
            _clean_text(row.get("单位")),
        )
        specs[key].append(row)

    spec_items = []
    for key, spec_rows in sorted(specs.items(), key=lambda item: item[0]):
        spec_items.append(
            {
                "project_name": key[0],
                "specification": key[1],
                "unit": key[2],
                "quantity_total": _number(_sum_decimal(spec_rows, "数量"), "0.001"),
                "arithmetic_average_unit_price_pretax": _number(purchase_reference_average_unit_price(spec_rows)),
                "arithmetic_average_unit_price_with_tax": _number(purchase_reference_average_unit_price_with_tax(spec_rows)),
                "weighted_average_unit_price_pretax": _number(stock_average_unit_price(spec_rows)),
                "weighted_average_unit_price_with_tax": _number(_weighted_unit_price_with_tax_from_total(spec_rows)),
            }
        )

    return {
        "project_name": project_name,
        "display_project_name": project_name or "未识别项目",
        "quantity_total": _number(_sum_decimal(project_rows, "数量"), "0.001"),
        "amount_pretax_total": _number(_sum_decimal(project_rows, "金额(除税)")),
        "total_with_tax": _number(_sum_decimal(project_rows, "价税合计")),
        "specs": spec_items,
    }


def invoice_cost_breakdown(
    rows: list[dict],
    invoice_number: object = "",
    source_path: object = "",
    source_file: object = "",
) -> dict:
    """Aggregate cost detail rows for one invoice detail page."""

    number = _clean_text(invoice_number)
    matched, match_strategy = _match_invoice_cost_rows(
        rows,
        invoice_numbers=[number],
        source_values=[source_path, source_file],
    )

    if not matched:
        return {
            "available": False,
            "match_strategy": "none",
            "invoice_number": number,
            "source_file": _clean_text(source_file) or Path(_clean_text(source_path)).name,
            "detail_count": 0,
            "projects": [],
        }

    projects: dict[str, list[dict]] = defaultdict(list)
    for row in matched:
        projects[_clean_text(row.get("内部项目名称"))].append(row)

    result_projects = []
    for project_name, project_rows in sorted(projects.items(), key=lambda item: item[0]):
        result_projects.append(_cost_project_breakdown(project_name, project_rows))

    return {
        "available": True,
        "match_strategy": match_strategy or "none",
        "invoice_number": number,
        "source_file": _clean_text(source_file) or Path(_clean_text(source_path)).name,
        "detail_count": len(matched),
        "projects": result_projects,
    }


def selection_cost_breakdown(rows: list[dict], invoice_families: list[dict]) -> dict:
    """Aggregate existing cost rows for de-duplicated selected invoice families."""

    matched_rows: list[dict] = []
    seen_rows: set[int] = set()
    matched_invoice_count = 0
    match_strategy_counts = {"invoice_number": 0, "source_file": 0}

    for family in invoice_families:
        sources = [*(family.get("source_paths") or []), *(family.get("source_files") or [])]
        family_rows, match_strategy = _match_invoice_cost_rows(
            rows,
            invoice_numbers=family.get("invoice_numbers") or [],
            source_values=sources,
        )
        if not family_rows:
            continue
        matched_invoice_count += 1
        if match_strategy in match_strategy_counts:
            match_strategy_counts[match_strategy] += 1
        for row in family_rows:
            row_identity = id(row)
            if row_identity in seen_rows:
                continue
            seen_rows.add(row_identity)
            matched_rows.append(row)

    projects: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for row in matched_rows:
        project_name = _clean_text(row.get("内部项目名称"))
        tax_rate = _tax_rate_text(row.get("税率"))
        projects[(project_name, tax_rate)].append(row)

    result_projects = []
    for (project_name, tax_rate), project_rows in sorted(
        projects.items(),
        key=lambda item: (item[0][0], item[0][1] or "\uffff"),
    ):
        project = _cost_project_breakdown(project_name, project_rows)
        project.update(
            {
                "tax_rate": tax_rate,
                "display_tax_rate": tax_rate or "税率未识别",
            }
        )
        result_projects.append(project)

    invoice_count = len(invoice_families)
    return {
        "available": bool(matched_rows),
        "matched_invoice_count": matched_invoice_count,
        "unmatched_invoice_count": max(0, invoice_count - matched_invoice_count),
        "detail_count": len(matched_rows),
        "match_strategy_counts": match_strategy_counts,
        "projects": result_projects,
    }


def _decimal_text(value: Decimal | None, places: str = "0.001") -> str:
    rounded = _round(value, places)
    if rounded is None:
        return "0"
    text = format(rounded, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")
    return text or "0"


def normalize_reference_markup_rate(value: object, default: object = DEFAULT_REFERENCE_MARKUP_RATE) -> Decimal:
    for candidate in (value, default, DEFAULT_REFERENCE_MARKUP_RATE):
        text = _clean_text(candidate).replace("％", "%")
        if not text:
            continue
        is_percent = text.endswith("%")
        if is_percent:
            text = text[:-1].strip()
        try:
            number = Decimal(text)
        except InvalidOperation:
            continue
        if not number.is_finite() or number < 0:
            continue
        if is_percent or number > 1:
            number = number / Decimal("100")
        return number
    return DEFAULT_REFERENCE_MARKUP_RATE


def normalize_reference_markup_percent(value: object) -> Decimal:
    text = _clean_text(value).replace("％", "%")
    if text.endswith("%"):
        text = text[:-1].strip()
    if not re.fullmatch(r"(?:[0-9]+(?:\.[0-9]*)?|\.[0-9]+)", text):
        raise ValueError(f"开票加价率不是有效数字: {value}")
    try:
        number = Decimal(text)
    except InvalidOperation as exc:
        raise ValueError(f"开票加价率不是有效数字: {value}") from exc
    if not number.is_finite() or number < 0:
        raise ValueError(f"开票加价率不是有效数字: {value}")
    return number / Decimal("100")


def reference_markup_multiplier(value: object) -> Decimal:
    return Decimal("1") + normalize_reference_markup_rate(value)


def reference_markup_percent_text(value: object) -> str:
    return _decimal_text(normalize_reference_markup_rate(value) * Decimal("100"), "0.001")


def reference_markup_display(value: object) -> str:
    return f"{reference_markup_percent_text(value)}%"


def _truthy(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return _clean_text(value).casefold() in {"1", "true", "yes", "y", "locked", "已锁定", "锁定"}


def _row_markup_meta(status_item: dict | None, fallback_rate: object, fallback_locked: bool = False) -> dict:
    status_item = status_item if isinstance(status_item, dict) else {}
    rate = normalize_reference_markup_rate(fallback_rate)
    if _clean_text(status_item.get("reference_markup_rate_percent")):
        rate = normalize_reference_markup_percent(status_item.get("reference_markup_rate_percent"))
    elif _clean_text(status_item.get("reference_markup_rate")):
        rate = normalize_reference_markup_rate(status_item.get("reference_markup_rate"), fallback_rate)
    elif _clean_text(status_item.get("markup_rate")):
        rate = normalize_reference_markup_rate(status_item.get("markup_rate"), fallback_rate)
    return {
        "rate": rate,
        "display": reference_markup_display(rate),
        "percent": reference_markup_percent_text(rate),
        "locked": _truthy(status_item.get("reference_markup_locked", fallback_locked)),
    }


def _clamp_decimal(value: Decimal, minimum: Decimal, maximum: Decimal) -> Decimal:
    if value < minimum:
        return minimum
    if value > maximum:
        return maximum
    return value


def _invoice_status(quantity: Decimal, invoiced: Decimal) -> str:
    if invoiced <= 0:
        return "未开具"
    if quantity > 0 and invoiced >= quantity:
        return "已开具"
    return "部分开具"


def _ratio_part(value: Decimal, quantity: Decimal, part_quantity: Decimal) -> Decimal:
    if quantity <= 0 or part_quantity <= 0:
        return Decimal("0")
    return value * part_quantity / quantity


def _locked_reference_part(
    status_item: dict,
    field: str,
    current_total: Decimal,
    quantity: Decimal,
    invoiced_quantity: Decimal,
) -> Decimal:
    if invoiced_quantity <= 0:
        return Decimal("0")
    value = _to_decimal(status_item.get(field))
    if value is not None:
        return max(value, Decimal("0"))
    return _ratio_part(current_total, quantity, invoiced_quantity)


def _remaining_reference_part(current_total: Decimal, locked_part: Decimal) -> Decimal:
    remaining = current_total - locked_part
    if remaining <= 0:
        return Decimal("0")
    return remaining


def _sum_decimal(rows: Iterable[dict], field: str) -> Decimal:
    total = Decimal("0")
    for row in rows:
        value = _to_decimal(row.get(field))
        if value is not None:
            total += value
    return total


def _as_words(raw_words: Iterable) -> list[WordBox]:
    words: list[WordBox] = []
    for raw in raw_words or []:
        if isinstance(raw, WordBox):
            box = raw
        elif isinstance(raw, dict):
            box = WordBox(
                float(raw.get("x0", 0)),
                float(raw.get("y0", 0)),
                float(raw.get("x1", raw.get("x0", 0))),
                float(raw.get("y1", raw.get("y0", 0))),
                _clean_text(raw.get("text")),
            )
        else:
            try:
                box = WordBox(float(raw[0]), float(raw[1]), float(raw[2]), float(raw[3]), _clean_text(raw[4]))
            except Exception:
                continue
        if box.text:
            words.append(box)
    return words


def _page_width(words: list[WordBox], explicit_width: float | None = None) -> float:
    if explicit_width and explicit_width > 0:
        return float(explicit_width)
    max_x = max((word.x1 for word in words), default=595.0)
    return max(595.0, max_x)


_HEADER_ALIASES = {
    "item_name": {
        "项目名称",
        "货物或应税劳务服务名称",
        "货物或应税劳务、服务名称",
        "货物名称",
        "商品名称",
        "服务名称",
        "出行人",
        "旅客姓名",
        "运输货物名称",
        "不动产名称",
    },
    "spec": {"规格型号", "规格"},
    "unit": {"单位", "面积单位"},
    "quantity": {"数量", "面积", "建筑面积"},
    "unit_price": {"单价", "不含税单价"},
    "amount": {"金额", "不含税金额", "金额(不含税)"},
    "tax_rate": {"税率", "税率/征收率", "征收率"},
    "tax_amount": {"税额"},
    "ignored": {
        "建筑服务发生地",
        "建筑项目名称",
        "发生地",
        "项目地址",
        "有效身份证件号",
        "出行日期",
        "出发地",
        "到达地",
        "交通工具类型",
        "交通工具等级",
        "等级",
        "运输工具种类",
        "运输工具牌号",
        "产权证书号",
        "产权证号",
        "不动产单元代码",
        "跨地市标志",
        "跨地(市)标志",
    },
}
_NORMALIZED_HEADER_ALIASES = {
    normalize_classification_text(alias).replace("、", "")
    .replace(",", "")
    .replace("，", ""): key
    for key, aliases in _HEADER_ALIASES.items()
    for alias in aliases
}
_REQUIRED_HEADER_KEYS = {"item_name", "amount", "tax_rate", "tax_amount"}


def _header_key(value: object) -> str:
    text = normalize_classification_text(value).replace("、", "").replace(",", "").replace("，", "")
    return _NORMALIZED_HEADER_ALIASES.get(text, "")


def _header_cells(words: list[WordBox]) -> list[tuple[float, float, str, str]]:
    ordered = sorted(words, key=lambda word: (word.x0, word.y0))
    cells: list[tuple[float, float, str, str]] = []
    index = 0
    while index < len(ordered):
        matched = False
        for size in range(min(4, len(ordered) - index), 0, -1):
            group = ordered[index : index + size]
            if max(word.y1 for word in group) - min(word.y0 for word in group) > 22:
                continue
            combined = "".join(word.text for word in group)
            key = _header_key(combined)
            if not key:
                continue
            cells.append((min(word.x0 for word in group), max(word.x1 for word in group), key, combined))
            index += size
            matched = True
            break
        if matched:
            continue
        word = ordered[index]
        cells.append((word.x0, word.x1, "", word.text))
        index += 1
    return cells


def _find_header_layout(words: list[WordBox], width: float) -> tuple[list[HeaderColumn], float, float]:
    recognized = [word for word in words if _header_key(word.text)]
    best: tuple[int, float, list[WordBox], list[tuple[float, float, str, str]]] | None = None
    for anchor in recognized:
        window = [
            word
            for word in words
            if abs(word.y0 - anchor.y0) <= 6
            and not word.text.lstrip().startswith("*")
            and _to_decimal(word.text) is None
        ]
        cells = _header_cells(window)
        keys = {key for _x0, _x1, key, _text in cells if key}
        if not _REQUIRED_HEADER_KEYS.issubset(keys):
            continue
        score = len(keys)
        header_top = min(word.y0 for word in window)
        candidate = (score, -header_top, window, cells)
        if best is None or candidate[:2] > best[:2]:
            best = candidate
    if best is None:
        return [], 0.0, 0.0

    _score, _negative_y, window, raw_cells = best
    centers = [((x0 + x1) / 2.0, key, text) for x0, x1, key, text in raw_cells]
    centers.sort(key=lambda item: item[0])
    columns: list[HeaderColumn] = []
    for index, (center, key, text) in enumerate(centers):
        left = 0.0 if index == 0 else (centers[index - 1][0] + center) / 2.0
        right = width if index == len(centers) - 1 else (center + centers[index + 1][0]) / 2.0
        columns.append(HeaderColumn(left, right, key, text))
    return columns, min(word.y0 for word in window), max(word.y1 for word in window)


def _column_for_word(word: WordBox, columns: list[HeaderColumn]) -> str:
    center = (word.x0 + word.x1) / 2.0
    for column in columns:
        if column.x0 <= center < column.x1:
            return column.key
    return ""


def _group_words_by_baseline(words: list[WordBox], tolerance: float = 3.0) -> list[list[WordBox]]:
    groups: list[list[WordBox]] = []
    centers: list[float] = []
    for word in sorted(words, key=lambda item: (((item.y0 + item.y1) / 2.0), item.x0)):
        center = (word.y0 + word.y1) / 2.0
        if not groups or abs(center - centers[-1]) > tolerance:
            groups.append([word])
            centers.append(center)
            continue
        groups[-1].append(word)
        centers[-1] = sum((item.y0 + item.y1) / 2.0 for item in groups[-1]) / len(groups[-1])
    for group in groups:
        group.sort(key=lambda item: item.x0)
    return groups


def _band_cells(words: list[WordBox], columns: list[HeaderColumn]) -> dict[str, list[str]]:
    cells: dict[str, list[str]] = defaultdict(list)
    for word in words:
        key = _column_for_word(word, columns)
        if key and key != "ignored":
            cells[key].append(word.text)
    return cells


def _find_header_y(words: list[WordBox]) -> float:
    width = _page_width(words)
    _columns, header_top, _header_bottom = _find_header_layout(words, width)
    return header_top


def _find_stop_y(words: list[WordBox], header_bottom: float, width: float) -> float | None:
    candidates = []
    for word in words:
        if word.y0 <= header_bottom + 4:
            continue
        if "价税合计" in word.text:
            candidates.append(word.y0)
            continue
        if word.x0 < width * 0.22 and word.text in {"合", "计", "合计"}:
            candidates.append(word.y0)
    return min(candidates) if candidates else None


def _split_item_name(raw_name: str) -> tuple[str, str]:
    text = re.sub(r"\s+", "", _clean_text(raw_name))
    match = re.match(r"^\*([^*]+)\*(.+)$", text)
    if not match:
        return "", text
    return match.group(1).strip(), match.group(2).strip()


def _first_decimal(values: list[str]) -> Decimal | None:
    for value in values:
        parsed = _to_decimal(value)
        if parsed is not None:
            return parsed
    return None


def _first_text(values: list[str]) -> str:
    return "".join(_clean_text(value) for value in values if _clean_text(value)).strip()


def _metadata_value(metadata: dict | None, *keys: str) -> str:
    if not isinstance(metadata, dict):
        return ""
    for key in keys:
        value = _clean_text(metadata.get(key))
        if value:
            return value
    return ""


def _normalize_date(value: str) -> str:
    text = _clean_text(value)
    match = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", text)
    if not match:
        match = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", text)
    if not match:
        return text
    return f"{match.group(1)}-{match.group(2).zfill(2)}-{match.group(3).zfill(2)}"


def _extract_project_name(page_text: str) -> str:
    for line in str(page_text or "").splitlines():
        text = _clean_text(line)
        match = re.search(r"项目名称[:：]\s*(.+)", text)
        if match:
            return match.group(1).strip()
    return ""


def _extract_basic_metadata(words: list[WordBox], page_text: str, page_width: float) -> dict:
    lines = [_clean_text(line) for line in str(page_text or "").splitlines() if _clean_text(line)]
    text = "\n".join(lines)
    invoice_number = ""
    date_value = ""
    invoice_pretax = None
    invoice_tax = None
    invoice_total = None

    number_match = re.search(r"(?<!\d)(\d{20})(?!\d)", text)
    if number_match:
        invoice_number = number_match.group(1)

    date_match = re.search(r"\d{4}年\d{1,2}月\d{1,2}日|\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", text)
    if date_match:
        date_value = _normalize_date(date_match.group(0))

    companies = []
    for line in lines:
        if "公司" not in line:
            continue
        if any(marker in line for marker in ("开户银行", "银行账号", "发票", "下载次数")):
            continue
        if line not in companies:
            companies.append(line)
    buyer = companies[0] if len(companies) >= 1 else ""
    seller = companies[1] if len(companies) >= 2 else ""

    money_words = [word for word in words if word.text.startswith(("¥", "￥"))]
    money_groups: dict[int, list[WordBox]] = defaultdict(list)
    header_y = _find_header_y(words)
    for word in money_words:
        if word.y0 > header_y:
            money_groups[int(round(word.y0))].append(word)
    for _, grouped in sorted(money_groups.items()):
        if len(grouped) < 2:
            continue
        for word in grouped:
            value = _to_decimal(word.text)
            ratio = word.x0 / max(page_width, 1.0)
            if 0.60 <= ratio < 0.80 and value is not None:
                invoice_pretax = value
            elif ratio >= 0.86 and value is not None:
                invoice_tax = value
        if invoice_pretax is not None and invoice_tax is not None:
            break
    for _, grouped in sorted(money_groups.items()):
        for word in grouped:
            value = _to_decimal(word.text)
            if value is None:
                continue
            if invoice_pretax is not None and invoice_tax is not None and _round(invoice_pretax + invoice_tax) == _round(value):
                invoice_total = value
                break
        if invoice_total is not None:
            break

    return {
        "销售方": seller,
        "购买方": buyer,
        "发票号码": invoice_number,
        "开票日期": date_value,
        "备注项目名称": _extract_project_name(page_text),
        "发票金额(除税)": _number(invoice_pretax),
        "发票税金": _number(invoice_tax),
        "价税合计": _number(invoice_total),
    }


def _finish_row(raw_row: dict, base_meta: dict, source_name: str) -> dict | None:
    raw_name = _first_text(raw_row.get("name_parts", []))
    category, item_name = _split_item_name(raw_name)
    if not item_name:
        return None

    cells = raw_row.get("cells", {})
    quantity = _first_decimal(cells.get("quantity", []))
    unit_price = _first_decimal(cells.get("unit_price", []))
    amount = _first_decimal(cells.get("amount", []))
    tax_amount = _first_decimal(cells.get("tax_amount", []))
    total = amount + tax_amount if amount is not None and tax_amount is not None else None

    if amount is None and tax_amount is None and quantity is None:
        return None

    return {
        "销售方": base_meta.get("销售方", ""),
        "购买方": base_meta.get("购买方", ""),
        "发票号码": base_meta.get("发票号码", ""),
        "开票日期": base_meta.get("开票日期", ""),
        "备注项目名称": base_meta.get("备注项目名称", ""),
        "内部项目名称": item_name,
        "规格型号": _first_text(cells.get("spec", [])),
        "单位": _first_text(cells.get("unit", [])),
        "数量": _number(quantity, "0.001"),
        "单价(除税)": _number(unit_price),
        "金额(除税)": _number(amount),
        "税率": _first_text(cells.get("tax_rate", [])),
        "税金": _number(tax_amount),
        "价税合计": _number(total),
        "发票代码(**内文字)": category,
        "源文件": source_name,
    }


def parse_cost_rows_from_words(
    raw_words: Iterable,
    page_text: str = "",
    source_name: str = "",
    metadata: dict | None = None,
    page_width: float | None = None,
) -> tuple[list[dict], dict]:
    words = _as_words(raw_words)
    width = _page_width(words, page_width)
    columns, _header_top, header_bottom = _find_header_layout(words, width)
    stop_y = _find_stop_y(words, header_bottom, width) if columns else None
    parsed_meta = _extract_basic_metadata(words, page_text, width)
    base_meta = {
        "销售方": _metadata_value(metadata, "销售方", "seller") or parsed_meta.get("销售方", ""),
        "购买方": _metadata_value(metadata, "购买方", "buyer") or parsed_meta.get("购买方", ""),
        "发票号码": _metadata_value(metadata, "发票号码", "invoice_number") or parsed_meta.get("发票号码", ""),
        "开票日期": _normalize_date(_metadata_value(metadata, "开票日期", "开票时间", "invoice_date") or parsed_meta.get("开票日期", "")),
        "备注项目名称": _metadata_value(metadata, "备注项目名称", "项目名称") or parsed_meta.get("备注项目名称", ""),
        "发票金额(除税)": parsed_meta.get("发票金额(除税)", "") or _metadata_value(metadata, "发票金额(除税)", "除税价"),
        "发票税金": parsed_meta.get("发票税金", "") or _metadata_value(metadata, "发票税金", "税金"),
        "价税合计": parsed_meta.get("价税合计", "") or _metadata_value(metadata, "价税合计", "开票金额"),
        "发票大类": _metadata_value(metadata, "发票大类", "发票类型", "invoice_type"),
        "特定业务类型": _metadata_value(metadata, "特定业务类型", "business_type"),
        "类型识别状态": _metadata_value(metadata, "类型识别状态", "classification_status"),
        "类型识别说明": _metadata_value(metadata, "类型识别说明", "classification_issue"),
    }

    if not columns:
        base_meta["_detail_parse_issue"] = "未识别到可靠明细表头（至少需要项目名称、金额、税率、税额）"
        return [], base_meta

    detail_words = [
        word
        for word in words
        if word.y0 > header_bottom - 0.5
        and (stop_y is None or word.y0 < stop_y - 2)
    ]

    current: dict | None = None
    pending: list[tuple[float, dict[str, list[str]]]] = []
    rows: list[dict] = []
    for band in _group_words_by_baseline(detail_words):
        band_y = sum((word.y0 + word.y1) / 2.0 for word in band) / len(band)
        cells = _band_cells(band, columns)
        name_parts = cells.pop("item_name", [])
        name_text = _first_text(name_parts)
        starts_item = any(part.lstrip().startswith("*") for part in name_parts)
        has_row_values = bool(cells.get("amount") or cells.get("tax_amount"))
        starts_item = starts_item or bool(name_text and has_row_values)

        if starts_item:
            if current is not None:
                finished = _finish_row(current, base_meta, source_name)
                if finished:
                    rows.append(finished)
            current = {
                "start_y": band_y,
                "name_parts": list(name_parts),
                "cells": defaultdict(list),
            }
            for pending_y, pending_cells in pending:
                if 0 <= band_y - pending_y <= 8:
                    for key, values in pending_cells.items():
                        current["cells"][key].extend(values)
            pending.clear()
            for key, values in cells.items():
                current["cells"][key].extend(values)
            continue

        if current is None:
            if cells:
                pending.append((band_y, cells))
                pending = [(pending_y, values) for pending_y, values in pending if band_y - pending_y <= 8]
            continue

        distance = band_y - float(current.get("start_y", 0))
        if name_parts and 0 <= distance <= 18:
            current["name_parts"].extend(name_parts)
        if 0 <= distance <= 8:
            for key, values in cells.items():
                current["cells"][key].extend(values)

    if current is not None:
        finished = _finish_row(current, base_meta, source_name)
        if finished:
            rows.append(finished)
    if not rows:
        base_meta["_detail_parse_issue"] = "已识别明细表头，但未识别到可靠成本明细行"
    return rows, base_meta


def _merge_issue_text(*values: object) -> str:
    issues: list[str] = []
    for value in values:
        for issue in str(value or "").split("；"):
            clean = issue.strip()
            if clean and clean not in issues:
                issues.append(clean)
    return "；".join(issues)


def _cost_validation(rows: list[dict], invoice: dict) -> dict:
    invoice_amount = _to_decimal(invoice.get("发票金额(除税)") or invoice.get("除税价"))
    invoice_tax = _to_decimal(invoice.get("发票税金") or invoice.get("税金"))
    parsed_amount = _sum_decimal(rows, "金额(除税)")
    parsed_tax = _sum_decimal(rows, "税金")
    amount_values_complete = bool(rows) and all(_to_decimal(row.get("金额(除税)")) is not None for row in rows)
    tax_values_complete = bool(rows) and all(_to_decimal(row.get("税金")) is not None for row in rows)
    amount_ok = (
        invoice_amount is not None
        and amount_values_complete
        and abs(parsed_amount - invoice_amount) <= Decimal("0.02")
    )
    tax_ok = (
        invoice_tax is not None
        and tax_values_complete
        and abs(parsed_tax - invoice_tax) <= Decimal("0.02")
    )
    issues: list[str] = []
    if not rows:
        issues.append("未识别到成本明细行")
    else:
        if invoice_amount is None:
            issues.append("票头除税金额缺失")
        elif not amount_values_complete:
            issues.append("明细除税金额缺失")
        elif not amount_ok:
            issues.append("除税金额校验不通过")
        if invoice_tax is None:
            issues.append("票头税额缺失")
        elif not tax_values_complete:
            issues.append("明细税额缺失")
        elif not tax_ok:
            issues.append("税额校验不通过")
    return {
        "invoice_amount": invoice_amount,
        "invoice_tax": invoice_tax,
        "parsed_amount": parsed_amount,
        "parsed_tax": parsed_tax,
        "amount_ok": amount_ok,
        "tax_ok": tax_ok,
        "score": int(amount_ok) + int(tax_ok),
        "issue": "；".join(issues),
    }


def _finalize_cost_analysis(analysis: dict) -> dict:
    result = dict(analysis)
    invoice = dict(result.get("invoice") or {})
    rows = list(result.get("rows") or [])
    validation = _cost_validation(rows, invoice)
    result["amount_validation_ok"] = validation["amount_ok"]
    result["tax_validation_ok"] = validation["tax_ok"]
    result["validation_score"] = validation["score"]
    result["no_detail_rows"] = not rows
    if result.get("status") in {"failed", "skipped"}:
        return result
    parse_issue = invoice.pop("_detail_parse_issue", "")
    result["invoice"] = invoice
    result["status"] = "ok" if validation["score"] == 2 else "needs_review"
    result["message"] = _merge_issue_text(result.get("message"), parse_issue, validation["issue"])
    return result


def analyze_pdf_costs(pdf_path: Path, metadata: dict | None = None) -> dict:
    try:
        import fitz
    except ImportError as exc:
        return {
            "source": Path(pdf_path).name,
            "rows": [],
            "invoice": dict(metadata or {}),
            "status": "skipped",
            "message": f"PyMuPDF unavailable: {exc}",
        }

    source_path = Path(pdf_path)
    all_rows: list[dict] = []
    invoice_meta: dict = _invoice_base_metadata(source_path, metadata)
    parse_issues: list[str] = []
    try:
        with fitz.open(str(source_path)) as doc:
            for page in doc:
                page_text = page.get_text()
                page_rows, page_meta = parse_cost_rows_from_words(
                    page.get_text("words"),
                    page_text=page_text,
                    source_name=source_path.name,
                    metadata=invoice_meta,
                    page_width=float(page.rect.width),
                )
                parse_issue = _clean_text(page_meta.pop("_detail_parse_issue", ""))
                if parse_issue:
                    parse_issues.append(parse_issue)
                invoice_meta = {**page_meta, **{key: value for key, value in invoice_meta.items() if _clean_text(value)}}
                all_rows.extend(page_rows)
    except Exception as exc:
        return {
            "source": source_path.name,
            "rows": [],
            "invoice": invoice_meta,
            "status": "failed",
            "message": str(exc),
        }

    return _finalize_cost_analysis(
        {
            "source": source_path.name,
            "rows": all_rows,
            "invoice": invoice_meta,
            "status": "parsed",
            "message": "；".join(dict.fromkeys(parse_issues)),
        }
    )


def _element_text(element: ET.Element) -> str:
    return "".join(element.itertext()).strip()


def _first_alias(mapping: dict[str, str], *aliases: str) -> str:
    for alias in aliases:
        value = _clean_text(mapping.get(alias.lower()))
        if value:
            return value
    return ""


def _cost_detail_row_from_values(values: dict[str, str], base_meta: dict, source_name: str) -> dict | None:
    raw_name = _first_alias(values, "itemname", "item_name", "goodsname", "productname", "xmmc", "hwmc")
    category, item_name = _split_item_name(raw_name)
    if not item_name:
        item_name = raw_name
    if not item_name:
        return None

    amount = _to_decimal(_first_alias(values, "amount", "taxexclusiveamount", "totalamwithouttax", "je"))
    tax_amount = _to_decimal(_first_alias(values, "comtaxam", "taxamount", "se"))
    total = _to_decimal(_first_alias(values, "totaltaxincludedamount", "totaltax-includedamount", "totalwithtax", "jshj"))
    if total is None and amount is not None and tax_amount is not None:
        total = amount + tax_amount
    quantity = _to_decimal(_first_alias(values, "quantity", "sl"))
    unit_price = _to_decimal(_first_alias(values, "unprice", "unitprice", "price", "dj"))
    tax_rate = _tax_rate_text(_first_alias(values, "taxrate", "taxscheme", "slv") or base_meta.get("税率"))

    if amount is None and tax_amount is None and quantity is None:
        return None

    return {
        "销售方": base_meta.get("销售方", ""),
        "购买方": base_meta.get("购买方", ""),
        "发票号码": base_meta.get("发票号码", ""),
        "开票日期": base_meta.get("开票日期", ""),
        "备注项目名称": base_meta.get("备注项目名称", ""),
        "内部项目名称": item_name,
        "规格型号": _first_alias(values, "specmod", "specification", "specificationmodel", "model", "ggxh"),
        "单位": _first_alias(values, "meaunits", "measurementdimension", "unit", "dw"),
        "数量": _quantity_text(quantity),
        "单价(除税)": _unit_price_text(unit_price),
        "金额(除税)": _money_text(amount),
        "税率": tax_rate,
        "税金": _money_text(tax_amount),
        "价税合计": _money_text(total),
        "发票代码(**内文字)": category,
        "源文件": source_name,
    }


def _xml_cost_row_values(path: Path) -> list[dict[str, str]]:
    try:
        root = ET.parse(path).getroot()
    except Exception:
        return []
    rows: list[dict[str, str]] = []
    container_names = {"issuiteminformation", "issueiteminformation", "iteminformation", "goodsinformation", "goodsitem", "detailitem"}
    for element in root.iter():
        if _local_name(element.tag) not in container_names:
            continue
        values = {_local_name(child.tag): _element_text(child) for child in list(element)}
        if not values:
            continue
        if not _first_alias(values, "itemname", "goodsname", "productname", "xmmc", "hwmc"):
            continue
        if not any(_first_alias(values, alias) for alias in ("amount", "taxexclusiveamount", "comtaxam", "taxamount", "quantity")):
            continue
        rows.append(values)
    return rows


def analyze_xml_costs(xml_path: Path, metadata: dict | None = None) -> dict:
    source_path = Path(xml_path)
    base_meta = _invoice_base_metadata(source_path, metadata)
    rows = [
        row
        for values in _xml_cost_row_values(source_path)
        if (row := _cost_detail_row_from_values(values, base_meta, source_path.name)) is not None
    ]
    return _finalize_cost_analysis(
        {
            "source": source_path.name,
            "rows": rows,
            "invoice": base_meta,
            "status": "parsed",
            "message": "" if rows else "未识别到XML结构化成本明细行",
        }
    )


def _ofd_text_objects_and_refs(path: Path) -> tuple[dict[str, str], dict[str, list[str]]]:
    text_by_id: dict[str, str] = {}
    refs: dict[str, list[str]] = defaultdict(list)
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            if not name.lower().endswith("content.xml"):
                continue
            try:
                root = ET.fromstring(archive.read(name))
            except Exception:
                continue
            for element in root.iter():
                if _local_name(element.tag) != "textobject":
                    continue
                object_id = element.attrib.get("ID")
                if not object_id:
                    continue
                text = "".join(child.text or "" for child in element.iter() if _local_name(child.tag) == "textcode").strip()
                if text:
                    text_by_id[object_id] = text

        for name in archive.namelist():
            if not name.lower().endswith("customtag.xml"):
                continue
            try:
                root = ET.fromstring(archive.read(name))
            except Exception:
                continue

            def walk(element: ET.Element, parts: list[str]) -> None:
                local = _local_name(element.tag)
                if local == "objectref":
                    object_id = (element.text or "").strip()
                    if object_id and parts:
                        refs["/".join(parts)].append(object_id)
                    return
                current = [*parts, local]
                for child in list(element):
                    walk(child, current)

            walk(root, [])
    return text_by_id, refs


def _ofd_values_by_leaf(text_by_id: dict[str, str], refs: dict[str, list[str]], *leaf_names: str) -> list[str]:
    names = {name.lower() for name in leaf_names}
    values: list[str] = []
    for field_path, object_ids in refs.items():
        leaf = field_path.rsplit("/", 1)[-1].lower()
        if leaf not in names:
            continue
        for object_id in object_ids:
            value = _clean_text(text_by_id.get(object_id))
            if value:
                values.append(value)
    return values


def _align_ofd_values(values: list[str], row_count: int) -> list[str]:
    if row_count <= 0:
        return []
    if len(values) == row_count:
        return values
    if len(values) > row_count and len(values) % row_count == 0:
        size = len(values) // row_count
        return ["".join(values[index * size : (index + 1) * size]) for index in range(row_count)]
    return [values[index] if index < len(values) else "" for index in range(row_count)]


def _ofd_cost_row_values(path: Path) -> list[dict[str, str]]:
    try:
        text_by_id, refs = _ofd_text_objects_and_refs(path)
    except Exception:
        return []
    if not text_by_id or not refs:
        return []
    columns = {
        "itemname": _ofd_values_by_leaf(text_by_id, refs, "item", "itemname", "goodsname"),
        "specmod": _ofd_values_by_leaf(text_by_id, refs, "specification", "specmod", "model"),
        "meaunits": _ofd_values_by_leaf(text_by_id, refs, "measurementdimension", "meaunits", "unit"),
        "quantity": _ofd_values_by_leaf(text_by_id, refs, "quantity"),
        "unprice": _ofd_values_by_leaf(text_by_id, refs, "price", "unprice", "unitprice"),
        "amount": _ofd_values_by_leaf(text_by_id, refs, "amount"),
        "taxrate": _ofd_values_by_leaf(text_by_id, refs, "taxscheme", "taxrate"),
        "comtaxam": _ofd_values_by_leaf(text_by_id, refs, "taxamount"),
    }
    row_count = max((len(columns[key]) for key in ("specmod", "meaunits", "quantity", "unprice", "amount", "comtaxam")), default=0)
    if row_count <= 0:
        return []
    aligned = {key: _align_ofd_values(values, row_count) for key, values in columns.items()}
    rows = []
    for index in range(row_count):
        row = {key: values[index] if index < len(values) else "" for key, values in aligned.items()}
        if row.get("itemname"):
            rows.append(row)
    return rows


def analyze_ofd_costs(ofd_path: Path, metadata: dict | None = None) -> dict:
    source_path = Path(ofd_path)
    base_meta = _invoice_base_metadata(source_path, metadata)
    rows = [
        row
        for values in _ofd_cost_row_values(source_path)
        if (row := _cost_detail_row_from_values(values, base_meta, source_path.name)) is not None
    ]
    return _finalize_cost_analysis(
        {
            "source": source_path.name,
            "rows": rows,
            "invoice": base_meta,
            "status": "parsed",
            "message": "" if rows else "未识别到OFD结构化成本明细行",
        }
    )


def analyze_cost_invoice(path: Path, metadata: dict | None = None) -> dict:
    suffix = Path(path).suffix.lower()
    if suffix == ".xml":
        return analyze_xml_costs(path, metadata=metadata)
    if suffix == ".ofd":
        return analyze_ofd_costs(path, metadata=metadata)
    return analyze_pdf_costs(path, metadata=metadata)


def _normalize_metadata_index(invoice_metadata: object) -> dict[str, dict]:
    index: dict[str, dict] = {}
    if invoice_metadata is None:
        return index
    if isinstance(invoice_metadata, dict):
        iterable = invoice_metadata.values() if all(isinstance(v, dict) for v in invoice_metadata.values()) else []
    else:
        iterable = invoice_metadata
    for row in iterable or []:
        if not isinstance(row, dict):
            continue
        path_value = row.get("文件路径") or row.get("file_path") or row.get("path")
        key = _canonical_path(path_value)
        if key:
            index[key] = dict(row)
    return index


def _load_reference_status(path: Path) -> dict[str, dict]:
    path = Path(path)
    if not path.exists():
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception:
        return {}
    raw_items = payload.get("items") if isinstance(payload, dict) else None
    if not isinstance(raw_items, dict):
        return {}
    items: dict[str, dict] = {}
    for key, item in raw_items.items():
        if not isinstance(item, dict):
            continue
        normalized_key = _clean_text(item.get("reference_key")) or _clean_text(key)
        if not normalized_key:
            continue
        normalized = dict(item)
        normalized.setdefault("reference_key", normalized_key)
        field_aliases = {
            "locked_reference_amount": "invoiced_reference_amount",
            "locked_reference_tax_amount": "invoiced_reference_tax",
            "locked_reference_total_with_tax": "invoiced_reference_total_with_tax",
        }
        for old_field, new_field in field_aliases.items():
            if _clean_text(normalized.get(new_field)) == "" and _clean_text(normalized.get(old_field)) != "":
                normalized[new_field] = normalized[old_field]
        if _clean_text(normalized.get("reference_markup_rate")) == "" and _clean_text(normalized.get("markup_rate")) != "":
            normalized["reference_markup_rate"] = normalized["markup_rate"]
        items[normalized_key] = normalized
    return items


def _write_reference_status(
    path: Path,
    items: dict[str, dict],
    reference_markup_rate: object = DEFAULT_REFERENCE_MARKUP_RATE,
    reference_markup_locked: bool = False,
) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    markup_rate = normalize_reference_markup_rate(reference_markup_rate)
    normalized = {
        key: dict(value)
        for key, value in sorted((items or {}).items())
        if _clean_text(key) and isinstance(value, dict)
    }
    payload = {
        "version": 1,
        "updated_at": _utc_now_text(),
        "reference_markup_rate": _decimal_text(markup_rate, "0.000001"),
        "reference_markup_rate_percent": reference_markup_percent_text(markup_rate),
        "reference_markup_locked": bool(reference_markup_locked),
        "items": normalized,
    }
    tmp_path = path.with_name(f"{path.name}.tmp")
    tmp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    os.replace(tmp_path, path)


def _load_workbook_reference_status(path: Path) -> dict[str, dict]:
    path = Path(path)
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "开票参考" not in wb.sheetnames:
            return {}
        ws = wb["开票参考"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        return {}
    finally:
        try:
            wb.close()
        except Exception:
            pass
    if not rows:
        return {}
    headers = [_clean_text(cell) for cell in rows[0]]
    result: dict[str, dict] = {}
    for values in rows[1:]:
        row = {headers[index]: values[index] if index < len(values) else "" for index in range(len(headers))}
        key = (
            _clean_text(row.get("发票代码(**内文字)")),
            _clean_text(row.get("内部项目名称")),
            _clean_text(row.get("规格型号")),
            _clean_text(row.get("单位")),
        )
        reference_key = _reference_key(*key)
        invoiced_quantity = _to_decimal(row.get("已开数量")) or Decimal("0")
        markup_text = _clean_text(row.get("参考加价率"))
        if not reference_key or (invoiced_quantity <= 0 and not markup_text):
            continue
        markup_rate = normalize_reference_markup_rate(markup_text)
        reference_amount = _to_decimal(row.get("金额(除税)合计")) or Decimal("0")
        reference_tax = _to_decimal(row.get("税金合计")) or Decimal("0")
        reference_total = _to_decimal(row.get("价税合计")) or Decimal("0")
        locked_total = _to_decimal(row.get("已开参考价税合计")) or Decimal("0")
        ratio = Decimal("0") if reference_total <= 0 else locked_total / reference_total
        result[reference_key] = {
            "reference_key": reference_key,
            "item_code": key[0],
            "item_name": key[1],
            "spec": key[2],
            "unit": key[3],
            "invoiced_quantity": _decimal_text(invoiced_quantity, "0.001"),
            "quantity_total": _decimal_text(_to_decimal(row.get("数量合计")) or Decimal("0"), "0.001"),
            "reference_amount_total": _decimal_text(reference_amount, "0.000001"),
            "reference_tax_total": _decimal_text(reference_tax, "0.000001"),
            "reference_total_with_tax": _decimal_text(reference_total, "0.000001"),
            "reference_markup_rate": _decimal_text(markup_rate, "0.000001"),
            "reference_markup_rate_percent": reference_markup_percent_text(markup_rate),
            "reference_markup_locked": False,
            "invoiced_reference_amount": _decimal_text(reference_amount * ratio, "0.000001"),
            "invoiced_reference_tax": _decimal_text(reference_tax * ratio, "0.000001"),
            "invoiced_reference_total_with_tax": _decimal_text(locked_total, "0.000001"),
            "updated_at": _clean_text(row.get("状态更新时间")),
        }
    return result


def _merge_reference_status(json_items: dict[str, dict], workbook_items: dict[str, dict]) -> tuple[dict[str, dict], bool]:
    merged = {key: dict(value) for key, value in (json_items or {}).items() if isinstance(value, dict)}
    changed = False
    for key, workbook_item in (workbook_items or {}).items():
        if not isinstance(workbook_item, dict):
            continue
        if key not in merged:
            merged[key] = dict(workbook_item)
            changed = True
            continue
        for field in (
            "reference_key",
            "item_code",
            "item_name",
            "spec",
            "unit",
            "quantity_total",
            "reference_amount_total",
            "reference_tax_total",
            "reference_total_with_tax",
            "reference_markup_rate",
            "reference_markup_rate_percent",
            "reference_markup_locked",
            "invoiced_reference_amount",
            "invoiced_reference_tax",
            "invoiced_reference_total_with_tax",
            "updated_at",
        ):
            if _clean_text(merged[key].get(field)) == "" and _clean_text(workbook_item.get(field)) != "":
                merged[key][field] = workbook_item[field]
                changed = True
    return merged, changed


def _reference_status_item(status_items: dict[str, dict] | None, key: str, *values: object) -> dict:
    if not isinstance(status_items, dict):
        return {}
    aliases = [key]
    if values:
        aliases.append(_legacy_plain_reference_key(*values))
        aliases.append("|".join(_reference_key_values(*values)))
    for alias in aliases:
        item = status_items.get(alias)
        if isinstance(item, dict):
            return item
    return {}


def _iter_cost_invoice_files(watch_folder: Path) -> list[Path]:
    files: list[Path] = []
    for root, dirs, names in os.walk(watch_folder, topdown=True):
        dirs[:] = [name for name in dirs if not any(name.startswith(prefix) for prefix in RUNTIME_DIR_PREFIXES)]
        root_path = Path(root)
        for name in names:
            path = root_path / name
            if path.name.startswith("~$"):
                continue
            if path.suffix.lower() in COST_SOURCE_EXTS:
                files.append(path)
    return sorted(files, key=lambda path: str(path).casefold())


def _iter_pdf_files(watch_folder: Path) -> list[Path]:
    return [path for path in _iter_cost_invoice_files(watch_folder) if path.suffix.lower() == ".pdf"]


def _cost_source_priority(path: Path) -> int:
    return {".xml": 0, ".ofd": 1, ".pdf": 2}.get(path.suffix.lower(), 99)


def _cost_family_key(path: Path, metadata: dict | None = None) -> str:
    metadata = metadata if isinstance(metadata, dict) else {}
    number = _valid_invoice_number(_metadata_value(metadata, "发票号码", "invoice_number"))
    if not number:
        number = _valid_invoice_number(path.stem)
    if number:
        return f"invoice:{number}"
    return f"path:{_canonical_path(path)}"


def _analysis_attempt_summary(attempts: list[dict]) -> str:
    parts = []
    for attempt in attempts:
        source_format = _clean_text(attempt.get("source_format")) or Path(_clean_text(attempt.get("source"))).suffix.lower().lstrip(".")
        status = _clean_text(attempt.get("status")) or "unknown"
        message = _clean_text(attempt.get("message"))
        parts.append(f"{source_format}:{status}{'(' + message + ')' if message else ''}")
    return " / ".join(parts)


def _select_cost_analysis(candidates: list[tuple[Path, dict]]) -> tuple[dict | None, list[dict]]:
    attempts: list[dict] = []
    for source_path, metadata in sorted(candidates, key=lambda item: (_cost_source_priority(item[0]), str(item[0]).casefold())):
        analysis = _finalize_cost_analysis(dict(analyze_cost_invoice(source_path, metadata=metadata)))
        analysis["source_format"] = source_path.suffix.lower().lstrip(".")
        analysis["source_path"] = str(source_path)
        attempts.append(analysis)
    if not attempts:
        return None, attempts
    selected = dict(
        max(
            attempts,
            key=lambda attempt: (
                int(attempt.get("validation_score") or 0),
                int(bool(attempt.get("rows"))),
                int(attempt.get("status") not in {"failed", "skipped"}),
                -_cost_source_priority(Path(f"candidate.{attempt.get('source_format') or ''}")),
            ),
        )
    )
    if len(attempts) > 1 and selected.get("status") != "ok":
        message = _clean_text(selected.get("message")) or "未识别到成本明细行"
        selected["message"] = f"{message}；同票候选：{_analysis_attempt_summary(attempts)}"
    return selected, attempts


def _seller_summary(rows: list[dict]) -> list[dict]:
    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[_clean_text(row.get("销售方")) or "未识别销售方"].append(row)
    result = []
    for seller, group_rows in sorted(grouped.items()):
        invoice_numbers = {_clean_text(row.get("发票号码")) for row in group_rows if _clean_text(row.get("发票号码"))}
        result.append(
            {
                "销售方": seller,
                "发票张数": len(invoice_numbers),
                "明细行数": len(group_rows),
                "数量合计": _number(_sum_decimal(group_rows, "数量"), "0.001"),
                "金额(除税)合计": _number(_sum_decimal(group_rows, "金额(除税)")),
                "税金合计": _number(_sum_decimal(group_rows, "税金")),
                "价税合计": _number(_sum_decimal(group_rows, "价税合计")),
            }
        )
    return result


def project_spec_summary(rows: list[dict]) -> list[dict]:
    grouped: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        key = (
            _clean_text(row.get("销售方")),
            _clean_text(row.get("发票代码(**内文字)")),
            _clean_text(row.get("内部项目名称")),
            _clean_text(row.get("规格型号")),
            _clean_text(row.get("单位")),
        )
        grouped[key].append(row)
    result = []
    for key, group_rows in sorted(grouped.items()):
        quantity = _sum_decimal(group_rows, "数量")
        amount = _sum_decimal(group_rows, "金额(除税)")
        total = _sum_decimal(group_rows, "价税合计")
        stock_avg_price = stock_average_unit_price(group_rows)
        stock_avg_price_with_tax = stock_average_unit_price_with_tax(group_rows)
        purchase_avg_price_with_tax = purchase_reference_average_unit_price_with_tax(group_rows)
        invoice_numbers = sorted({_clean_text(row.get("发票号码")) for row in group_rows if _clean_text(row.get("发票号码"))})
        result.append(
            {
                "销售方": key[0],
                "发票代码(**内文字)": key[1],
                "内部项目名称": key[2],
                "规格型号": key[3],
                "单位": key[4],
                "数量合计": _number(quantity, "0.001"),
                "平均单价(除税)": _number(stock_avg_price),
                "平均单价(含税)": _number(stock_avg_price_with_tax),
                "库存平均单价(除税)": _number(stock_avg_price),
                "库存平均单价(含税)": _number(stock_avg_price_with_tax),
                "采购参考平均单价(含税)": _number(purchase_avg_price_with_tax),
                "金额(除税)合计": _number(amount),
                "税金合计": _number(_sum_decimal(group_rows, "税金")),
                "价税合计": _number(total),
                "涉及发票号码": ",".join(invoice_numbers),
            }
        )
    return result


_project_spec_summary = project_spec_summary


def _invoice_reference_summary(
    rows: list[dict],
    status_items: dict[str, dict] | None = None,
    reference_markup_rate: object = DEFAULT_REFERENCE_MARKUP_RATE,
) -> list[dict]:
    fallback_markup_rate = normalize_reference_markup_rate(reference_markup_rate)
    grouped: dict[tuple, list[dict]] = defaultdict(list)
    for row in rows:
        key = (
            _clean_text(row.get("发票代码(**内文字)")),
            _clean_text(row.get("内部项目名称")),
            _clean_text(row.get("规格型号")),
            _clean_text(row.get("单位")),
        )
        grouped[key].append(row)
    result = []
    for key, group_rows in sorted(grouped.items()):
        quantity = _sum_decimal(group_rows, "数量")
        amount = _sum_decimal(group_rows, "金额(除税)")
        avg_price = stock_average_unit_price(group_rows)
        status_key = _reference_key(*key)
        status_item = _reference_status_item(status_items, status_key, *key)
        markup = _row_markup_meta(status_item, fallback_markup_rate)
        markup_multiplier = reference_markup_multiplier(markup["rate"])
        reference_avg_price_with_tax = reference_output_average_unit_price_with_tax(avg_price, markup["rate"])
        reference_base_amount = avg_price * quantity if avg_price is not None and quantity > 0 else Decimal("0")
        reference_amount = reference_base_amount * markup_multiplier
        reference_tax = reference_amount * REFERENCE_OUTPUT_TAX_RATE
        reference_total = reference_amount + reference_tax
        invoiced_quantity = _to_decimal(status_item.get("invoiced_quantity")) or Decimal("0")
        invoiced_quantity = _clamp_decimal(invoiced_quantity, Decimal("0"), quantity if quantity > 0 else Decimal("0"))
        uninvoiced_quantity = quantity - invoiced_quantity
        status_text = _invoice_status(quantity, invoiced_quantity)
        invoiced_reference_amount = _locked_reference_part(
            status_item,
            "invoiced_reference_amount",
            reference_amount,
            quantity,
            invoiced_quantity,
        )
        invoiced_reference_tax = _locked_reference_part(
            status_item,
            "invoiced_reference_tax",
            reference_tax,
            quantity,
            invoiced_quantity,
        )
        invoiced_reference_total = _locked_reference_part(
            status_item,
            "invoiced_reference_total_with_tax",
            reference_total,
            quantity,
            invoiced_quantity,
        )
        uninvoiced_reference_amount = _remaining_reference_part(reference_amount, invoiced_reference_amount)
        uninvoiced_reference_tax = _remaining_reference_part(reference_tax, invoiced_reference_tax)
        uninvoiced_reference_total = _remaining_reference_part(reference_total, invoiced_reference_total)
        result.append(
            {
                "发票代码(**内文字)": key[0],
                "内部项目名称": key[1],
                "规格型号": key[2],
                "单位": key[3],
                "数量合计": _number(quantity, "0.001"),
                "已开数量": _number(invoiced_quantity, "0.001"),
                "未开数量": _number(uninvoiced_quantity, "0.001"),
                "开票状态": status_text,
                "平均单价(除税)": _number(avg_price * markup_multiplier if avg_price is not None else None),
                "平均单价(含税)": _number(reference_avg_price_with_tax),
                "金额(除税)合计": _number(reference_amount),
                "税金合计": _number(reference_tax),
                "价税合计": _number(reference_total),
                "已开参考价税合计": _number(invoiced_reference_total),
                "未开参考金额(除税)": _number(uninvoiced_reference_amount),
                "未开参考税金": _number(uninvoiced_reference_tax),
                "未开参考价税合计": _number(uninvoiced_reference_total),
                "参考加价率": markup["display"],
                "状态更新时间": _clean_text(status_item.get("updated_at")),
            }
        )
    return result


def _check_rows(rows: list[dict], analyses: list[dict]) -> list[dict]:
    rows_by_source: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        rows_by_source[_clean_text(row.get("源文件"))].append(row)

    result = []
    for analysis in analyses:
        source = _clean_text(analysis.get("source"))
        invoice = dict(analysis.get("invoice") or {})
        detail_rows = rows_by_source.get(source, [])
        validation = _cost_validation(detail_rows, invoice)
        parsed_amount = validation["parsed_amount"]
        parsed_tax = validation["parsed_tax"]
        parsed_total = _sum_decimal(detail_rows, "价税合计")
        invoice_amount = validation["invoice_amount"]
        invoice_tax = validation["invoice_tax"]
        diff_amount = parsed_amount - invoice_amount if invoice_amount is not None else None
        diff_tax = parsed_tax - invoice_tax if invoice_tax is not None else None
        classification_ok = _clean_text(invoice.get("类型识别状态")) == "ok"
        ok = validation["amount_ok"] and validation["tax_ok"] and classification_ok
        status = "通过" if ok else "待核对"
        classification_issue = _clean_text(invoice.get("类型识别说明"))
        if not classification_ok and not classification_issue:
            classification_issue = "发票类型识别待核对"
        message = _merge_issue_text(analysis.get("message"), validation["issue"], classification_issue)
        result.append(
            {
                "源文件": source,
                "发票大类": invoice.get("发票大类", ""),
                "特定业务类型": invoice.get("特定业务类型", ""),
                "类型识别状态": invoice.get("类型识别状态", ""),
                "类型识别说明": invoice.get("类型识别说明", ""),
                "销售方": invoice.get("销售方", ""),
                "发票号码": invoice.get("发票号码", ""),
                "开票日期": invoice.get("开票日期", "") or invoice.get("开票时间", ""),
                "明细行数": len(detail_rows),
                "发票金额(除税)": _number(invoice_amount),
                "解析金额(除税)": _number(parsed_amount),
                "差异(除税)": _number(diff_amount),
                "发票税金": _number(invoice_tax),
                "解析税金": _number(parsed_tax),
                "差异(税金)": _number(diff_tax),
                "价税合计": _number(parsed_total),
                "校验状态": status,
                "说明": message,
            }
        )
    return result


def _write_csv(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DETAIL_HEADERS)
        writer.writeheader()
        for row in detail_rows_with_tax_average(rows):
            writer.writerow({header: row.get(header, "") for header in DETAIL_HEADERS})


DETAIL_INVOICE_MERGE_HEADERS = ("销售方", "购买方", "发票号码", "开票日期", "备注项目名称", "源文件")


def _detail_invoice_group_key(row: dict) -> str:
    invoice_number = _clean_text(row.get("发票号码"))
    source_file = _clean_text(row.get("源文件"))
    if invoice_number:
        return f"invoice:{invoice_number}\x1fsource:{source_file}"
    if source_file:
        return f"source:{source_file}"
    return ""


def _merge_detail_invoice_cells(ws, headers: list[str], rows: list[dict]) -> None:
    if len(rows) < 2:
        return
    merge_columns = [headers.index(header) + 1 for header in DETAIL_INVOICE_MERGE_HEADERS if header in headers]
    if not merge_columns:
        return

    def merge_range(start_row: int, end_row: int, key: str) -> None:
        if not key or end_row <= start_row:
            return
        for col_idx in merge_columns:
            ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            ws.cell(row=start_row, column=col_idx).alignment = Alignment(vertical="center", wrap_text=True)

    current_key = _detail_invoice_group_key(rows[0])
    group_start = 2
    for row_offset, row in enumerate(rows[1:], start=3):
        key = _detail_invoice_group_key(row)
        if key != current_key:
            merge_range(group_start, row_offset - 1, current_key)
            current_key = key
            group_start = row_offset
    merge_range(group_start, len(rows) + 1, current_key)


def _append_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title=title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header, "") for header in headers])

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "#,##0.00"
    widths = {
        "销售方": 32,
        "购买方": 32,
        "发票号码": 24,
        "开票日期": 12,
        "备注项目名称": 32,
        "内部项目名称": 14,
        "规格型号": 12,
        "单位": 10,
        "数量": 12,
        "单价(除税)": 14,
        "平均单价(除税)": 16,
        "平均单价(含税)": 16,
        "库存平均单价(除税)": 18,
        "库存平均单价(含税)": 18,
        "采购参考平均单价(含税)": 20,
        "金额(除税)": 16,
        "金额(除税)合计": 18,
        "税率": 10,
        "税金": 14,
        "税金合计": 14,
        "价税合计": 16,
        "已开参考价税合计": 18,
        "发票代码(**内文字)": 22,
        "源文件": 48,
        "发票大类": 20,
        "特定业务类型": 22,
        "类型识别状态": 16,
        "类型识别说明": 42,
        "校验状态": 12,
        "说明": 36,
    }
    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = widths.get(header, 16)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    if title == "发票明细":
        _merge_detail_invoice_cells(ws, headers, rows)


def _write_workbook(
    path: Path,
    detail_rows: list[dict],
    check_rows: list[dict],
    status_items: dict[str, dict] | None = None,
    reference_markup_rate: object = DEFAULT_REFERENCE_MARKUP_RATE,
) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    _append_sheet(wb, "发票明细", DETAIL_HEADERS, detail_rows_with_tax_average(detail_rows))
    _append_sheet(wb, "销售方汇总", SELLER_SUMMARY_HEADERS, _seller_summary(detail_rows))
    _append_sheet(wb, "项目规格汇总", PROJECT_SPEC_SUMMARY_HEADERS, _project_spec_summary(detail_rows))
    _append_sheet(
        wb,
        "开票参考",
        INVOICE_REFERENCE_HEADERS,
        _invoice_reference_summary(detail_rows, status_items=status_items, reference_markup_rate=reference_markup_rate),
    )
    _append_sheet(wb, "发票校验", CHECK_HEADERS, check_rows)
    wb.save(path)


def build_cost_analysis_outputs(
    watch_folder: Path,
    output_folder: Path,
    invoice_metadata: object = None,
    reference_markup_rate: object = DEFAULT_REFERENCE_MARKUP_RATE,
) -> dict:
    watch_folder = Path(watch_folder)
    output_folder = Path(output_folder)
    metadata_index = _normalize_metadata_index(invoice_metadata)
    analyses = []
    detail_rows: list[dict] = []
    all_attempts: list[dict] = []
    source_files = _iter_cost_invoice_files(watch_folder)
    format_counts = {
        "pdf": len([path for path in source_files if path.suffix.lower() == ".pdf"]),
        "ofd": len([path for path in source_files if path.suffix.lower() == ".ofd"]),
        "xml": len([path for path in source_files if path.suffix.lower() == ".xml"]),
    }
    source_groups: dict[str, list[tuple[Path, dict]]] = defaultdict(list)

    for source_path in source_files:
        metadata = metadata_index.get(_canonical_path(source_path), {})
        source_groups[_cost_family_key(source_path, metadata)].append((source_path, metadata))

    for candidates in source_groups.values():
        analysis, attempts = _select_cost_analysis(candidates)
        all_attempts.extend(attempts)
        if analysis is None:
            continue
        analyses.append(analysis)
        detail_rows.extend(analysis.get("rows") or [])

    check_rows = _check_rows(detail_rows, analyses)
    detail_csv = output_folder / COST_DETAIL_CSV_NAME
    summary_xlsx = output_folder / COST_SUMMARY_XLSX_NAME
    status_json = output_folder / COST_REFERENCE_STATUS_NAME
    status_payload = {}
    if status_json.exists():
        try:
            raw_payload = json.loads(status_json.read_text(encoding="utf-8-sig"))
            status_payload = raw_payload if isinstance(raw_payload, dict) else {}
        except Exception:
            status_payload = {}
    markup_rate = normalize_reference_markup_rate(status_payload.get("reference_markup_rate", reference_markup_rate))
    markup_locked = _truthy(status_payload.get("reference_markup_locked", False))
    status_items, status_changed = _merge_reference_status(_load_reference_status(status_json), _load_workbook_reference_status(summary_xlsx))
    if status_changed:
        _write_reference_status(status_json, status_items, reference_markup_rate=markup_rate, reference_markup_locked=markup_locked)
    _write_csv(detail_csv, detail_rows)
    _write_workbook(summary_xlsx, detail_rows, check_rows, status_items=status_items, reference_markup_rate=markup_rate)

    failed = [item for item in analyses if item.get("status") == "failed"]
    skipped = [item for item in analyses if item.get("status") == "skipped"]
    needs_review = [item for item in analyses if item.get("status") == "needs_review"]
    no_detail = [item for item in analyses if not item.get("rows")]
    return {
        "detail_csv": str(detail_csv),
        "summary_xlsx": str(summary_xlsx),
        "reference_status_json": str(status_json),
        "pdf_count": format_counts["pdf"],
        "source_invoice_count": len(source_groups),
        "analyzed_source_count": len(analyses),
        "attempted_source_count": len(all_attempts),
        "format_counts": format_counts,
        "detail_count": len(detail_rows),
        "seller_count": len({row.get("销售方", "") for row in detail_rows if row.get("销售方", "")}),
        "check_count": len(check_rows),
        "check_mismatch_count": len([row for row in check_rows if row.get("校验状态") != "通过"]),
        "failed_count": len(failed),
        "skipped_count": len(skipped),
        "needs_review_count": len(needs_review),
        "no_detail_count": len(no_detail),
    }
