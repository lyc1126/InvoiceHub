from __future__ import annotations

import os
import re
from copy import copy
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from invoice_hub.extraction import extract_invoice_record, supported_invoice_files
from invoice_hub.projections.cost_analysis import analyze_cost_invoice


DOCUMENT_TEMPLATE_DIR = Path(__file__).resolve().parent / "document_templates"
INBOUND_TEMPLATE_PATH = DOCUMENT_TEMPLATE_DIR / "电子入库单模板.xlsx"
OUTBOUND_TEMPLATE_PATH = DOCUMENT_TEMPLATE_DIR / "电子出库单模板.xlsx"

INBOUND_DEFAULT_FIELDS = ("采购员", "负责人", "仓管员", "制表人")
OUTBOUND_DEFAULT_FIELDS = ("收货单位", "地址", "电话", "联系人", "编辑人", "收货人", "项目负责人")

SUPPORTED_DOCUMENT_INVOICE_EXTS = {".pdf", ".ofd", ".xml"}

RMB_DIGITS = "零壹贰叁肆伍陆柒捌玖"
RMB_UNITS = ("", "拾", "佰", "仟")
RMB_BIG_UNITS = ("", "万", "亿", "兆")
RMB_FRACTION_UNITS = ("角", "分")


class DocumentError(ValueError):
    pass


def clean_document_defaults(payload: dict | None = None) -> dict[str, dict[str, str]]:
    payload = payload if isinstance(payload, dict) else {}
    return {
        "inbound": {key: _clean_text((payload.get("inbound") or {}).get(key)) for key in INBOUND_DEFAULT_FIELDS},
        "outbound": {key: _clean_text((payload.get("outbound") or {}).get(key)) for key in OUTBOUND_DEFAULT_FIELDS},
    }


def merge_document_defaults(saved: dict | None, current: dict | None = None) -> dict[str, dict[str, str]]:
    result = clean_document_defaults(saved)
    current = current if isinstance(current, dict) else {}
    for kind, allowed in (("inbound", INBOUND_DEFAULT_FIELDS), ("outbound", OUTBOUND_DEFAULT_FIELDS)):
        source = current.get(kind)
        if not isinstance(source, dict):
            continue
        for key in allowed:
            if key in source:
                result[kind][key] = _clean_text(source.get(key))
    return result


def inbound_invoice_options(detail_rows: list[dict]) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}
    for row in detail_rows:
        number = _clean_text(row.get("发票号码"))
        if not number:
            continue
        item = grouped.setdefault(
            number,
            {
                "invoice_number": number,
                "invoice_date": "",
                "seller": "",
                "buyer": "",
                "row_count": 0,
                "total_with_tax": Decimal("0"),
            },
        )
        item["row_count"] += 1
        if not item["invoice_date"]:
            item["invoice_date"] = _clean_text(row.get("开票日期"))
        if not item["seller"]:
            item["seller"] = _clean_text(row.get("销售方"))
        if not item["buyer"]:
            item["buyer"] = _clean_text(row.get("购买方"))
        total = _line_total(row)
        if total is not None:
            item["total_with_tax"] += total
    result = []
    for item in grouped.values():
        total = item.pop("total_with_tax")
        item["total_with_tax"] = _money_text(total)
        item["label"] = _invoice_option_label(item)
        result.append(item)
    return sorted(result, key=lambda item: (str(item.get("invoice_date") or ""), str(item.get("invoice_number") or "")), reverse=True)


def build_inbound_preview(detail_rows: list[dict], invoice_number: str, defaults: dict | None = None) -> dict[str, Any]:
    number = _clean_text(invoice_number)
    if not number:
        raise DocumentError("发票号码不能为空")
    rows = [dict(row) for row in detail_rows if _clean_text(row.get("发票号码")) == number]
    if not rows:
        raise KeyError(number)
    defaults = clean_document_defaults(defaults).get("inbound", {})
    preview_rows = []
    total_with_tax = Decimal("0")
    for row in rows:
        line_total = _line_total(row)
        if line_total is not None:
            total_with_tax += line_total
        preview_rows.append(
            {
                "code": "",
                "item_name": _clean_text(row.get("内部项目名称")),
                "spec": _clean_text(row.get("规格型号")),
                "unit": _clean_text(row.get("单位")),
                "quantity": _quantity_text(row.get("数量")),
                "unit_price": _money_text_or_blank(row.get("单价(除税)")),
                "amount": _money_text_or_blank(row.get("金额(除税)")),
                "tax_amount": _money_text_or_blank(row.get("税金")),
                "tax_rate": _tax_rate_text(row.get("税率")),
                "remark": "",
                "total_with_tax": _money_text_or_blank(line_total),
            }
        )
    first = rows[0]
    return {
        "ok": True,
        "document_type": "inbound",
        "title": "入库单",
        "invoice_number": number,
        "invoice_date": _clean_text(first.get("开票日期")),
        "supplier": _clean_text(first.get("销售方")),
        "buyer": _clean_text(first.get("购买方")),
        "rows": preview_rows,
        "row_count": len(preview_rows),
        "total_with_tax": _money_text(total_with_tax),
        "total_with_tax_upper": rmb_uppercase(total_with_tax),
        "defaults": defaults,
    }


def outbound_invoice_options(outbound_dir: Path | None) -> list[dict[str, Any]]:
    if not outbound_dir:
        return []
    directory = Path(outbound_dir)
    if not directory.exists() or not directory.is_dir():
        return []
    grouped: dict[str, dict[str, Any]] = {}
    for path in _iter_supported_invoice_files(directory):
        try:
            record = extract_invoice_record(path)
        except Exception:
            continue
        number = _clean_text(record.invoice_number) or _invoice_number_from_name(path.name)
        if not number:
            continue
        item = grouped.setdefault(
            number,
            {
                "invoice_number": number,
                "invoice_date": "",
                "seller": "",
                "buyer": "",
                "amount": "",
                "file_count": 0,
                "formats": [],
                "source_files": [],
            },
        )
        item["file_count"] += 1
        if record.file_type and record.file_type not in item["formats"]:
            item["formats"].append(record.file_type)
        item["source_files"].append(str(path))
        if not item["invoice_date"]:
            item["invoice_date"] = _clean_text(record.invoice_date)
        if not item["seller"]:
            item["seller"] = _clean_text(record.seller)
        if not item["buyer"]:
            item["buyer"] = _clean_text(record.buyer)
        if not item["amount"]:
            item["amount"] = _money_text_or_blank(record.amount)
    result = []
    for item in grouped.values():
        item["formats"] = sorted(item["formats"])
        item["label"] = _invoice_option_label(item)
        result.append(item)
    return sorted(result, key=lambda item: (str(item.get("invoice_date") or ""), str(item.get("invoice_number") or "")), reverse=True)


def build_outbound_preview(outbound_dir: Path, invoice_number: str, defaults: dict | None = None) -> dict[str, Any]:
    number = _clean_text(invoice_number)
    if not number:
        raise DocumentError("发票号码不能为空")
    directory = Path(outbound_dir)
    if not directory.exists() or not directory.is_dir():
        raise DocumentError(f"开具发票目录不可用: {directory}")

    selected: tuple[Path, dict, dict] | None = None
    first_match: tuple[Path, dict, dict] | None = None
    for path in _matching_invoice_files(directory, number):
        try:
            record = extract_invoice_record(path)
        except Exception:
            record = None
        metadata = _metadata_from_record(record, fallback_invoice_number=number)
        analysis = dict(analyze_cost_invoice(path, metadata=metadata))
        if first_match is None:
            first_match = (path, metadata, analysis)
        if analysis.get("rows"):
            selected = (path, metadata, analysis)
            break
    if selected is None:
        selected = first_match
    if selected is None:
        raise KeyError(number)

    source_path, metadata, analysis = selected
    defaults = clean_document_defaults(defaults).get("outbound", {})
    preview_rows = []
    total_with_tax = Decimal("0")
    for index, row in enumerate(analysis.get("rows") or [], start=1):
        line_total = _line_total(row)
        if line_total is not None:
            total_with_tax += line_total
        preview_rows.append(
            {
                "index": index,
                "item_name": _clean_text(row.get("内部项目名称")),
                "spec": _clean_text(row.get("规格型号")),
                "unit": _clean_text(row.get("单位")),
                "quantity": _quantity_text(row.get("数量")),
                "unit_price": _money_text_or_blank(_line_unit_price_with_tax(row)),
                "amount": _money_text_or_blank(line_total),
                "remark": "",
            }
        )

    invoice_date = _clean_text(metadata.get("开票日期")) or _first_row_text(analysis.get("rows") or [], "开票日期")
    return {
        "ok": True,
        "document_type": "outbound",
        "title": "出库单",
        "invoice_number": number,
        "invoice_date": invoice_date,
        "seller": _clean_text(metadata.get("销售方")) or _first_row_text(analysis.get("rows") or [], "销售方"),
        "buyer": _clean_text(metadata.get("购买方")) or _first_row_text(analysis.get("rows") or [], "购买方"),
        "source_file": source_path.name,
        "source_path": str(source_path),
        "source_format": source_path.suffix.lower().lstrip("."),
        "rows": preview_rows,
        "row_count": len(preview_rows),
        "message": _clean_text(analysis.get("message")),
        "total_with_tax": _money_text(total_with_tax),
        "total_with_tax_upper": rmb_uppercase(total_with_tax),
        "defaults": defaults,
    }


def inbound_export_path(watch_dir: Path, preview: dict[str, Any]) -> Path:
    return Path(watch_dir) / "入库单" / f"入库单-{_safe_filename_part(preview.get('invoice_number'))}-{_safe_date_part(preview.get('invoice_date'))}.xlsx"


def outbound_export_path(outbound_dir: Path, preview: dict[str, Any]) -> Path:
    return Path(outbound_dir) / "出库单" / f"出库单-{_safe_filename_part(preview.get('invoice_number'))}-{_safe_date_part(preview.get('invoice_date'))}.xlsx"


def write_inbound_workbook(preview: dict[str, Any], output_path: Path) -> Path:
    _ensure_template(INBOUND_TEMPLATE_PATH)
    wb = load_workbook(INBOUND_TEMPLATE_PATH)
    try:
        ws = wb["入库单"] if "入库单" in wb.sheetnames else wb.active
        rows = list(preview.get("rows") or [])
        extra = _ensure_detail_rows(ws, detail_start=5, fixed_detail_rows=11, summary_row=16, detail_count=len(rows))
        summary_row = 16 + extra
        footer_row = 17 + extra

        ws["B3"] = _clean_text(preview.get("supplier"))
        ws["E3"] = _clean_text(preview.get("invoice_date"))
        ws["G3"] = f"NO：{_clean_text(preview.get('invoice_number'))}"

        _clear_range(ws, 5, 5 + max(11, len(rows)) - 1, 1, 10)
        for offset, row in enumerate(rows):
            target = 5 + offset
            values = [
                "",
                row.get("item_name", ""),
                row.get("spec", ""),
                row.get("unit", ""),
                row.get("quantity", ""),
                row.get("unit_price", ""),
                row.get("amount", ""),
                row.get("tax_amount", ""),
                row.get("tax_rate", ""),
                row.get("remark", ""),
            ]
            _write_row(ws, target, values, money_columns={6, 7, 8}, quantity_columns={5})

        total = _decimal_value(preview.get("total_with_tax")) or Decimal("0")
        _write_inbound_summary(ws, summary_row, total)
        defaults = preview.get("defaults") if isinstance(preview.get("defaults"), dict) else {}
        ws.cell(footer_row, 2).value = _clean_text(defaults.get("采购员"))
        ws.cell(footer_row, 4).value = _clean_text(defaults.get("负责人"))
        ws.cell(footer_row, 7).value = _clean_text(defaults.get("仓管员"))
        ws.cell(footer_row, 10).value = _clean_text(defaults.get("制表人"))
        _save_workbook(wb, Path(output_path))
    finally:
        wb.close()
    return Path(output_path)


def write_outbound_workbook(preview: dict[str, Any], output_path: Path) -> Path:
    _ensure_template(OUTBOUND_TEMPLATE_PATH)
    wb = load_workbook(OUTBOUND_TEMPLATE_PATH)
    try:
        ws = wb["出库单"] if "出库单" in wb.sheetnames else wb.active
        rows = list(preview.get("rows") or [])
        extra = _ensure_detail_rows(ws, detail_start=6, fixed_detail_rows=10, summary_row=16, detail_count=len(rows))
        summary_row = 16 + extra
        footer_row = 18 + extra
        defaults = preview.get("defaults") if isinstance(preview.get("defaults"), dict) else {}

        ws["A3"] = f"收货单位：{_clean_text(defaults.get('收货单位'))}"
        ws["C3"] = f"开单日期：{_clean_text(preview.get('invoice_date'))}"
        ws["F3"] = f"单据编号：{_clean_text(preview.get('invoice_number'))}"
        ws["A4"] = f"地    址：{_clean_text(defaults.get('地址'))}"
        ws["C4"] = f"电    话：{_clean_text(defaults.get('电话'))}"
        ws["F4"] = f"联 系 人：{_clean_text(defaults.get('联系人'))}"

        _clear_range(ws, 6, 6 + max(10, len(rows)) - 1, 1, 8)
        for offset, row in enumerate(rows):
            target = 6 + offset
            values = [
                row.get("index", offset + 1),
                row.get("item_name", ""),
                row.get("spec", ""),
                row.get("unit", ""),
                row.get("quantity", ""),
                row.get("unit_price", ""),
                row.get("amount", ""),
                row.get("remark", ""),
            ]
            _write_row(ws, target, values, money_columns={6, 7}, quantity_columns={5})

        total = _decimal_value(preview.get("total_with_tax")) or Decimal("0")
        _write_outbound_summary(ws, summary_row, total)
        ws.cell(footer_row, 2).value = _clean_text(defaults.get("编辑人"))
        ws.cell(footer_row, 5).value = _clean_text(defaults.get("收货人"))
        ws.cell(footer_row, 8).value = _clean_text(defaults.get("项目负责人"))
        _save_workbook(wb, Path(output_path))
    finally:
        wb.close()
    return Path(output_path)


def _write_inbound_summary(ws, row_index: int, total: Decimal) -> None:
    ws.cell(row_index, 1).value = "合计（大写）"
    ws.cell(row_index, 6).value = "合计（小写）"
    upper_cell = _prepare_merged_value_cell(ws, row_index, 2, 5)
    amount_cell = _prepare_merged_value_cell(ws, row_index, 7, 10)
    upper_cell.value = rmb_uppercase(total)
    amount_cell.value = total
    amount_cell.number_format = "#,##0.00"
    _set_horizontal_alignment(amount_cell, "left")


def _write_outbound_summary(ws, row_index: int, total: Decimal) -> None:
    ws.cell(row_index, 1).value = "合计(大写)"
    ws.cell(row_index, 5).value = "合计（小写）"
    upper_cell = _prepare_merged_value_cell(ws, row_index, 2, 4)
    amount_cell = _prepare_merged_value_cell(ws, row_index, 6, 8)
    upper_cell.value = rmb_uppercase(total)
    amount_cell.value = total
    amount_cell.number_format = "#,##0.00"
    _set_horizontal_alignment(amount_cell, "left")


def _prepare_merged_value_cell(ws, row_index: int, min_col: int, max_col: int):
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row <= row_index <= merged_range.max_row and not (
            merged_range.max_col < min_col or merged_range.min_col > max_col
        ):
            ws.unmerge_cells(str(merged_range))
    for column in range(min_col, max_col + 1):
        ws.cell(row_index, column).value = None
    if max_col > min_col:
        ws.merge_cells(
            f"{get_column_letter(min_col)}{row_index}:"
            f"{get_column_letter(max_col)}{row_index}"
        )
    return ws.cell(row_index, min_col)


def rmb_uppercase(value: object) -> str:
    number = _decimal_value(value) or Decimal("0")
    if number < 0:
        return "负" + rmb_uppercase(-number)
    cents = int((number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) * 100).to_integral_value())
    if cents == 0:
        return "人民币零元整"
    integer = cents // 100
    fraction = cents % 100
    result = "人民币" + _integer_to_rmb(integer) + "元"
    jiao = fraction // 10
    fen = fraction % 10
    if jiao == 0 and fen == 0:
        return result + "整"
    if jiao:
        result += RMB_DIGITS[jiao] + RMB_FRACTION_UNITS[0]
    elif integer:
        result += "零"
    if fen:
        result += RMB_DIGITS[fen] + RMB_FRACTION_UNITS[1]
    return result


def _integer_to_rmb(value: int) -> str:
    if value == 0:
        return "零"
    groups: list[int] = []
    while value:
        groups.append(value % 10000)
        value //= 10000
    parts: list[str] = []
    zero_pending = False
    for group_index in range(len(groups) - 1, -1, -1):
        group = groups[group_index]
        if group == 0:
            zero_pending = bool(parts)
            continue
        if parts and (zero_pending or group < 1000):
            parts.append("零")
            zero_pending = False
        text = _four_digit_to_rmb(group)
        if group_index:
            text += RMB_BIG_UNITS[group_index]
        parts.append(text)
    return "".join(parts).rstrip("零") or "零"


def _four_digit_to_rmb(value: int) -> str:
    text = ""
    zero_pending = False
    digits = [value // 1000, value // 100 % 10, value // 10 % 10, value % 10]
    for index, digit in enumerate(digits):
        unit = RMB_UNITS[3 - index]
        if digit == 0:
            if text:
                zero_pending = True
            continue
        if zero_pending:
            text += "零"
            zero_pending = False
        text += RMB_DIGITS[digit] + unit
    return text


def _clean_text(value: object) -> str:
    return str(value or "").replace("\u00a0", " ").strip()


def _decimal_value(value: object) -> Decimal | None:
    if isinstance(value, Decimal):
        return value if value.is_finite() else None
    text = _clean_text(value).replace(",", "").replace("¥", "").replace("￥", "").replace("%", "")
    if not text:
        return None
    try:
        number = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    return number if number.is_finite() else None


def _money_text(value: object) -> str:
    number = _decimal_value(value) or Decimal("0")
    return format(number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")


def _money_text_or_blank(value: object) -> str:
    number = _decimal_value(value)
    if number is None:
        return ""
    return format(number.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP), "f")


def _quantity_text(value: object) -> str:
    number = _decimal_value(value)
    if number is None:
        return ""
    rounded = number.quantize(Decimal("0.001"), rounding=ROUND_HALF_UP)
    text = format(rounded, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def _tax_rate_text(value: object) -> str:
    text = _clean_text(value).replace("％", "%")
    if not text:
        return ""
    if text.endswith("%"):
        number = _decimal_value(text[:-1])
        return f"{_quantity_text(number)}%" if number is not None else text
    number = _decimal_value(text)
    if number is None:
        return text
    if Decimal("0") <= number <= Decimal("1"):
        number *= Decimal("100")
    return f"{_quantity_text(number)}%"


def _line_total(row: dict) -> Decimal | None:
    amount = _decimal_value(row.get("金额(除税)"))
    tax = _decimal_value(row.get("税金"))
    if amount is None or tax is None:
        return None
    return amount + tax


def _line_unit_price_with_tax(row: dict) -> Decimal | None:
    quantity = _decimal_value(row.get("数量"))
    total = _line_total(row)
    if quantity is None or quantity <= 0 or total is None:
        return None
    return total / quantity


def _metadata_from_record(record: object, fallback_invoice_number: str = "") -> dict[str, str]:
    if record is None:
        return {"发票号码": fallback_invoice_number}
    return {
        "销售方": _clean_text(getattr(record, "seller", "")),
        "购买方": _clean_text(getattr(record, "buyer", "")),
        "发票号码": _clean_text(getattr(record, "invoice_number", "")) or fallback_invoice_number,
        "开票日期": _clean_text(getattr(record, "invoice_date", "")),
        "开票金额": _clean_text(getattr(record, "amount", "")),
        "除税价": _clean_text(getattr(record, "pretax_amount", "")),
        "税金": _clean_text(getattr(record, "tax_amount", "")),
        "税率": _clean_text(getattr(record, "tax_rate", "")),
    }


def _iter_supported_invoice_files(directory: Path) -> list[Path]:
    return [path for path in supported_invoice_files(directory) if path.suffix.lower() in SUPPORTED_DOCUMENT_INVOICE_EXTS]


def _matching_invoice_files(directory: Path, invoice_number: str) -> list[Path]:
    matches = []
    for path in _iter_supported_invoice_files(directory):
        try:
            record = extract_invoice_record(path)
            number = _clean_text(record.invoice_number) or _invoice_number_from_name(path.name)
        except Exception:
            number = _invoice_number_from_name(path.name)
        if number == invoice_number:
            matches.append(path)
    return sorted(matches, key=lambda path: (_source_priority(path), str(path).casefold()))


def _source_priority(path: Path) -> int:
    return {".xml": 0, ".ofd": 1, ".pdf": 2}.get(path.suffix.lower(), 99)


def _invoice_number_from_name(value: object) -> str:
    match = re.search(r"(?<!\d)(\d{8,20})(?!\d)", _clean_text(value))
    return match.group(1) if match else ""


def _invoice_option_label(item: dict[str, Any]) -> str:
    number = _clean_text(item.get("invoice_number"))
    date = _clean_text(item.get("invoice_date"))
    seller = _clean_text(item.get("seller"))
    count = _clean_text(item.get("row_count") or item.get("file_count"))
    parts = [part for part in (number, date, seller) if part]
    if count:
        parts.append(f"{count}行" if "row_count" in item else f"{count}个文件")
    return " · ".join(parts) or number


def _first_row_text(rows: Iterable[dict], key: str) -> str:
    for row in rows:
        text = _clean_text(row.get(key))
        if text:
            return text
    return ""


def _safe_filename_part(value: object) -> str:
    text = _clean_text(value) or "未识别"
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", text)
    text = re.sub(r"\s+", "_", text).strip("._ ")
    return text[:80] or "未识别"


def _safe_date_part(value: object) -> str:
    return _safe_filename_part(value or "未识别日期")


def _ensure_template(path: Path) -> None:
    if not path.exists():
        raise DocumentError(f"单据模板不存在: {path}")


def _ensure_detail_rows(ws, detail_start: int, fixed_detail_rows: int, summary_row: int, detail_count: int) -> int:
    extra = max(0, int(detail_count) - fixed_detail_rows)
    if extra <= 0:
        return 0
    shifted_merges = []
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= summary_row:
            shifted_merges.append(
                (
                    merged_range.min_col,
                    merged_range.min_row + extra,
                    merged_range.max_col,
                    merged_range.max_row + extra,
                )
            )
            ws.unmerge_cells(str(merged_range))
    ws.insert_rows(summary_row, extra)
    for min_col, min_row, max_col, max_row in shifted_merges:
        ws.merge_cells(
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{max_row}"
        )
    template_row = summary_row - 1
    for row in range(summary_row, summary_row + extra):
        _copy_row_style(ws, template_row, row)
    return extra


def _copy_row_style(ws, source_row: int, target_row: int) -> None:
    ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        if source.font:
            target.font = copy(source.font)
        if source.fill:
            target.fill = copy(source.fill)
        if source.border:
            target.border = copy(source.border)
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.number_format:
            target.number_format = source.number_format
        if source.protection:
            target.protection = copy(source.protection)


def _clear_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row, col).value = ""


def _write_row(ws, row_index: int, values: list[Any], money_columns: set[int], quantity_columns: set[int]) -> None:
    for offset, value in enumerate(values, start=1):
        cell = ws.cell(row_index, offset)
        if offset in money_columns:
            cell.value = _decimal_for_cell(value)
            cell.number_format = "#,##0.00"
        elif offset in quantity_columns:
            cell.value = _decimal_for_cell(value)
            cell.number_format = "0.###"
        else:
            cell.value = value


def _set_horizontal_alignment(cell, horizontal: str) -> None:
    current = cell.alignment or Alignment()
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=current.vertical,
        text_rotation=current.textRotation,
        wrap_text=current.wrap_text,
        shrink_to_fit=current.shrink_to_fit,
        indent=current.indent,
        relativeIndent=current.relativeIndent,
        justifyLastLine=current.justifyLastLine,
        readingOrder=current.readingOrder,
    )


def _decimal_for_cell(value: object) -> Decimal | str:
    number = _decimal_value(value)
    return number if number is not None else ""


def _save_workbook(wb, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    temp_path = output_path.with_name(f".{output_path.stem}.{os.getpid()}.tmp.xlsx")
    wb.save(temp_path)
    os.replace(temp_path, output_path)
