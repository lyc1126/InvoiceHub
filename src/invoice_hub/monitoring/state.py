from __future__ import annotations

import errno
import json
import os
import shutil
import subprocess
import threading
import time
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterator

from openpyxl import load_workbook

from invoice_hub.domain import TargetProfile
from invoice_hub.domain.models import utc_now_text
from invoice_hub.extraction.parsers import SUPPORTED_EXTS, supported_invoice_files
from invoice_hub.projections.summary import SUMMARY_HEADERS, write_summary_xlsx
from invoice_hub.storage.files import atomic_write_json, read_csv_rows, read_json_object, write_csv_rows

STOP_FLAG_NAME = ".invoice_stop"
LOCK_FILE_NAME = ".invoice_monitor.lock"
SYNC_WRITE_LOCK_FILE_NAME = ".invoice_sync.lock"
PROCESSED_FILE_NAME = "processed_files.json"
MANUAL_OVERRIDES_FILE_NAME = "manual_overrides.json"
STATUS_FILE_NAME = "monitor_status.json"
BUSINESS_LOG_NAME = "文件变化监控日志.txt"
EDITABLE_FIELDS = ("销售方", "开票金额", "发票号码")
SYNC_WRITE_LOCK_POLL_SECONDS = 0.1


_SYNC_THREAD_LOCKS_GUARD = threading.Lock()
_SYNC_THREAD_LOCKS: dict[str, threading.RLock] = {}
_SYNC_HELD_LOCKS = threading.local()


def _sync_thread_lock(key: str) -> threading.RLock:
    with _SYNC_THREAD_LOCKS_GUARD:
        return _SYNC_THREAD_LOCKS.setdefault(key, threading.RLock())


def _is_windows_lock_contention(exc: OSError) -> bool:
    return (
        exc.errno in {errno.EACCES, errno.EAGAIN, errno.EDEADLK}
        or getattr(exc, "winerror", None) in {32, 33}
    )


def _acquire_sync_os_lock(handle) -> None:
    if os.name == "nt":
        import msvcrt

        # Do not read the byte before taking the lock: Windows denies that
        # read while another process holds it. File metadata is sufficient to
        # initialize the one-byte lock region, then every contender locks the
        # same byte at offset zero.
        handle.seek(0, os.SEEK_END)
        if handle.tell() == 0:
            handle.write(b"0")
            handle.flush()
        handle.seek(0)
        # LK_LOCK gives up after a fixed retry window. Keep retrying a
        # non-blocking request while the owning synchronizer is alive so a
        # long rebuild remains serialized instead of becoming a false failure.
        while True:
            try:
                msvcrt.locking(handle.fileno(), msvcrt.LK_NBLCK, 1)
                return
            except OSError as exc:
                if not _is_windows_lock_contention(exc):
                    raise
                time.sleep(SYNC_WRITE_LOCK_POLL_SECONDS)
        return
    import fcntl

    fcntl.flock(handle.fileno(), fcntl.LOCK_EX)


def _release_sync_os_lock(handle) -> None:
    if os.name == "nt":
        import msvcrt

        handle.seek(0)
        msvcrt.locking(handle.fileno(), msvcrt.LK_UNLCK, 1)
        return
    import fcntl

    fcntl.flock(handle.fileno(), fcntl.LOCK_UN)


def is_pid_alive(pid: int | str | None) -> bool:
    try:
        value = int(pid or 0)
    except Exception:
        return False
    if value <= 0:
        return False
    if os.name != "nt":
        try:
            os.kill(value, 0)
            return True
        except OSError:
            return False

    try:
        import ctypes
        from ctypes import wintypes

        kernel32 = ctypes.WinDLL("kernel32", use_last_error=True)
        open_process = kernel32.OpenProcess
        open_process.argtypes = [wintypes.DWORD, wintypes.BOOL, wintypes.DWORD]
        open_process.restype = wintypes.HANDLE
        get_exit_code = kernel32.GetExitCodeProcess
        get_exit_code.argtypes = [wintypes.HANDLE, ctypes.POINTER(wintypes.DWORD)]
        get_exit_code.restype = wintypes.BOOL
        close_handle = kernel32.CloseHandle
        close_handle.argtypes = [wintypes.HANDLE]
        close_handle.restype = wintypes.BOOL

        handle = open_process(0x1000, False, value)
        if not handle:
            return ctypes.get_last_error() == 5
        try:
            exit_code = wintypes.DWORD()
            if not get_exit_code(handle, ctypes.byref(exit_code)):
                return False
            return exit_code.value == 259
        finally:
            close_handle(handle)
    except (AttributeError, OSError, ValueError):
        return False


def canonical_path(path: Path | str) -> str:
    try:
        return str(Path(path).resolve())
    except OSError:
        return str(Path(path))


def file_signature(path: Path) -> dict[str, Any]:
    try:
        stat = path.stat()
    except OSError:
        return {"exists": False, "mtime_ns": 0, "size": 0}
    return {"exists": True, "mtime_ns": int(stat.st_mtime_ns), "size": int(stat.st_size)}


def _now_local_text() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


@dataclass(frozen=True)
class SourceChangeSet:
    added: list[str]
    updated: list[str]
    deleted: list[str]

    @property
    def changed(self) -> bool:
        return bool(self.added or self.updated or self.deleted)

    def as_counts(self) -> dict[str, int]:
        return {"added": len(self.added), "updated": len(self.updated), "deleted": len(self.deleted)}


class MonitorState:
    def __init__(self, profile: TargetProfile, db_path: Path | None = None, sync_interval_seconds: int = 60):
        self.profile = profile
        self.db_path = Path(db_path) if db_path else None
        self.sync_interval_seconds = int(sync_interval_seconds or 60)
        self.watch_dir = Path(profile.watch_dir)
        self.workspace_dir = Path(profile.workspace_dir)
        self.state_dir = Path(profile.state_dir)
        self.localappdata_dir = Path(profile.localappdata_dir)
        self.summary_csv = self.workspace_dir / "发票汇总.csv"
        self.summary_xlsx = self.workspace_dir / "发票汇总.xlsx"
        self.log_path = self.workspace_dir / BUSINESS_LOG_NAME
        self.stop_file = self.workspace_dir / STOP_FLAG_NAME
        self.lock_file = self.state_dir / LOCK_FILE_NAME
        self.sync_write_lock_file = self.state_dir / SYNC_WRITE_LOCK_FILE_NAME
        self.processed_file = self.state_dir / PROCESSED_FILE_NAME
        self.manual_overrides_file = self.state_dir / MANUAL_OVERRIDES_FILE_NAME
        self.status_file = self.state_dir / STATUS_FILE_NAME
        self.stdout_path = self.state_dir / "bridge_stdout.log"
        self.stderr_path = self.state_dir / "bridge_stderr.log"
        self._lock_acquired = False
        self.ensure_dirs()

    def ensure_dirs(self) -> None:
        for path in (self.watch_dir, self.workspace_dir, self.state_dir, self.localappdata_dir):
            path.mkdir(parents=True, exist_ok=True)

    @contextmanager
    def sync_write_lock(self) -> Iterator[None]:
        """Serialize projection and monitor-state writes for this TargetProfile."""
        key = canonical_path(self.state_dir)
        thread_lock = _sync_thread_lock(key)
        with thread_lock:
            held = getattr(_SYNC_HELD_LOCKS, "counts", {})
            if held.get(key, 0):
                held[key] += 1
                _SYNC_HELD_LOCKS.counts = held
                try:
                    yield
                finally:
                    held[key] -= 1
                return

            self.state_dir.mkdir(parents=True, exist_ok=True)
            with self.sync_write_lock_file.open("a+b") as handle:
                _acquire_sync_os_lock(handle)
                held[key] = 1
                _SYNC_HELD_LOCKS.counts = held
                try:
                    yield
                finally:
                    held.pop(key, None)
                    _release_sync_os_lock(handle)

    def log_event(self, action: str, message: str = "", level: str = "INFO", payload: dict | None = None) -> None:
        self.log_path.parent.mkdir(parents=True, exist_ok=True)
        suffix = ""
        if payload:
            suffix = " " + json.dumps(payload, ensure_ascii=False, sort_keys=True)
        line = f"[{_now_local_text()}] [{level}] {action} {message}{suffix}".rstrip()
        with self.log_path.open("a", encoding="utf-8") as handle:
            handle.write(line + "\n")

    def read_lock(self) -> dict:
        return read_json_object(self.lock_file, {})

    def lock_pid(self) -> int:
        try:
            return int(self.read_lock().get("pid") or 0)
        except Exception:
            return 0

    def cleanup_stale_lock(self) -> bool:
        if not self.lock_file.exists():
            return False
        pid = self.lock_pid()
        if pid and is_pid_alive(pid):
            return False
        try:
            backup = self.lock_file.with_name(f"{self.lock_file.name}.stale.{int(time.time())}.bak")
            shutil.move(str(self.lock_file), str(backup))
            self.log_event("STALE_LOCK_CLEANED", f"pid={pid} backup={backup}", level="WARN")
            return True
        except OSError as exc:
            self.log_event("STALE_LOCK_CLEAN_FAILED", str(exc), level="WARN")
            return False

    def acquire_lock(self) -> bool:
        self.cleanup_stale_lock()
        if self.lock_file.exists():
            pid = self.lock_pid()
            if pid and is_pid_alive(pid):
                self.log_event("LOCK_REFUSED", f"existing_pid={pid}", level="ERROR")
                return False
        payload = {
            "pid": os.getpid(),
            "watch_dir": str(self.watch_dir),
            "workspace_dir": str(self.workspace_dir),
            "state_dir": str(self.state_dir),
            "target_id": self.profile.id,
            "started_at": utc_now_text(),
            "sync_interval_seconds": self.sync_interval_seconds,
        }
        self.lock_file.parent.mkdir(parents=True, exist_ok=True)
        try:
            fd = os.open(str(self.lock_file), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            with os.fdopen(fd, "w", encoding="utf-8") as handle:
                json.dump(payload, handle, ensure_ascii=False, indent=2)
            self._lock_acquired = True
            return True
        except FileExistsError:
            self.log_event("LOCK_REFUSED", "lock file appeared during startup", level="ERROR")
            return False
        except OSError as exc:
            self.log_event("LOCK_WRITE_FAILED", str(exc), level="ERROR")
            return False

    def release_lock(self) -> None:
        if not self._lock_acquired:
            return
        try:
            if self.lock_file.exists():
                self.lock_file.unlink()
        except OSError as exc:
            self.log_event("LOCK_RELEASE_FAILED", str(exc), level="WARN")
        finally:
            self._lock_acquired = False

    def clear_stop_flag(self) -> None:
        try:
            if self.stop_file.exists():
                self.stop_file.unlink()
        except OSError:
            pass

    def request_stop(self) -> None:
        self.stop_file.parent.mkdir(parents=True, exist_ok=True)
        self.stop_file.write_text("stop\n", encoding="ascii")

    def stop_requested(self) -> bool:
        return self.stop_file.exists()

    def update_status(self, **updates: Any) -> dict:
        # status_file is a read/modify/write projection shared by the daemon,
        # startup child, and manual rebuild. Its whole transaction is locked.
        with self.sync_write_lock():
            payload = read_json_object(self.status_file, {})
            payload.update(updates)
            payload.update(
                {
                    "target_id": self.profile.id,
                    "watch_dir": str(self.watch_dir),
                    "workspace_dir": str(self.workspace_dir),
                    "state_dir": str(self.state_dir),
                    "log_path": str(self.log_path),
                    "sync_interval_seconds": self.sync_interval_seconds,
                    "updated_at": utc_now_text(),
                }
            )
            atomic_write_json(self.status_file, payload)
            return payload

    def read_status(self) -> dict:
        return read_json_object(self.status_file, {})

    def load_processed(self) -> dict[str, dict]:
        if not self.processed_file.exists():
            return {}
        try:
            payload = json.loads(self.processed_file.read_text(encoding="utf-8-sig"))
        except Exception as exc:
            backup = self.processed_file.with_name(f"{self.processed_file.stem}.broken.{int(time.time())}{self.processed_file.suffix}")
            try:
                shutil.copy2(self.processed_file, backup)
                self.log_event("PROCESSED_REBUILT", f"broken state backed up to {backup}: {exc}", level="WARN")
            except OSError:
                self.log_event("PROCESSED_REBUILT", f"broken state could not be backed up: {exc}", level="WARN")
            rebuilt = self.rebuild_processed_from_summary()
            self.save_processed(rebuilt)
            return rebuilt
        if not isinstance(payload, dict):
            return {}
        raw_files = payload.get("files") if isinstance(payload.get("files"), dict) else payload
        files: dict[str, dict] = {}
        for raw_path, item in raw_files.items():
            if isinstance(item, dict):
                files[canonical_path(raw_path)] = dict(item)
        return files

    def save_processed(self, files: dict[str, dict]) -> None:
        atomic_write_json(self.processed_file, {"version": 1, "files": files, "updated_at": utc_now_text()})

    def source_snapshot(self) -> dict[str, dict]:
        result: dict[str, dict] = {}
        for path in supported_invoice_files(self.watch_dir):
            key = canonical_path(path)
            sig = file_signature(path)
            result[key] = {
                "path": key,
                "name": path.name,
                "suffix": path.suffix.lower(),
                "mtime_ns": sig["mtime_ns"],
                "size": sig["size"],
            }
        return result

    def detect_source_changes(self) -> SourceChangeSet:
        previous = self.load_processed()
        current = self.source_snapshot()
        added: list[str] = []
        updated: list[str] = []
        deleted: list[str] = []
        for key, sig in current.items():
            old = previous.get(key)
            if not old:
                added.append(key)
            elif int(old.get("mtime_ns") or old.get("mtime") or 0) != int(sig.get("mtime_ns") or 0) or int(old.get("size") or 0) != int(sig.get("size") or 0):
                updated.append(key)
        for key in previous:
            if key not in current:
                deleted.append(key)
        return SourceChangeSet(added=sorted(added), updated=sorted(updated), deleted=sorted(deleted))

    def rebuild_processed_from_summary(self) -> dict[str, dict]:
        rows = read_csv_rows(self.summary_csv)
        current = self.source_snapshot()
        rebuilt: dict[str, dict] = {}
        for row in rows:
            key = canonical_path(row.get("文件路径") or "")
            if not key:
                continue
            sig = current.get(key) or file_signature(Path(key))
            rebuilt[key] = {
                "invoice_no": str(row.get("发票号码") or "").strip(),
                "seller": str(row.get("销售方") or "").strip(),
                "amount": str(row.get("开票金额") or "").strip(),
                "processed_at": utc_now_text(),
                "mtime_ns": int(sig.get("mtime_ns") or 0),
                "size": int(sig.get("size") or 0),
            }
        return rebuilt

    def load_manual_overrides(self) -> dict[str, dict]:
        payload = read_json_object(self.manual_overrides_file, {})
        raw_items = payload.get("items") if isinstance(payload.get("items"), dict) else payload
        items: dict[str, dict] = {}
        if not isinstance(raw_items, dict):
            return items
        for key, raw in raw_items.items():
            if not isinstance(raw, dict):
                continue
            source_path = raw.get("source_path") or key
            normalized_key = canonical_path(source_path)
            fields = raw.get("fields") if isinstance(raw.get("fields"), dict) else {}
            cleaned = {field: str(fields.get(field) or "").strip() for field in EDITABLE_FIELDS if field in fields}
            if cleaned:
                items[normalized_key] = {"source_path": normalized_key, "fields": cleaned, "updated_at": str(raw.get("updated_at") or "")}
        return items

    def save_manual_overrides(self, items: dict[str, dict]) -> None:
        atomic_write_json(self.manual_overrides_file, {"version": 1, "items": items, "updated_at": utc_now_text()})

    def apply_manual_overrides_to_summary(self) -> int:
        rows = read_csv_rows(self.summary_csv)
        overrides = self.load_manual_overrides()
        applied = 0
        for row in rows:
            key = canonical_path(row.get("文件路径") or "")
            item = overrides.get(key)
            fields = item.get("fields") if isinstance(item, dict) else None
            if not isinstance(fields, dict):
                continue
            changed = False
            for field, value in fields.items():
                if field in EDITABLE_FIELDS and str(row.get(field) or "") != str(value or ""):
                    row[field] = str(value or "")
                    changed = True
            if changed:
                row["手改状态"] = "已手改"
                applied += 1
        if applied:
            write_csv_rows(self.summary_csv, SUMMARY_HEADERS, rows)
            write_summary_xlsx(self.summary_xlsx, rows)
            self.log_event("MANUAL_EDIT_APPLIED", f"applied={applied}")
        return applied

    def _read_summary_xlsx_rows(self) -> list[dict[str, str]]:
        if not self.summary_xlsx.exists():
            return []
        wb = load_workbook(self.summary_xlsx, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(cell or "").strip() for cell in rows[0]]
        result: list[dict[str, str]] = []
        for values in rows[1:]:
            result.append({headers[i]: str(values[i] or "").strip() if i < len(values) else "" for i in range(len(headers)) if headers[i]})
        return result

    def sync_excel_manual_edits(self) -> int:
        if not self.summary_csv.exists() or not self.summary_xlsx.exists():
            return 0
        try:
            xlsx_rows = self._read_summary_xlsx_rows()
        except Exception as exc:
            self.log_event("MANUAL_SYNC_GUARD_BLOCK", f"summary workbook locked or unreadable: {exc}", level="WARN")
            return 0
        csv_rows = read_csv_rows(self.summary_csv)
        csv_by_key = {canonical_path(row.get("文件路径") or ""): row for row in csv_rows if row.get("文件路径")}
        overrides = self.load_manual_overrides()
        changed = 0
        for xrow in xlsx_rows:
            key = canonical_path(xrow.get("文件路径") or "")
            if not key or key not in csv_by_key:
                continue
            target = csv_by_key[key]
            fields: dict[str, str] = {}
            for field in EDITABLE_FIELDS:
                incoming = str(xrow.get(field) or "").strip()
                current = str(target.get(field) or "").strip()
                if incoming != current:
                    fields[field] = incoming
            if not fields:
                continue
            self.log_event("MANUAL_EDIT_DETECTED", f"path={key} fields={','.join(sorted(fields))}")
            for field, value in fields.items():
                target[field] = value
            target["手改状态"] = "已手改"
            overrides[key] = {"source_path": key, "fields": {**overrides.get(key, {}).get("fields", {}), **fields}, "updated_at": utc_now_text()}
            changed += 1
        if not changed:
            return 0
        self.log_event("MANUAL_SYNC_GUARD_PASS", f"changed_rows={changed}")
        write_csv_rows(self.summary_csv, SUMMARY_HEADERS, csv_rows)
        write_summary_xlsx(self.summary_xlsx, csv_rows)
        self.save_manual_overrides(overrides)
        self.log_event("MANUAL_EDIT_AUTO_SYNC", f"changed_rows={changed}")
        return changed

    def notify_invoice_change(self, trigger: str, counts: dict[str, int]) -> None:
        total = sum(int(counts.get(key, 0)) for key in ("added", "updated", "deleted", "blocked"))
        if total <= 0:
            self.log_event("NOTIFY_SKIP", f"zero change trigger={trigger}")
            return
        if os.environ.get("INVOICE_HUB_DISABLE_NOTIFY") == "1":
            self.log_event("NOTIFY_SKIP", f"disabled by env trigger={trigger}")
            return
        title = "发票文件变化"
        message = f"新增 {counts.get('added', 0)}，更新 {counts.get('updated', 0)}，删除 {counts.get('deleted', 0)}"
        self.log_event("NOTIFY_ATTEMPT", f"title={title} message={message}")
        if os.name != "nt":
            self.log_event("NOTIFY_FALLBACK", "non-Windows environment, logged only", level="WARN")
            return
        command = [
            "powershell.exe",
            "-NoLogo",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            (
                "Add-Type -AssemblyName System.Windows.Forms;"
                f"$n=New-Object System.Windows.Forms.NotifyIcon;$n.Icon=[System.Drawing.SystemIcons]::Information;"
                f"$n.Visible=$true;$n.ShowBalloonTip(5000,'{title}','{message}',[System.Windows.Forms.ToolTipIcon]::Info);"
                "Start-Sleep -Milliseconds 800;$n.Dispose()"
            ),
        ]
        try:
            result = subprocess.run(command, check=False, capture_output=True, text=True, timeout=8, creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0))
        except Exception as exc:
            self.log_event("NOTIFY_FAIL", str(exc), level="WARN")
            return
        if result.returncode == 0:
            self.log_event("NOTIFY_SENT", "Windows balloon notification dispatched")
        else:
            self.log_event("NOTIFY_FAIL", (result.stderr or result.stdout or "unknown failure").strip(), level="WARN")

    def notify_self_test(self) -> dict:
        self.notify_invoice_change("self_test", {"added": 1, "updated": 0, "deleted": 0, "blocked": 0})
        return {"ok": True, "log_path": str(self.log_path)}

    def supported_path(self, path: Path | str) -> bool:
        value = Path(path)
        return value.suffix.lower() in SUPPORTED_EXTS and not value.name.startswith("~$")
